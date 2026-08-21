# -*- coding: utf-8 -*-
"""软包墙围分阶段创建：计划、预览与结果追踪。

Commands
--------
plan      从登记表 + 赛狐/EN 快照生成计划并写/更新阶段记录 xlsx
preview   对计划行调用 EN get_bundle_serial_preview（只读），写入预计 TJ#
status    打印阶段记录汇总
record    更新单条通途SKU的阶段状态和结果（用于 apply 后登记）

示例
----
uv run --project .. python soft_wall_stage.py plan
uv run --project .. python soft_wall_stage.py preview
uv run --project .. python soft_wall_stage.py status
uv run --project .. python soft_wall_stage.py record --sku TT0031084K0063339-12pcs --stage-status EN已建
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from combo_en import EnRestClient, make_en_session
from combo_en import bundle_from_docs
from combo_reconcile import validate_en_bundle
from sellfox_combo_ops import cmd_create, make_client

HERE = Path(__file__).resolve().parent
DATA = HERE / "数据源"
REGISTER = DATA / "未配对产品登记表0821.xlsx"
KEYWORD = "软包墙围"
TRACKER_NAME = "软包墙围阶段记录.xlsx"
SNAPSHOT_PREFIX = "软包墙围快照_"
TRACKER = DATA / TRACKER_NAME


def configure(product: str) -> None:
    global KEYWORD, TRACKER_NAME, SNAPSHOT_PREFIX, TRACKER
    KEYWORD = product
    TRACKER_NAME = f"{product}阶段记录.xlsx"
    SNAPSHOT_PREFIX = f"{product}快照_"
    TRACKER = DATA / TRACKER_NAME

HEADERS = [
    "序号",
    "阶段",
    "通途SKU",
    "数量",
    "底层EN物料",
    "赛狐底层SKU",
    "赛狐底层ID",
    "底层赛狐存在",
    "物料名称",
    "ASIN",
    "店铺",
    "预计TJ#",
    "EN套件名称",
    "阶段状态",
    "EN结果",
    "客户物料号结果",
    "赛狐结果",
    "完成时间",
    "备注",
]

STATUS_COLUMNS = ["阶段状态", "EN结果", "客户物料号结果", "赛狐结果", "完成时间", "备注"]


def latest_snapshot() -> Path:
    files = sorted(DATA.glob(f"{SNAPSHOT_PREFIX}*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise SystemExit(f"未找到快照: {DATA}/{SNAPSHOT_PREFIX}*.json，先跑 soft_wall_lookup.py --product {KEYWORD}")
    return files[0]


def normalize(text: str) -> str:
    return str(text or "").strip()


def read_register_rows() -> list[tuple[int, str, str, str, object]]:
    workbook = load_workbook(REGISTER, read_only=True, data_only=True)
    rows = list(workbook["Sheet1"].iter_rows(values_only=True))
    result: list[tuple[int, str, str, str, object]] = []
    for index, row in enumerate(rows[1:], start=2):
        asin = normalize(row[0])
        shop = normalize(row[1])
        sku = normalize(row[2])
        name = normalize(row[3])
        qty = row[4]
        if not name or KEYWORD not in name:
            continue
        result.append((index, asin, shop, sku, name, qty))
    return result


def build_plan_rows(*, full: bool = False) -> list[dict]:
    snapshot = json.loads(latest_snapshot().read_text(encoding="utf-8"))
    ref_to_items: dict[str, list[str]] = {}
    base_names: dict[str, str] = {}
    base_infos: dict[str, list[str]] = {}
    for item in snapshot.get("en_items") or []:
        if str(item.get("item_group") or "") != "软包墙围":
            continue
        code = str(item.get("item_code") or "")
        base_names[code] = str(item.get("item_name") or "")
        for customer in item.get("customer_items") or []:
            ref_raw = str(customer.get("ref_code") or "").strip()
            ref = ref_raw.casefold()
            if ref:
                ref_to_items.setdefault(ref, []).append(code)
                base_infos.setdefault(code, []).append(ref_raw)
    sellfox = snapshot.get("sellfox_by_sku") or {}
    rows: dict[str, dict] = {}
    for index, asin, shop, sku, name, qty in read_register_rows():
        if not sku or sku == "无捆绑SKU":
            continue
        match = re.search(r"-(\d+)pcs$", sku, re.IGNORECASE)
        if match:
            parsed_qty = int(match.group(1))
        else:
            try:
                parsed_qty = int(qty)
            except (TypeError, ValueError):
                parsed_qty = None
        base_sku = re.sub(r"-(\d+)pcs$", "", sku, flags=re.IGNORECASE)
        base_items = list(dict.fromkeys(ref_to_items.get(base_sku.casefold(), [])))
        row = rows.setdefault(
            sku,
            {
                "阶段": "全部",
                "通途SKU": sku,
                "数量": parsed_qty,
                "底层EN物料": " | ".join(base_items),
                "赛狐底层SKU": base_items[0] if base_items else "",
                "赛狐底层ID": "",
                "底层赛狐存在": "是" if base_items else "否",
                "物料名称": name,
                "ASIN": [],
                "店铺": [],
                "预计TJ#": "",
                "EN套件名称": "",
                "阶段状态": "待创建",
                "EN结果": "",
                "客户物料号结果": "",
                "赛狐结果": "",
                "完成时间": "",
                "备注": "",
            },
        )
        row["ASIN"].append(asin)
        row["店铺"].append(shop)
    for row in rows.values():
        row["ASIN"] = " | ".join(sorted(set(row["ASIN"])))
        row["店铺"] = " | ".join(sorted(set(row["店铺"])))
        base_code = row["赛狐底层SKU"]
        if base_code:
            sx = sellfox.get(base_code) or {}
            row["赛狐底层ID"] = str(sx.get("id") or "")
            row["底层赛狐存在"] = "是" if sx else "否"
    if full:
        for code, refs in base_infos.items():
            base_ref = refs[0] if refs else ""
            if not base_ref:
                continue
            base_name = base_names.get(code, code)
            for qty in (4, 6, 9, 12):
                full_sku = f"{base_ref}-{qty}pcs"
                if full_sku.casefold() in {str(sku).casefold() for sku in rows}:
                    continue
                sx = sellfox.get(code) or {}
                rows[full_sku] = {
                    "阶段": "全部",
                    "通途SKU": full_sku,
                    "数量": qty,
                    "底层EN物料": code,
                    "赛狐底层SKU": code,
                    "赛狐底层ID": str(sx.get("id") or ""),
                    "底层赛狐存在": "是" if sx else "否",
                    "物料名称": base_name,
                    "ASIN": "无登记表",
                    "店铺": "无登记表",
                    "预计TJ#": "",
                    "EN套件名称": "",
                    "阶段状态": "待创建",
                    "EN结果": "",
                    "客户物料号结果": "",
                    "赛狐结果": "",
                    "完成时间": "",
                    "备注": "全量补齐：登记表无此通途SKU",
                }
    return list(rows.values())


def existing_tracker_rows() -> dict[str, dict]:
    if not TRACKER.exists():
        return {}
    workbook = load_workbook(TRACKER, data_only=True)
    if "软包墙围计划" not in workbook.sheetnames:
        return {}
    sheet = workbook["软包墙围计划"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h or "") for h in rows[0]]
    result: dict[str, dict] = {}
    for raw in rows[1:]:
        row = {header[i]: raw[i] if i < len(raw) else None for i in range(len(header))}
        sku = normalize(row.get("通途SKU"))
        if sku:
            result[sku] = row
    return result


def tracker_rows_sorted() -> list[dict]:
    if TRACKER.exists():
        rows = list(existing_tracker_rows().values())
        rows.sort(key=lambda row: int(row.get("序号") or 0))
        return rows
    return build_plan_rows()


def write_tracker(rows: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "软包墙围计划"
    sheet.append(HEADERS)
    for index, row in enumerate(rows, start=1):
        values = [row.get(header, "") for header in HEADERS]
        values[0] = index
        sheet.append(values)
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(TRACKER)
    print(f"阶段记录已写: {TRACKER}")


def cmd_plan(_args: argparse.Namespace) -> int:
    rows = build_plan_rows(full=_args.full)
    old = existing_tracker_rows()
    for row in rows:
        previous = old.get(row["通途SKU"])
        if not previous:
            continue
        for column in STATUS_COLUMNS + ["预计TJ#", "EN套件名称", "阶段"]:
            if previous.get(column) not in (None, ""):
                row[column] = previous.get(column)
    write_tracker(rows)
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["阶段状态"]] = counts.get(row["阶段状态"], 0) + 1
    print(json.dumps({"rows": len(rows), "counts": counts}, ensure_ascii=False, indent=2))
    return 0


def cmd_preview(args: argparse.Namespace) -> int:
    rows = tracker_rows_sorted()
    root = HERE.parent
    base, session = make_en_session(root, args.env)
    en = EnRestClient(base, session)
    for row in rows:
        if row["预计TJ#"]:
            continue
        base_sku = row["赛狐底层SKU"]
        qty = row["数量"]
        if not base_sku or not qty:
            row["备注"] = (row["备注"] + " | " if row["备注"] else "") + "缺底层SKU或数量，无法预览"
            continue
        preview = en.preview([(base_sku, int(qty))])
        print(json.dumps({"通途SKU": row["通途SKU"], "preview": preview}, ensure_ascii=False))
        row["预计TJ#"] = str(preview.get("new_item_code") or "")
        row["EN套件名称"] = str(preview.get("new_item_name") or "")
        if preview.get("is_duplicate"):
            row["阶段状态"] = "EN已存在"
            row["备注"] = (row["备注"] + " | " if row["备注"] else "") + f"duplicate={preview.get('existing_bundle')}"
        elif row["阶段状态"] == "待创建":
            row["阶段状态"] = "已预览"
    write_tracker(rows)
    return 0


def cmd_status(_args: argparse.Namespace) -> int:
    rows = tracker_rows_sorted()
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["阶段状态"]] = counts.get(row["阶段状态"], 0) + 1
    print(json.dumps({"rows": len(rows), "counts": counts}, ensure_ascii=False, indent=2))
    for row in rows:
        print(
            f"{row['通途SKU']} | {row['数量']} | {row['赛狐底层SKU']} | "
            f"{row['预计TJ#']} | {row['阶段状态']} | {row['备注']}"
        )
    return 0


def cmd_record(args: argparse.Namespace) -> int:
    rows = tracker_rows_sorted()
    found = False
    for row in rows:
        if row["通途SKU"] != args.sku:
            continue
        found = True
        if args.stage_status:
            row["阶段状态"] = args.stage_status
        if args.phase:
            row["阶段"] = args.phase
        if args.tj_code:
            row["预计TJ#"] = args.tj_code
        if args.en_name:
            row["EN套件名称"] = args.en_name
        if args.en_result is not None:
            row["EN结果"] = args.en_result
        if args.customer_result is not None:
            row["客户物料号结果"] = args.customer_result
        if args.sellfox_result is not None:
            row["赛狐结果"] = args.sellfox_result
        if args.note is not None:
            row["备注"] = args.note if args.note_replace else (row["备注"] + " | " if row["备注"] else "") + args.note
        if args.complete:
            row["完成时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row["阶段状态"] = "完成"
    if not found:
        raise SystemExit(f"计划中不存在该通途SKU: {args.sku}")
    write_tracker(rows)
    return 0


def canonical_ref_code(sku: str) -> str:
    match = re.search(r"-(\d+)(?i:pcs)$", sku)
    if match:
        return sku[: match.start()] + f"-{match.group(1)}pcs"
    return sku


def _apply_row(
    row: dict, args: argparse.Namespace, en: EnRestClient, client: object
) -> None:
    base_sku = str(row.get("赛狐底层SKU") or "").strip()
    full_sku = str(row.get("通途SKU") or "").strip()
    if not base_sku or not full_sku:
        raise RuntimeError(f"缺少底层SKU或通途SKU: {row.get('通途SKU')}")
    try:
        qty = int(row.get("数量"))
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f"数量无法解析: {row.get('数量')!r}") from exc
    preview = en.preview([(base_sku, qty)])
    if preview.get("is_duplicate"):
        row["阶段状态"] = "EN已存在"
        row["备注"] = (row.get("备注") or "") + f" | duplicate={preview.get('existing_bundle')}"
        return
    row["预计TJ#"] = str(preview.get("new_item_code") or "")
    row["EN套件名称"] = str(preview.get("new_item_name") or "")
    if not args.apply:
        row["阶段状态"] = "已预览"
        return
    created = en.create_bundle([(base_sku, qty)])
    bundle_name = str(created.get("name") or created.get("new_item_code") or "")
    bundle_doc = en.get_product_bundle(bundle_name)
    item = en.get_item(str(bundle_doc.get("new_item_code") or bundle_name))
    bundle = bundle_from_docs(bundle_doc, item)
    problems = validate_en_bundle(bundle)
    if problems:
        raise RuntimeError(f"EN 回读断言失败: {', '.join(problems)}")
    tj_code = bundle.new_item_code
    en_name = bundle.new_item_code_name
    row["预计TJ#"] = tj_code
    row["EN套件名称"] = en_name
    row["EN结果"] = "创建成功并回读"
    item_detail = en.get_item_customer_items(tj_code)
    existing = list(item_detail.get("customer_items") or [])
    ref_code = canonical_ref_code(full_sku)
    existing_refs = {str(c.get("ref_code") or "").casefold() for c in existing}
    if ref_code.casefold() not in existing_refs:
        en.put_customer_items(
            tj_code, [*existing, {"customer_group": "美国公司", "ref_code": ref_code}]
        )
        verified = en.get_item_customer_items(tj_code)
        verified_refs = {str(c.get("ref_code") or "").casefold() for c in (verified.get("customer_items") or [])}
        if ref_code.casefold() not in verified_refs:
            raise RuntimeError(f"客户物料号回读失败: {tj_code} -> {ref_code}")
    row["客户物料号结果"] = f"已登记并回读 ({ref_code})"
    namespace = argparse.Namespace(
        sku=tj_code,
        name=en_name,
        child=[f"{base_sku}:{qty}"],
        full_cid="428697-",
        auto_calc_weight="true",
        apply=True,
        ops_cache=None,
        bottom_cache=None,
    )
    if cmd_create(client, namespace) != 0:
        raise RuntimeError(f"赛狐创建失败: {tj_code}")
    row["赛狐结果"] = "已创建并回读"
    row["阶段状态"] = "完成"
    row["完成时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def cmd_apply(args: argparse.Namespace) -> int:
    rows = tracker_rows_sorted()
    selected = {s.strip() for s in (args.only or "").split(",") if s.strip()}
    root = HERE.parent
    base, session = make_en_session(root, args.env)
    en = EnRestClient(base, session)
    client = make_client() if args.apply else None
    failed = 0
    for row in rows:
        sku = str(row.get("通途SKU") or "")
        if selected and sku not in selected:
            continue
        if row.get("阶段状态") in {"完成", "EN已存在"}:
            continue
        try:
            _apply_row(row, args, en, client)
        except (SystemExit, RuntimeError, ValueError) as exc:
            failed += 1
            row["阶段状态"] = "失败"
            row["备注"] = (row.get("备注") or "") + f" | 失败: {exc}"
        write_tracker(rows)
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.get("阶段状态") or ""] = counts.get(row.get("阶段状态") or "", 0) + 1
    print(json.dumps({
        "mode": "apply" if args.apply else "dry-run",
        "selected": sorted(selected) if selected else "all_pending",
        "counts": counts,
        "failed": failed,
    }, ensure_ascii=False, indent=2))
    return 1 if failed else 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)
    plan_cmd = sub.add_parser("plan", help="生成/更新阶段记录 xlsx")
    plan_cmd.add_argument("--full", action="store_true", help="补齐 6 底层物料 x 4 数量 = 24 行")
    plan_cmd.set_defaults(func=cmd_plan)
    preview = sub.add_parser("preview", help="调用 EN 预览套件编号（只读）")
    preview.add_argument("--env", default="prod", choices=["prod", "test"], help="EN 环境，默认 prod")
    preview.set_defaults(func=cmd_preview)
    sub.add_parser("status", help="打印阶段汇总").set_defaults(func=cmd_status)
    apply_cmd = sub.add_parser("apply", help="批量创建 EN → 客户物料号 → 赛狐组合（默认 dry-run）")
    apply_cmd.add_argument("--only", help="逗号分隔的通途SKU，默认全部待创建")
    apply_cmd.add_argument("--env", default="prod", choices=["prod", "test"], help="EN 环境，默认 prod")
    apply_cmd.add_argument("--apply", action="store_true", help="实际写入 EN 与赛狐")
    apply_cmd.set_defaults(func=cmd_apply)
    record = sub.add_parser("record", help="更新单行阶段状态和结果")
    record.add_argument("--sku", required=True, help="完整通途SKU")
    record.add_argument("--stage-status", help="阶段状态")
    record.add_argument("--en-result", help="EN 创建结果")
    record.add_argument("--customer-result", help="客户物料号登记结果")
    record.add_argument("--sellfox-result", help="赛狐组合创建结果")
    record.add_argument("--note", help="追加备注")
    record.add_argument("--note-replace", action="store_true", help="用 --note 覆盖备注而不是追加")
    record.add_argument("--phase", help="阶段：测试/全部")
    record.add_argument("--tj-code", help="实际 TJ# SKU")
    record.add_argument("--en-name", help="EN 套件名称")
    record.add_argument("--complete", action="store_true", help="标记完成并写完成时间")
    record.set_defaults(func=cmd_record)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
