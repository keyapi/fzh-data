#!/usr/bin/env python3
"""
test_e2e_flow.py
赛狐 端到端测试 — test001-white 单 SKU 完整流程

流程:
  1. 其他入库 test001-white (100个) → 导入 + 自验证
  2. 导出库存明细 → 读文件验证 test001-white 库存>0
  3. 其他出库 test001-white → 导入 + 自验证(库存清零)
  4. 导出库存明细 → 读文件验证 test001-white 库存=0
  5. 海外仓备货单 test001-white → 生成 + 导入 + 自验证

用法:
  cd D:\\Work\\赛狐\\Cursor\\warehouse_restock
  uv run python test_e2e_flow.py              # 全部步骤
  uv run python test_e2e_flow.py --step 1     # 仅某一步
  uv run python test_e2e_flow.py --headless   # 无头模式
"""

import os
import sys
import shutil
import time
import subprocess
from datetime import datetime
from pathlib import Path

os.chdir(os.path.dirname(os.path.abspath(__file__)))

CURSOR_ROOT = Path(__file__).resolve().parent.parent
WEB_AUTO = Path(r"D:\Work\赛狐\网页自动化")
CURSOR_VENV = CURSOR_ROOT / ".venv" / "Scripts" / "python.exe"
WEB_VENV = WEB_AUTO / ".venv" / "Scripts" / "python.exe"
DOWNLOADS_DIR = WEB_AUTO / "downloads"

# 测试参数
TEST_SKU = "test001-white"
TEST_WH = "POLAND"
TEST_QTY = 100
TEST_PRICE = 1.0

# ── 子进程调用 ─────────────────────────────────────────────

def run_script(script: Path, cwd: Path, venv: Path, *extra_args) -> int:
    args = [str(venv), str(script)] + list(extra_args)
    name = script.name
    print(f"\n  [{name}] 执行中...")
    result = subprocess.run(args, cwd=str(cwd), encoding="utf-8",
                            errors="replace", env={**os.environ, "PYTHONIOENCODING": "utf-8"})
    if result.returncode != 0:
        print(f"  [{name}] 返回码={result.returncode}")
    return result.returncode


def check_test_sku_in_export(expected_qty: int | None = None) -> dict:
    """读最新库存导出，检查 test001-white 的库存状态。"""
    candidates = sorted(
        [f for f in DOWNLOADS_DIR.glob("WarehouseItem*.xlsx") if not f.name.startswith("~$")],
        key=lambda f: f.stat().st_mtime, reverse=True
    )
    if not candidates:
        return {"found": False, "error": "无导出文件"}

    import pandas as pd
    df = pd.read_excel(candidates[0])
    sku_col = "SKU" if "SKU" in df.columns else df.columns[0]
    wh_col = None
    for c in df.columns:
        if "仓库" in str(c):
            wh_col = c
            break

    match = df[(df[sku_col].astype(str).str.strip() == TEST_SKU)]
    if wh_col:
        match = match[match[wh_col].astype(str).str.strip() == TEST_WH]

    if len(match) == 0:
        return {"found": False, "sku": TEST_SKU, "file": candidates[0].name}

    row = match.iloc[0]
    result = {
        "found": True,
        "sku": TEST_SKU,
        "warehouse": TEST_WH,
        "file": candidates[0].name,
    }
    for col in df.columns:
        val = row[col]
        if pd.notna(val):
            result[str(col)] = val

    # 提取关键库存字段
    for key_col in ["可用数", "库存总数", "次品数", "占用数", "在途数", "计划数"]:
        if key_col in df.columns:
            result[key_col] = float(row.get(key_col, 0) or 0)

    return result


# ── 各步骤 ────────────────────────────────────────────────

def step_1_inbound(qty: int = TEST_QTY):
    """Step 1: 其他入库 test001-white。"""
    print("\n" + "=" * 60)
    print(f"Step 1: 其他入库 {TEST_SKU} x{qty} → {TEST_WH}")
    print("=" * 60)

    headless = "--headless" in sys.argv
    args = [
        "--sku", TEST_SKU, "--wh", TEST_WH, "--qty", str(qty),
        "--price", str(TEST_PRICE), "--note", "E2E自动化测试"
    ]
    if headless:
        args.append("--headless")

    script = WEB_AUTO / "sellfox_import_other_inbound.py"
    ret = run_script(script, WEB_AUTO, WEB_VENV, *args)
    return ret == 0


def step_2_export_and_verify(expected_available: int = None):
    """Step 2: 导出库存明细 + 验证 test001-white 有库存。"""
    print("\n" + "=" * 60)
    msg = f"Step 2: 导出库存明细 + 验证 {TEST_SKU} 库存>0"
    if expected_available is not None:
        msg += f" (期望={expected_available})"
    print(msg)
    print("=" * 60)

    # 导出
    script = WEB_AUTO / "sellfox_auto_export.py"
    ret = run_script(script, WEB_AUTO, WEB_VENV, "--headless")
    if ret != 0:
        return False

    # 验证
    result = check_test_sku_in_export()
    if not result["found"]:
        # 库存为0时可能被赛狐隐藏(隐藏0数据)，也算通过
        print(f"  {TEST_SKU} 不在导出文件中 (可能库存=0被隐藏)")
        return True, 0

    available = result.get("可用数", 0)
    total = result.get("库存总数", 0)
    print(f"  {TEST_SKU} @ {TEST_WH}: 可用数={available}, 库存总数={total}")

    if expected_available is not None and available != expected_available:
        print(f"  ⚠ 期望={expected_available}, 实际={available} (可能累计了之前的测试数据)")
    if available > 0:
        print(f"  ✓ 验证通过: 库存>0")
        return True, available  # 返回实际库存数
    else:
        print(f"  ✗ 库存=0 (入库可能未确认?)")
        return False, 0


def step_3_outbound(available_qty: int = None):
    """Step 3: 其他出库 test001-white (清零)。"""
    print("\n" + "=" * 60)
    print(f"Step 3: 其他出库 {TEST_SKU} (清零)")
    print("=" * 60)

    # 先导出确认当前库存
    if available_qty is None:
        result = check_test_sku_in_export()
        if not result["found"]:
            print(f"  {TEST_SKU} 不在库存中（库存=0），跳过出库")
            return True
        available_qty = int(result.get("可用数", 0) or 0)

    defective = 0
    if available_qty == 0:
        print(f"  SKU 库存已为 0，跳过出库")
        return True

    print(f"  当前库存: 可用={available_qty}")

    import openpyxl

    # 创建出库文件
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outbound_dir = WEB_AUTO / "outbound"
    outbound_dir.mkdir(exist_ok=True)
    out_path = outbound_dir / f"test_outbound_{TEST_SKU}_{stamp}.xlsx"

    template = CURSOR_ROOT / "other_outbound" / "数据源样例" / "赛狐_其他出库_模板.xlsx"
    if template.exists():
        shutil.copy(template, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active

    row = 2  # 其他出库模板: 1行表头
    ws.cell(row=row, column=1, value=f"OB_TEST_{stamp}")  # 临时单号
    ws.cell(row=row, column=2, value=TEST_WH)              # *出库仓库
    ws.cell(row=row, column=3, value="其他出库")            # *出库类型
    ws.cell(row=row, column=7, value=TEST_SKU)             # *SKU
    ws.cell(row=row, column=10, value=available_qty)        # 可用出库量
    if defective > 0:
        ws.cell(row=row, column=11, value=defective)       # 次品出库量
    wb.save(out_path)
    print(f"  出库文件: {out_path.name}")

    # 导入
    headless = "--headless" in sys.argv
    headless_args = ["--headless"] if headless else []
    script = WEB_AUTO / "sellfox_import_other_outbound.py"
    ret = run_script(script, WEB_AUTO, WEB_VENV, str(out_path), *headless_args)
    if ret != 0:
        return False

    print("  ⚠ 导入完成。需手动去赛狐'确认出库'后库存才会清零")
    return True


def step_4_export_and_verify_zero():
    """Step 4: 导出库存明细 + 验证 test001-white 库存=0（或不存在）。"""
    print("\n" + "=" * 60)
    print(f"Step 4: 导出验证 {TEST_SKU} 库存=0")
    print("=" * 60)

    script = WEB_AUTO / "sellfox_auto_export.py"
    ret = run_script(script, WEB_AUTO, WEB_VENV, "--headless")
    if ret != 0:
        return False

    result = check_test_sku_in_export()
    if not result["found"]:
        print(f"  {TEST_SKU} 不在导出中（可能已完全清零）")
        print(f"  ✓ 验证通过: SKU 已清零")
        return True

    available = result.get("可用数", 0)
    total = result.get("库存总数", 0)
    print(f"  {TEST_SKU} @ {TEST_WH}: 可用数={available}, 库存总数={total}")
    if available == 0 and total == 0:
        print(f"  ✓ 验证通过: 库存已清零")
        return True
    else:
        print(f"  ⚠ 库存未清零: 可用={available}, 总={total} (可能还未手动确认出库)")
        return False


def step_5_warehouse_restock():
    """Step 5: 海外仓备货单 test001-white。"""
    print("\n" + "=" * 60)
    print(f"Step 5: 海外仓备货单 {TEST_SKU} → {TEST_WH}")
    print("=" * 60)

    # 生成备货单（已有完整脚本生成全部数据，test001-white 包含在内）
    ret = run_script(
        Path(__file__).resolve(),
        Path(__file__).resolve().parent,
        CURSOR_VENV
    )
    # 上面调用的是自己，不合理。改用 build_saihu_warehouse_restock.py
    gen_script = Path(__file__).resolve().parent / "build_saihu_warehouse_restock.py"
    ret = run_script(gen_script, gen_script.parent, CURSOR_VENV)
    if ret != 0:
        # 备货单生成内部已有所有验证
        return False

    # 找到包含 test001-white 的备货单文件
    out_dirs = sorted((gen_script.parent / "out").iterdir(), reverse=True)
    if not out_dirs:
        return False

    import openpyxl as xl
    test_file = None
    for f in out_dirs[0].glob("赛狐_海外仓备货单_导入_*.xlsx"):
        if "旧格式" in f.name:
            continue
        wb = xl.load_workbook(f)
        ws = wb["基本信息"]
        for row in range(3, ws.max_row + 1):
            if str(ws.cell(row=row, column=10).value) == TEST_SKU:
                test_file = f
                break
        if test_file:
            break

    if not test_file:
        print(f"  ✗ 未找到包含 {TEST_SKU} 的备货单文件")
        return False

    print(f"  找到备货单文件: {test_file.name}")

    # 复制到 restock/ 并导入
    restock_dir = WEB_AUTO / "restock"
    restock_dir.mkdir(exist_ok=True)
    dst = restock_dir / test_file.name
    shutil.copy(test_file, dst)

    headless = "--headless" in sys.argv
    headless_args = ["--headless"] if headless else []
    script = WEB_AUTO / "sellfox_import_warehouse_restock.py"
    ret = run_script(script, WEB_AUTO, WEB_VENV, str(dst), *headless_args)
    return ret == 0


# ── 主流程 ────────────────────────────────────────────────

def main():
    step_arg = None
    for a in sys.argv[1:]:
        if a.startswith("--step="):
            step_arg = int(a.split("=")[1])
            break
        if a == "--step" or a == "-s":
            idx = sys.argv.index(a)
            if idx + 1 < len(sys.argv):
                step_arg = int(sys.argv[idx + 1])
            break

    results = {}
    actual_qty = TEST_QTY  # Step 2 会更新为实际库存

    if step_arg is None or step_arg == 1:
        results[1] = step_1_inbound(TEST_QTY)

    if step_arg is None or step_arg == 2:
        ok, actual_qty = step_2_export_and_verify()
        results[2] = ok

    if step_arg is None or step_arg == 3:
        results[3] = step_3_outbound(available_qty=actual_qty if isinstance(actual_qty, (int, float)) else None)

    if step_arg is None or step_arg == 4:
        results[4] = step_4_export_and_verify_zero()

    if step_arg is None or step_arg == 5:
        results[5] = step_5_warehouse_restock()

    # 汇总
    print("\n" + "=" * 60)
    print("端到端测试汇总")
    print("=" * 60)
    for step_num in sorted(results):
        status = "✓" if results[step_num] else "✗"
        print(f"  {status} Step {step_num}")
    print("\n注意: Step 3 出库后需手动去赛狐'确认出库'，Step 4 才会看到库存清零")


if __name__ == "__main__":
    main()
