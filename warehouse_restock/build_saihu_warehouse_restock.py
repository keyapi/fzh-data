"""
build_saihu_warehouse_restock.py
EN BOM Cost List → 三成本拆分 → 赛狐海外仓备货单导入

数据流:
  EN BOM成本列表 → 成本借用 → 三成本拆分
    → 通途仓库映射 → 赛狐SKU白名单
    → 填入海外仓备货单模板 → 输出

使用:
  python build_saihu_warehouse_restock.py
"""
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

os.chdir(os.path.dirname(os.path.abspath(__file__)))

DATA_DIR = Path("数据源")
OUT_BASE = Path("out")
TEMPLATE_FILE = Path("数据源样例/赛狐_海外仓备货单_模板.xlsx")

WAREHOUSE_CONFIG = {
    "USNJ": {
        "aliases": {"USNJ", "CENTRADE", "美东", "美东USNJ"},
        "label": "CENTRADE",
        "processing": "美东加工成本, USNJ",
        "cover_freight": "头程皮壳运费, 美东USNJ",
        "sfg_freight":   "头程半成品运费, 美东USNJ",
        "fg_freight":    "头程成品运费, 美东USNJ",
        "cover_total":   "发皮壳尾程前成本, 美东USNJ",
        "fg_total":      "发成品尾程前成本, 美东USNJ",
    },
    "USTX": {
        "aliases": {"USTX", "DANEEY", "美中", "美中USTX"},
        "label": "DANEEY",
        "processing": "美中加工成本, USTX",
        "cover_freight": "头程皮壳运费, 美中USTX",
        "sfg_freight":   "头程半成品运费, 美中USTX",
        "fg_freight":    "头程成品运费, 美中USTX",
        "cover_total":   "发皮壳尾程前成本, 美中USTX",
        "fg_total":      "发成品尾程前成本, 美中USTX",
    },
    "PL": {
        "aliases": {"PL", "POLAND", "波兰", "波兰PL"},
        "label": "POLAND",
        "processing": "波兰加工成本, PL",
        "cover_freight": "头程皮壳运费, 波兰PL",
        "sfg_freight":   "头程半成品运费, 波兰PL",
        "fg_freight":    "头程成品运费, 波兰PL",
        "cover_total":   "发皮壳尾程前成本, 波兰PL",
        "fg_total":      "发成品尾程前成本, 波兰PL",
    },
}

SAFE_QTY = 1000  # 备货数量

# 通途 SKU 已知人工后缀（匹配时剥离）
# -out/-淘汰: 通途人工标记，安全去除
# -Cover/-Foam: 通途为拆多包裹发货加的皮壳/海绵标记，
#   EN上成品定义是 物料组-面料-尺寸-颜色，不含半成品后缀
KNOWN_SKU_SUFFIXES = ["-淘汰", "-out", "-Cover", "-Foam"]

def _clean_sku(sku: str) -> str:
    """剥离已知人工后缀，用于跨系统 SKU 匹配。不影响输出 SKU。"""
    s = str(sku).strip()
    for suffix in KNOWN_SKU_SUFFIXES:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
            break
    return s

# ── 工具函数 ──────────────────────────────────────────

def norm_col(name: str) -> str:
    return (
        str(name).strip()
        .replace("\n", ",").replace("\r", "")
        .replace(" ", "").replace("-", ",")
    )

def build_col_map(df: pd.DataFrame) -> dict:
    return {norm_col(c): c for c in df.columns}

def _fmt(v):
    """格式化 — 整数不显示小数点。"""
    if pd.isna(v) or v is None:
        return None
    f = float(v)
    return int(f) if f == int(f) else round(f, 2)

def auto_select(pattern: str) -> Path:
    d = DATA_DIR
    candidates = [f for f in d.glob(pattern) if not f.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"未找到匹配 '{pattern}' 的文件于 {d}")
    return max(candidates, key=lambda f: f.stat().st_mtime)


# ── 成本借用 ──────────────────────────────────────────

def borrow_costs(df: pd.DataFrame, cost_columns: list[str]) -> tuple[pd.DataFrame, list[dict]]:
    """同重量模板内借用: 0/空 → 同组同列非0最大值。"""
    records = []
    # 用子串匹配找重量模板列
    col_wt = None
    for c in df.columns:
        if "重量模板" in str(c):
            col_wt = c
            break
    if col_wt is None:
        return df, records

    df = df.copy()
    df["_wt_key"] = df[col_wt].apply(
        lambda x: str(x).replace("ZLMB#", "").strip() if pd.notna(x) else None
    )

    for col in cost_columns:
        if col not in df.columns:
            continue
        for wt_key, grp in df.groupby("_wt_key"):
            max_val = grp[col].replace(0, pd.NA).dropna().max()
            if pd.isna(max_val) or max_val <= 0:
                continue
            mask = (df["_wt_key"] == wt_key) & (df[col].fillna(0) == 0)
            for idx in df[mask].index:
                records.append({
                    "产品编号": df.at[idx, "产品编号"],
                    "重量模板": df.at[idx, col_wt],
                    "借用列": col,
                    "借用值": max_val,
                })
                df.at[idx, col] = max_val

    df.drop(columns=["_wt_key"], inplace=True)
    return df, records


# ── 三成本拆分 ────────────────────────────────────────

def split_cost_for_row(row, col_map: dict, wh_key: str) -> dict | None:
    wh = WAREHOUSE_CONFIG[wh_key]
    method = str(row.get(col_map.get(norm_col("绍兴发货方式"), ""), "")).strip()
    if not method:
        return None  # 发货方式空 → 无法拆分

    def _v(label):
        c = col_map.get(norm_col(label))
        return float(row.get(c, 0) or 0) if c else 0.0

    cost_fg = _v("皮壳成本")
    cost_sfg = _v("绍兴包装半成品成本")
    sx_total = _v("绍兴总成本")
    processing = _v(wh["processing"])

    if method == "成品":
        sx_cost = sx_total
        freight = _v(wh["fg_freight"])
        process_cost = 0.0
    elif method == "半成品":
        sx_cost = cost_fg + cost_sfg
        freight = _v(wh["sfg_freight"])
        if freight == 0:
            freight = _v(wh["cover_freight"])  # 半成品头程缺失时用皮壳
        process_cost = processing
    else:  # 皮壳 或空
        sx_cost = cost_fg
        freight = _v(wh["cover_freight"])
        process_cost = processing

    return {
        "sku": row.get(col_map.get(norm_col("产品编号"), ""), ""),
        "warehouse": wh["label"],
        "shipping_method": method,
        "sx_shipping_cost": _fmt(sx_cost),          # 绍兴发货成本 (单价)
        "first_freight": _fmt(freight),              # 头程运费 (单价)
        "overseas_processing": _fmt(process_cost),   # 国外加工成本 (单价)
    }


# ── 通途仓库映射 ─────────────────────────────────────

WAREHOUSE_REVERSE = {
    "CENTRADE": "CENTRADE",
    "FZHPoland-covers": "POLAND",
    "FZH-DANEEY-皮壳仓库": "DANEEY",
    "FZH-DANEEY-退货产品仓": "DANEEY",
    "FZH-DANEEY-成品仓": "DANEEY",
    "FZH-DANEEY-半成品仓": "DANEEY",
}

def get_warehouse_sku_map(df_tongtu: pd.DataFrame) -> tuple[dict[str, set[str]], list[dict]]:
    """通途 → {通途SKU(清理后): {仓库集合}} + 后缀清理记录。"""
    result = {}
    cleaned_map = {}  # {原始SKU: (清理后SKU, {仓库集合})}
    for _, row in df_tongtu.iterrows():
        sku = str(row.get("SKU", "")).strip()
        wh_raw = str(row.get("仓库", "")).strip()
        wh = WAREHOUSE_REVERSE.get(wh_raw)
        if not sku or not wh:
            continue
        clean = _clean_sku(sku)
        result.setdefault(clean, set()).add(wh)
        if clean != sku:
            cleaned_map.setdefault(sku, {})["clean"] = clean
            cleaned_map.setdefault(sku, {}).setdefault("warehouses", set()).add(wh)
    cleaned = [
        {"通途原始SKU": orig, "清理后SKU": info["clean"],
         "仓库": ", ".join(sorted(info["warehouses"]))}
        for orig, info in cleaned_map.items()
    ]
    return result, cleaned


# ── 填入模板 ──────────────────────────────────────────

# 格式 → 文件名后缀
FMT_SUFFIX = {
    1: "1三成本分开",
    2: "2加工并入采购",
    3: "3三成本全并入采购",
}


def _build_col_index(ws) -> dict[str, int]:
    """读取模板第2行表头，返回 {列名: 列号} 映射。"""
    idx = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=2, column=col).value or "").strip()
        if val:
            idx[val] = col
    return idx


def _fill_single_file(rows: list[dict], out_path: Path, fmt_version: int = 2):
    """复制模板 → 写入数据 → 保存。按列名查找列号，不硬编码。"""
    shutil.copy(TEMPLATE_FILE, out_path)
    import openpyxl
    wb = openpyxl.load_workbook(out_path)
    ws = wb["基本信息"]
    ci = _build_col_index(ws)  # {列名: 列号}

    for i, r in enumerate(rows):
        row_idx = 3 + i
        ws.cell(row=row_idx, column=ci["*收货仓库"], value=r["warehouse"])
        ws.cell(row=row_idx, column=ci["*头程分摊方式"], value="自定义")
        ws.cell(row=row_idx, column=ci["*税费分摊方式"], value="自定义")
        ws.cell(row=row_idx, column=ci["*SKU"], value=r["sku"])
        ws.cell(row=row_idx, column=ci["*备货数量"], value=SAFE_QTY)

        sx = r.get("sx_shipping_cost") or 0
        freight = r.get("first_freight") or 0
        proc = r.get("overseas_processing") or 0

        if fmt_version == 3:
            # 兜底: 三成本全累加到采购单价, 头程留空
            ws.cell(row=row_idx, column=ci["指定采购单价"], value=_fmt(sx + freight + proc))
        elif fmt_version == 2:
            # 默认: 绍兴+加工→采购单价, 头程→单个头程费用
            ws.cell(row=row_idx, column=ci["指定采购单价"], value=_fmt(sx + proc))
            ws.cell(row=row_idx, column=ci["单个头程费用"], value=_fmt(freight))
        else:
            # 旧格式: 物流费用+其他费用×1000, 赛狐自动算头程
            ws.cell(row=row_idx, column=ci["指定采购单价"], value=r.get("sx_shipping_cost"))
            ws.cell(row=row_idx, column=ci["物流费用"], value=_fmt(freight * SAFE_QTY))
            ws.cell(row=row_idx, column=ci["其他费用"], value=_fmt(proc * SAFE_QTY))

    wb.save(out_path)


MAX_BATCH = 500  # 赛狐限制: 单个备货单不超过500条

def fill_templates_by_warehouse(rows: list[dict], out_dir: Path, stamp: str, fmt: str = "2") -> list[Path]:
    """按仓库拆文件，超过500条拆分批次。
    fmt='1'/'2'/'3' 仅该格式, 'all' 三种都输出。"""
    out_paths = []
    for wh in ["CENTRADE", "DANEEY", "POLAND"]:
        wh_rows = [r for r in rows if r["warehouse"] == wh]
        if not wh_rows:
            continue
        batches = [wh_rows[i:i+MAX_BATCH] for i in range(0, len(wh_rows), MAX_BATCH)]
        for bi, batch in enumerate(batches):
            tag = f"_{wh}" if len(batches) == 1 else f"_{wh}_p{bi+1}"
            base = out_dir / f"赛狐_海外仓备货单_导入{tag}_{stamp}"

            for fv in [1, 2, 3]:
                if fmt not in (str(fv), "all"):
                    continue
                path = Path(f"{base}_{FMT_SUFFIX[fv]}.xlsx")
                _fill_single_file(batch, path, fmt_version=fv)
                out_paths.append(path)
                print(f"  {wh}: {len(batch)} 条 → {path.name}")

    return out_paths


# ── 主流程 ────────────────────────────────────────────

def main():
    print("=" * 60)
    print("赛狐 海外仓备货单导入 — 生成中")
    print("=" * 60)

    # 1. 读取数据源
    f_bom = auto_select("EN产品BOM成本列表*.xlsx")
    f_tongtu = auto_select("通途合并库存结存清单*.xlsx")
    f_saihu = auto_select("商品导出*.xlsx")

    print(f"BOM:  {f_bom.name}")
    print(f"通途: {f_tongtu.name}")
    print(f"赛狐: {f_saihu.name}")

    df_bom = pd.read_excel(f_bom)
    df_tongtu = pd.read_excel(f_tongtu)
    df_saihu = pd.read_excel(f_saihu)

    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"模板不存在: {TEMPLATE_FILE}")

    # 2. 赛狐 SKU 白名单
    sai_sku_col = "SKU" if "SKU" in df_saihu.columns else df_saihu.columns[0]
    sai_whitelist = set(df_saihu[sai_sku_col].dropna().astype(str).str.strip())
    print(f"赛狐SKU白名单: {len(sai_whitelist)} 个")

    # 3. 通途 → SKU 仓库映射
    wh_map, suffix_cleaned = get_warehouse_sku_map(df_tongtu)
    print(f"通途 SKU-仓库 映射: {len(wh_map)} 个 SKU")
    if suffix_cleaned:
        print(f"后缀清理匹配: {len(suffix_cleaned)} 条 (详见问题报告)")

    # 4. BOM 表头标准化
    col_map = build_col_map(df_bom)

    # 5. 成本借用
    all_cost_cols = []
    for wh_key in WAREHOUSE_CONFIG:
        wh = WAREHOUSE_CONFIG[wh_key]
        all_cost_cols.extend([wh["processing"], wh["cover_freight"], wh["sfg_freight"], wh["fg_freight"]])
    all_cost_cols.extend(["皮壳成本", "绍兴包装半成品成本", "绍兴总成本", "绍兴包装成品成本"])
    all_cost_cols = [c for c in all_cost_cols if c in df_bom.columns]

    df_bom, borrow_records = borrow_costs(df_bom, all_cost_cols)
    print(f"成本借用: {len(borrow_records)} 个单元格")

    # 6. 逐行拆分三成本（按 SKU+仓库 粒度：每个目标仓要么输出、要么报告跳过原因）
    TARGET_WH = ["CENTRADE", "DANEEY", "POLAND"]
    results = []
    issues = []  # [(SKU, 原因, 仓库, 原因详情, 客户物料号)]
    suffix_records = []  # [(通途原始SKU, 清理后SKU, 客户物料号, 匹配的BOM产品编号)]

    for _, row in df_bom.iterrows():
        sku = str(row.get(col_map.get(norm_col("产品编号"), ""), "")).strip()
        if not sku or sku == "nan":
            continue

        # 发货方式为空 → 3仓全跳过
        method_col = col_map.get(norm_col("绍兴发货方式"))
        if method_col is None or pd.isna(row.get(method_col)):
            for wh in TARGET_WH:
                issues.append((sku, "发货方式为空", wh, "BOM绍兴发货方式列为空", ""))
            continue

        # 赛狐白名单 → 3仓全跳过
        if sku not in sai_whitelist:
            for wh in TARGET_WH:
                issues.append((sku, "不在赛狐商品列表", wh, "赛狐商品导出中无此SKU", ""))
            continue

        # 通过 BOM 的 客户物料号 匹配通途 SKU → 找仓库
        cust_col = None
        for c in df_bom.columns:
            if "客户物料号" in str(c):
                cust_col = c
                break
        cust_sku = str(row.get(cust_col, "")).strip() if cust_col else ""
        clean_cust = _clean_sku(cust_sku) if cust_sku else ""
        wh_set = wh_map.get(clean_cust, set()) if clean_cust else set()

        # 逐仓处理
        for wh_label in TARGET_WH:
            if wh_label not in wh_set:
                # 该仓通途无此 SKU 记录
                reason_detail = "客户物料号为空" if not cust_sku else f"客户物料号={cust_sku}"
                issues.append((sku, "通途无此仓SKU记录", wh_label, reason_detail, cust_sku))
                continue

            wh_key = None
            for k, v in WAREHOUSE_CONFIG.items():
                if v["label"] == wh_label:
                    wh_key = k
                    break
            if wh_key is None:
                continue

            cost = split_cost_for_row(row, col_map, wh_key)
            if cost is None:
                issues.append((sku, "无法拆分成本", wh_label, "发货方式无法识别", cust_sku))
                continue
            results.append(cost)

    print(f"拆分结果: {len(results)} 条")

    # 去重：同一 (SKU, 仓库) 只保留一条（同 SKU 可能通过多通途 SKU 命中同仓）
    seen = set()
    deduped = []
    for r in results:
        key = (r["sku"], r["warehouse"])
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    dup_count = len(results) - len(deduped)
    if dup_count:
        print(f"去重: 移除 {dup_count} 条重复 → 剩余 {len(deduped)} 条")
    results = deduped

    # 可选参数解析
    filter_sku = None
    output_fmt = "2"  # 默认仅格式2
    for a in sys.argv[1:]:
        if a.startswith("--sku="):
            filter_sku = a.split("=", 1)[1]
        elif a.startswith("--fmt="):
            output_fmt = a.split("=", 1)[1]  # "1", "2", "3", "all"
    if filter_sku:
        results = [r for r in results if r["sku"] == filter_sku]
        print(f"SKU 过滤: --sku={filter_sku} → 剩余 {len(results)} 条")

    # 7. 输出（按仓库拆文件）
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    OUT_DIR = OUT_BASE / stamp
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    out_paths = fill_templates_by_warehouse(results, OUT_DIR, stamp, fmt=output_fmt)

    # 8. 问题报告（始终生成，含后缀清理记录）
    report_path = OUT_DIR / f"warehouse_restock_问题报告_{stamp}.xlsx"
    df_issues = pd.DataFrame(issues, columns=["SKU", "原因", "仓库", "原因详情", "客户物料号"]) if issues else pd.DataFrame(columns=["SKU", "原因", "仓库", "原因详情", "客户物料号"])
    df_borrow = pd.DataFrame(borrow_records) if borrow_records else pd.DataFrame(columns=["产品编号", "重量模板", "借用列", "借用值"])
    df_suffix = pd.DataFrame(suffix_cleaned) if suffix_cleaned else pd.DataFrame(columns=["通途原始SKU", "清理后SKU", "仓库"])
    with pd.ExcelWriter(report_path) as writer:
        df_summary = df_issues["原因"].value_counts().reset_index()
        df_summary.columns = ["原因", "数量"]
        df_summary.to_excel(writer, sheet_name="汇总", index=False)
        df_issues.to_excel(writer, sheet_name="跳过明细", index=False)
        df_borrow.to_excel(writer, sheet_name="成本借用记录", index=False)
        df_suffix.to_excel(writer, sheet_name="后缀清理匹配", index=False)
    print(f"\n问题报告: {report_path} (跳过{len(issues)}条, 后缀清理{len(suffix_cleaned)}条, 成本借用{len(borrow_records)}个)")

    # 9. 汇总
    print(f"\n{'=' * 60}")
    print("赛狐 海外仓备货单 — 生成完成")
    print(f"{'=' * 60}")
    total = len(results)
    for wh in ["CENTRADE", "DANEEY", "POLAND"]:
        n = sum(1 for r in results if r["warehouse"] == wh)
        print(f"  {wh}: {n} 条")
    print(f"  合计: {total} 条")
    print(f"  跳过: {len(issues)} 条")
    if borrow_records:
        print(f"  成本借用: {len(borrow_records)} 个单元格")


if __name__ == "__main__":
    main()
