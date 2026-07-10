#!/usr/bin/env python3
"""生成月度工单排查报告 Excel v4 — 统一列格式 + open_material_qty全覆盖

读取 fetch.py 生成的标准 JSON 文件，输出分类 Excel 报告。

用法:
  python erpnext/scripts/gen_report.py              # 默认 /tmp/erpnext_*.json
  python erpnext/scripts/gen_report.py --month 2026-06  # 从 --month 派生路径
  python erpnext/scripts/gen_report.py --wo /path/to/wo.json --jc /path/to/jc.json
"""
import argparse, json, os, sys, urllib.request, ssl
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Default paths (from fetch.py output) ─────────────────
BASE = Path(__file__).resolve().parent.parent
DATA_DIR = BASE / "data"
TMP_DIR = Path("/tmp")

def resolve_paths(month: str | None = None, wo_path: str | None = None, jc_path: str | None = None):
    """Resolve input JSON paths. Priority: explicit args > month-derived > defaults."""
    if wo_path and jc_path:
        return Path(wo_path), Path(jc_path)

    if month:
        return (
            TMP_DIR / f"erpnext_wo_data_{month}.json",  # future-proofing
            TMP_DIR / f"erpnext_jc_data_{month}.json",
        )

    # Default: look for files from fetch.py
    wo_file = TMP_DIR / "erpnext_wo_data.json"
    jc_file = TMP_DIR / "erpnext_jc_data.json"
    if wo_file.is_file() and jc_file.is_file():
        return wo_file, jc_file

    # Fallback to legacy paths from conversational workflow
    wo_file = TMP_DIR / "wo_full.json"
    jc_file = TMP_DIR / "all_job_cards3.json"
    if wo_file.is_file():
        return wo_file, jc_file

    raise FileNotFoundError(
        "未找到数据文件。请先运行:\n"
        "  uv run python erpnext/scripts/fetch.py --month 2026-06\n"
        f"或手动指定: --wo <path> --jc <path>"
    )

# ── Load data ─────────────────────────────────────────────
def load_data(wo_path, jc_path):
    """Load WO and JC data from JSON files."""
    with open(wo_path, 'r', encoding='utf-8') as f:
        wo_full = json.load(f)
    with open(jc_path, 'r', encoding='utf-8') as f:
        all_jc = json.load(f)
    return wo_full, all_jc

# ── Parse CLI args ────────────────────────────────────────
parser = argparse.ArgumentParser(description="生成 ERPNext 工单排查报告")
parser.add_argument("--month", help="目标月份 (如 2026-06)")
parser.add_argument("--wo", help="工单 JSON 路径")
parser.add_argument("--jc", help="Job Card JSON 路径")
parser.add_argument("--output", help="输出路径 (默认 erpnext/data/)")
args = parser.parse_args()

wo_path, jc_path = resolve_paths(args.month, args.wo, args.jc)
print(f"读取数据: WO={wo_path}, JC={jc_path}")

wo_full, all_jc = load_data(wo_path, jc_path)
print(f"工单: {len(wo_full)} 条, JC数据: {len(all_jc)} 个工作单")

# Derive creation dates from wo_full
creation_data = {}
for wo_name, d in wo_full.items():
    cre = d.get('creation', '')
    if cre:
        creation_data[wo_name] = cre

# zero_mat_detail = all WOs with ops data (wo_full already has per-WO detail)
zero_mat_detail = {wo: d for wo, d in wo_full.items() if isinstance(d, dict)}

# Categories (dynamically from data, not month-specific)
fg_wos = [wo for wo, d in zero_mat_detail.items()
          if isinstance(d, dict) and d.get('ops_count', len(d.get('operations', d.get('ops', [])))) == 0]
semi_wos = [wo for wo, d in zero_mat_detail.items()
            if isinstance(d, dict) and d.get('ops_count', len(d.get('operations', d.get('ops', [])))) > 0]

# mixed and nc_touched: WOs with both real and virtual JCs
def classify_jc_owner(wo_name):
    jcs = all_jc.get(wo_name, [])
    if not jcs: return 'no_jc'
    has_yang = any('yangyisen' in jc.get('owner','') for jc in jcs)
    has_real = any('yangyisen' not in jc.get('owner','') for jc in jcs)
    if has_yang and has_real: return 'mixed'
    if has_yang: return 'virtual_only'
    return 'real_only'

mixed_wos = [wo for wo in wo_full if classify_jc_owner(wo) == 'mixed' and wo not in semi_wos and wo not in fg_wos]
nc_touched = [wo for wo in wo_full
              if wo not in fg_wos and wo not in semi_wos
              and classify_jc_owner(wo) == 'virtual_only'
              and wo_full[wo].get('status', '') != 'Completed']
scan_clean = [w for w in wo_full if w not in nc_touched and w not in mixed_wos and w not in fg_wos and w not in semi_wos]

# Fallback: if no WOs classified, include all
if not fg_wos and not semi_wos:
    fg_wos = [wo for wo, d in zero_mat_detail.items() if isinstance(d, dict)]
    semi_wos = []

# Universal column headers (consistent across ALL sheets)
HDR_ALL = ['工单号', '创建时间', '产品类型', '产品编码', '状态', '工序数',
           '计划量', '开料量', '产出量', 'JC数', 'JC类型', '数据可信度', '处理建议/说明']
W_ALL = [16, 12, 14, 50, 14, 8, 10, 10, 10, 8, 20, 38, 60]

# Styles
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(bold=True, size=11, color='FFFFFF')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ylw_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
grn_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
org_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')  # orange for special flags
def border(): return Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
def hdr(ws, row=1):
    for c, h in enumerate(HDR_ALL, 1):
        cl = ws.cell(row=row, column=c, value=h)
        cl.font, cl.fill, cl.alignment, cl.border = hdr_font, hdr_fill, Alignment(horizontal='center', vertical='center', wrap_text=True), border()
def srow(ws, row, vals, fill=None):
    for c, v in enumerate(vals, 1):
        cl = ws.cell(row=row, column=c, value=v)
        cl.border, cl.alignment = border(), Alignment(vertical='center', wrap_text=True)
        if fill: cl.fill = fill
def setw(ws):
    for i, w in enumerate(W_ALL, 1): ws.column_dimensions[get_column_letter(i)].width = w

# ---- Helpers ----
def jc_analysis(wo_name):
    jcs = all_jc.get(wo_name, [])
    if not jcs: return '无JC', '', 0
    owners = {}
    for jc in jcs:
        o = jc.get('owner','?'); owners[o] = owners.get(o,0)+1
    has_yang = any('yangyisen' in o for o in owners)
    has_real = any('yangyisen' not in o for o in owners)
    total = sum(owners.values())
    if has_yang and not has_real:
        return '纯虚拟(杨义森)', f'全部{total}条JC,HR-EMP-00001', total
    elif has_yang and has_real:
        rc = sum(c for o,c in owners.items() if 'yangyisen' not in o)
        yc = sum(c for o,c in owners.items() if 'yangyisen' in o)
        rn = ', '.join([o.split('@')[0] for o in owners if 'yangyisen' not in o])
        return '混合(真实+虚拟)', f'真实{rc}({rn})+杨义森{yc}', total
    else:
        return '纯真实工人', f'全部{total}条真实工人', total

def get_item(wo_name):
    for src in [wo_full, zero_mat_detail]:
        d = src.get(wo_name, {})
        if isinstance(d, dict):
            item = d.get('item','') or d.get('production_item','')
            if item: return item
    return ''

def get_wo_data(wo_name):
    if wo_name in wo_full: return wo_full[wo_name]
    if wo_name in zero_mat_detail: return zero_mat_detail[wo_name]
    return {}

def fmt_cre(wo_name):
    c = creation_data.get(wo_name, '?'); return c[:10] if len(c) > 10 else c

def ptype(item):
    if item.startswith('PK#'): return '半成品(皮壳)'
    elif item.startswith('ND#'): return '半成品(内胆)'
    elif item.startswith('KS'): return '成品(fg)'
    return '其他'

def ops_count(d):
    ops = d.get('operations') or d.get('ops', [])
    return len(ops) if ops else d.get('ops_count', 0)

def ops_str(d):
    ops = d.get('operations') or d.get('ops', [])
    if not ops: return str(d.get('ops_count', 0))
    return ', '.join(o.get('name','') or o.get('operation','') for o in ops[:5])

wb = Workbook()

# ===== Sheet 1: 总览 =====
ws = wb.active; ws.title = '总览'; hdr(ws); row = 2
for cat, fill, wo_list in [
    ('成品fg-正常完工', grn_fill, fg_wos),
    ('正常扫码-工序瓶颈', None, scan_clean),
    ('非Completed+一键完工', ylw_fill, nc_touched),
    ('混合-真实+虚拟JC', ylw_fill, mixed_wos),
    ('半成品-一键完工', red_fill, semi_wos),
]:
    for wo_name in wo_list:
        d = get_wo_data(wo_name); item = get_item(wo_name)
        jc_type, jc_detail, jc_count = jc_analysis(wo_name)
        cre = fmt_cre(wo_name); qty = d.get('qty',0); omq = d.get('open_material_qty',0); prod = d.get('produced_qty',0)
        ops_n = ops_count(d); st = d.get('status',''); ops_names = ops_str(d)
        has_sewing = '缝制' in str(ops_names)

        if cat == '成品fg-正常完工':
            trust = '可信(成品无工序,合理手动完工)'
            note = '成品fg,0工序,open_mat=0正常; 待补扫码流程'
        elif cat == '正常扫码-工序瓶颈':
            pending = [o for o in d.get('operations',[]) if o.get('status')!='Completed' and o.get('completed_qty',0)==0]
            rp = [o for o in pending if o['name'] not in ('质检发现问题','通用返工')]
            if rp: note = f'瓶颈工序: {rp[0]["name"]}'
            elif omq > 0 and omq < qty: note = f'裁剪不足(差{qty-omq:.0f})'
            else: note = '工序已完成到质检'
            trust = '可信(扫码报工)'
        elif cat == '半成品-一键完工':
            if has_sewing:
                trust = '不可信(遗留虚拟工序+纯虚拟JC)'
                note = f'⚠工艺路线错误: 含"缝制"假工序({ops_names}); open_mat={omq}异常(应为>0); 需更新BOM+物理盘点'
            else:
                trust = '不可信(纯虚拟JC)'
                note = f'⚠ 开料={omq}(异常,半成品应>0); 从未扫码报工; 工序={ops_names}; 需物理盘点'
        elif cat == '混合-真实+虚拟JC':
            trust = '开料可信,工序量被覆盖'
            diff = omq - qty if omq else 0
            note = f'真实工人JC+杨义森虚拟JC混合; 开料={omq:.0f}(差{diff:+.0f}); 以真实JC为准'
        else:  # nc+杨义森
            trust = '虚报(纯虚拟JC)'
            gap = abs(omq - qty) if omq else qty
            note = f'open_mat={omq:.0f}, 虚报差{gap:.0f}; ' + jc_detail

        srow(ws, row, [wo_name, cre, ptype(item), item, st, ops_n, qty, omq, prod, jc_count, jc_type, trust, note], fill)
        row += 1
setw(ws)

# ===== Sheet 2: 半成品-一键完工(23条) =====
ws2 = wb.create_sheet('半成品-一键完工'); hdr(ws2); row = 2
for wo_name in sorted(semi_wos):
    d = zero_mat_detail.get(wo_name, {}); item = d.get('production_item','')
    jc_type, jc_detail, jc_count = jc_analysis(wo_name)
    cre = fmt_cre(wo_name); qty = d.get('qty',0); omq = d.get('open_material_qty',0); prod = d.get('produced_qty',0)
    ops_n = d.get('ops_count',0); st = d.get('status',''); ops_names = ops_str(d)
    has_sewing = '缝制' in str(ops_names)

    # Determine fill: red for normal, orange for special flags
    if has_sewing:
        fill_use = org_fill
        trust = '遗留虚拟工序: "缝制"'
        note = f'⚠ BOM/工艺路线未更新, 含假工序"缝制"; ops={ops_names}; open_mat=0(异常)'
    else:
        fill_use = red_fill
        trust = '纯虚拟JC'
        note = f'⚠ 开料={omq}(异常! 半成品应>0); 从未扫码报工; 工序={ops_names}; 需核实为何开料为0'

    srow(ws2, row, [wo_name, cre, ptype(item), item, st, ops_n, qty, omq, prod, jc_count, trust, '不可信', note], fill_use)
    row += 1
setw(ws2)

# ===== Sheet 3: 成品fg-正常完工(20条) =====
ws3 = wb.create_sheet('成品fg-正常完工'); hdr(ws3); row = 2
for wo_name in sorted(fg_wos):
    d = zero_mat_detail.get(wo_name, {}); item = d.get('production_item','')
    cre = fmt_cre(wo_name); qty = d.get('qty',0); omq = d.get('open_material_qty',0); prod = d.get('produced_qty',0)
    st = d.get('status','')
    srow(ws3, row, [wo_name, cre, '成品(fg)', item, st, 0, qty, omq, prod, 0, '无JC(成品无工序)',
        '可信', '成品组装(皮壳+内胆+充棉), open_mat=0正常; 待郭昌坤扫码完工流程'], grn_fill)
    row += 1
setw(ws3)

# ===== Sheet 4: 正常扫码-工序瓶颈(20条) =====
ws4 = wb.create_sheet('正常扫码-工序瓶颈'); hdr(ws4); row = 2
for wo_name, d in sorted(wo_full.items()):
    if wo_name in nc_touched: continue
    item = get_item(wo_name); cre = fmt_cre(wo_name); qty = d.get('qty',0); omq = d.get('open_material_qty',0); prod = d.get('produced_qty',0)
    ops_n = ops_count(d); st = d.get('status',''); pct = prod/qty*100 if qty>0 else 0
    pending = [o for o in d.get('operations',[]) if o.get('status')!='Completed' and o.get('completed_qty',0)==0]
    rp = [o for o in pending if o['name'] not in ('质检发现问题','通用返工')]
    if rp: note = f'瓶颈工序: {rp[0]["name"]}'
    elif omq > 0 and omq < qty: note = f'裁剪不足(差{qty-omq:.0f})'
    else: note = '工序已完成到质检'
    srow(ws4, row, [wo_name, cre, ptype(item), item, st, ops_n, qty, omq, prod, 0, '纯真实工人', '可信', note])
    row += 1
setw(ws4)

# ===== Sheet 5: 非Completed+一键完工(4条) =====
ws5 = wb.create_sheet('非Completed+一键完工'); hdr(ws5); row = 2
detail = {
    'WO-26-00082': '真实98JC(李清君等83+8sy14+yj0 1)+杨义森7; SE入库10批共216; 真实产量约285-298',
    'WO-26-00254': '真实16JC(1hu-2c76aikeww)+杨义森1; 开料=计划=50, 混合型',
    'WO-26-00748': '真实5JC(李清君等)+杨义森5; 开料17≠计划19, 混合型',
    'WO-26-01349': '真实8JC(李清君等)+杨义森8; 开料18≠计划20, 混合型'
}
for wo_name in nc_touched:
    d = wo_full.get(wo_name,{}); item = get_item(wo_name); cre = fmt_cre(wo_name)
    jc_type, jc_detail, jc_count = jc_analysis(wo_name)
    qty = d.get('qty',0); omq = d.get('open_material_qty',0); prod = d.get('produced_qty',0)
    ops_n = ops_count(d); st = d.get('status','')
    srow(ws5, row, [wo_name, cre, ptype(item), item, st, ops_n, qty, omq, prod, jc_count, jc_type, '虚报/混合', detail.get(wo_name,'')], ylw_fill)
    row += 1
setw(ws5)

# ===== Sheet 6: 混合-真实+虚拟JC(4条) =====
ws6 = wb.create_sheet('混合-真实+虚拟JC'); hdr(ws6); row = 2
detail6 = {
    'WO-26-01532': '真实48JC(105-prd4qxz8w9)+杨义森1; 开料76>计划75(多裁1)',
    'WO-26-01539': '真实90JC(105-prd4qxz8w9)+杨义森4; 开料148>计划140(多裁8)',
    'WO-26-01540': '开料104>计划100(多裁4), 真实+杨义森混合',
    'WO-26-01611': '开料152>计划150(多裁2), 真实+杨义森混合',
}
for wo_name in mixed_wos:
    d = get_wo_data(wo_name); item = get_item(wo_name); cre = fmt_cre(wo_name)
    jc_type, jc_detail, jc_count = jc_analysis(wo_name)
    qty = d.get('qty',0); omq = d.get('open_material_qty',0); prod = d.get('produced_qty',0)
    ops_n = ops_count(d); st = d.get('status','')
    diff = omq - qty if omq else 0
    note = f'开料={omq:.0f}(差{diff:+.0f}); {jc_detail}; 以真实JC为准'
    srow(ws6, row, [wo_name, cre, ptype(item), item, st, ops_n, qty, omq, prod, jc_count, jc_type, '开料可信,工序量被覆盖', note], ylw_fill)
    row += 1
setw(ws6)

# ===== Sheet 7: 探案方法 (8-step) =====
ws7 = wb.create_sheet('探案方法')
hdr8 = ['步骤', '检查项', '数据源/API', '判断逻辑', '典型示例']
def hdr8f(ws, row=1):
    for c, h in enumerate(hdr8, 1):
        cl = ws.cell(row=row, column=c, value=h)
        cl.font, cl.fill, cl.alignment, cl.border = hdr_font, hdr_fill, Alignment(horizontal='center', vertical='center', wrap_text=True), border()
def srow8(ws, row, vals):
    for c, v in enumerate(vals, 1):
        cl = ws.cell(row=row, column=c, value=v)
        cl.border, cl.alignment = border(), Alignment(vertical='center', wrap_text=True)
hdr8f(ws7); row = 2
for m in [
    ['1', '产品类型', 'production_item', 'PK#/ND# → 半成品(有工序,应扫码)\nKS开头 → 成品fg(0工序,组装品)', 'PK#KS0001-HLR: 半成品10道工序\nKS0194-HLR-60: 成品0工序'],
    ['2', '遗留虚拟工序', 'operations', '"缝制"→一键完工时代假工序\n含"缝制"→BOM/工艺路线需更新', 'WO-26-01563: [裁剪,缝制,质检] — BOM未更新'],
    ['3', 'Version活动', 'GET Version\nfilter: owner=yangyisen92', '有杨义森记录→疑似一键完工\n痕迹: 草稿→未开始, end_date 0.017s迭代', 'WO-26-00146: 8次end_date迭代'],
    ['4', 'open_material_qty', 'Work Order.open_material_qty', '半成品=0→⚠异常(裁剪未报开料!)\n半成品>0→真实裁剪量\n成品fg=0→正常(不需开料)', '23条半成品全部open_mat=0(异常!)\n20条成品fg open_mat=0(正常)'],
    ['5', 'JC time_logs.employee', 'GET Job Card/{name}\n→ time_logs[].employee', 'HR-EMP-00001→虚拟员工\n其他→真实员工\n无JC→成品fg或未开工', 'WO-26-00146: 8JC全HR-EMP-00001\nWO-26-00082: 李清君JC=HR-EMP-00109'],
    ['6', 'JC owner(辅助)', 'GET Job Card\nfields: owner', 'owner=杨义森→虚拟JC\nowner=真实用户→扫码JC\n注意: owner≠employee!', 'WO-26-00146: 全部yangyisen92\nWO-26-01532: 48条105-prd4qxz8w9+1条杨义森'],
    ['7', 'Stock Entry', 'GET Stock Entry\nstock_entry_type=Manufacture', 'owner=杨义森→入库量=计划量(不可信)\nowner≠杨义森→真实入库', 'WO-26-00082: yj0入库216(真实)\nWO-26-00146: 杨义森入库100(不可信)'],
    ['8', '交叉验证+分类', '综合以上7步', '成品fg+0工序→正常\n半成品+纯虚拟+open_mat=0→全假+异常\n半成品+开料=0→特殊标记(需排查原因)\n混合→以真实JC/SE为准', '成品20条→正常; 半成品23条→全虚拟+全开料=0(异常); 混合4条→开料可信; 非Completed 4条→虚报'],
]:
    srow8(ws7, row, m); row += 1
for i, w in enumerate([6, 20, 45, 55, 60], 1): ws7.column_dimensions[get_column_letter(i)].width = w

# Determine output filename from month or current data
if args.month:
    month_str = args.month
elif args.output:
    month_str = None
else:
    month_str = "unknown"
output_name = args.output or (f'{month_str}_工单排查报告.xlsx' if month_str else '工单排查报告.xlsx')
DATA_DIR.mkdir(parents=True, exist_ok=True)
output = str(DATA_DIR / output_name)
wb.save(output)
print(f'OK: {output}')
print(f'Sheets: {wb.sheetnames}')
