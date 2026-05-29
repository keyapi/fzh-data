"""
build_saihu_other_outbound.py
赛狐库存明细导出 → 其他出库导入文件（清零库存）

用法:
  python build_saihu_other_outbound.py               # 全量生成
  python build_saihu_other_outbound.py --test SKU     # 仅生成指定 SKU 的测试文件
"""
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

os.chdir(os.path.dirname(os.path.abspath(__file__)))

DATA_DIR = Path("数据源")
OUT_BASE = Path("out")
TEMPLATE_FILE = Path("数据源样例/赛狐_其他出库_模板.xlsx")

TARGET_WH = ["CENTRADE", "DANEEY", "POLAND"]
OUTBOUND_TYPE = "其他出库"


def auto_select(pattern: str) -> Path:
    d = DATA_DIR
    candidates = [f for f in d.glob(pattern) if not f.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"未找到匹配 '{pattern}' 的文件于 {d}")
    return max(candidates, key=lambda f: f.stat().st_mtime)


def generate_outbound_rows(df_stock: pd.DataFrame) -> tuple[list[dict], list[dict]]:
    """从库存明细中提取需要出库的行。返回 (正常行, 跳过的组合商品行)。"""
    rows = []
    skipped = []
    for _, r in df_stock.iterrows():
        sku = str(r.get("SKU", "")).strip()
        wh = str(r.get("仓库", "")).strip()
        out_qty = int(r.get("可用数", 0) or 0)
        defective = int(r.get("次品数", 0) or 0)

        if wh not in TARGET_WH:
            continue
        if out_qty == 0 and defective == 0:
            continue

        # 过滤组合商品 SKU（-ALL 后缀，其他出库不支持）
        if sku.endswith("-ALL"):
            skipped.append({
                "SKU": sku,
                "仓库": wh,
                "可用数": available,
                "原因": "组合商品SKU（-ALL后缀），其他出库不支持",
            })
            continue

        shop = str(r.get("店铺", "")).strip() if pd.notna(r.get("店铺")) else ""
        fnsku = str(r.get("FNSKU", "")).strip() if pd.notna(r.get("FNSKU")) else ""

        rows.append({
            "sku": sku,
            "warehouse": wh,
            "店铺": shop if shop and shop != "nan" else "",
            "FNSKU": fnsku if fnsku and fnsku != "nan" else "",
            "可用出库量": out_qty,
            "次品出库量": defective,
        })
    return rows, skipped


def _fill_single_file(rows: list[dict], out_path: Path, batch_id: str = ""):
    shutil.copy(TEMPLATE_FILE, out_path)
    import openpyxl
    wb = openpyxl.load_workbook(out_path)
    ws = wb["sheet1"]

    order_no = f"OB{datetime.now().strftime('%Y%m%d%H%M%S')}{batch_id}"
    for i, r in enumerate(rows):
        row_idx = 2 + i
        ws.cell(row=row_idx, column=1, value=order_no)
        ws.cell(row=row_idx, column=2, value=r["warehouse"])    # *出库仓库
        ws.cell(row=row_idx, column=3, value=OUTBOUND_TYPE)     # *出库类型
        ws.cell(row=row_idx, column=7, value=r["sku"])          # *SKU
        ws.cell(row=row_idx, column=8, value=r.get("店铺", ""))  # 店铺
        ws.cell(row=row_idx, column=9, value=r.get("FNSKU", "")) # FNSKU
        ws.cell(row=row_idx, column=10, value=r["可用出库量"])   # 可用出库量
        if r["次品出库量"] > 0:
            ws.cell(row=row_idx, column=11, value=r["次品出库量"])

    wb.save(out_path)


def main():
    test_sku = None
    if len(sys.argv) > 2 and sys.argv[1] == "--test":
        test_sku = sys.argv[2]

    f_stock = auto_select("赛狐库存明细*.xlsx")
    df_stock = pd.read_excel(f_stock)
    print(f"库存数据: {f_stock.name} ({len(df_stock)} 行)")

    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"模板不存在: {TEMPLATE_FILE}")

    if test_sku:
        df_stock = df_stock[df_stock["SKU"] == test_sku]
        if df_stock.empty:
            print(f"SKU '{test_sku}' 未在库存数据中找到")
            return
        print(f"测试模式: SKU={test_sku}")

    rows, skipped_combo = generate_outbound_rows(df_stock)
    print(f"出库条目: {len(rows)}")
    if skipped_combo:
        print(f"  跳过的组合商品 SKU (-ALL): {len(skipped_combo)} 个")
        for s in skipped_combo:
            print(f"    {s['仓库']} {s['SKU']} (库存={s['可用数']})")
    if not rows:
        print("无需要出库的条目")
        return

    MAX_PER_BATCH = 1000  # 赛狐限制: 同一临时单号不超过1000条

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    OUT_DIR = OUT_BASE / stamp
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for wh in TARGET_WH:
        wh_rows = [r for r in rows if r["warehouse"] == wh]
        if not wh_rows:
            continue
        # 超过1000条拆批，每个批次用不同临时单号
        batches = [wh_rows[i:i+MAX_PER_BATCH] for i in range(0, len(wh_rows), MAX_PER_BATCH)]
        tag = f"_TEST_{test_sku}" if test_sku else ""
        for bi, batch in enumerate(batches):
            batch_tag = f"_{wh}{tag}" if len(batches) == 1 else f"_{wh}_p{bi+1}{tag}"
            out_path = OUT_DIR / f"赛狐_其他出库_导入{batch_tag}_{stamp}.xlsx"
            _fill_single_file(batch, out_path, batch_id=f"_p{bi+1}" if len(batches) > 1 else "")
            avail = sum(r["可用出库量"] for r in batch)
            defect = sum(r["次品出库量"] for r in batch)
            print(f"  {wh}: {len(batch)} 条, 可用={avail}, 次品={defect} → {out_path.name}")

    if skipped_combo:
        df_skipped = pd.DataFrame(skipped_combo)
        report_path = OUT_DIR / f"other_outbound_跳过组合商品_{stamp}.xlsx"
        df_skipped.to_excel(report_path, sheet_name="组合商品跳过", index=False)
        print(f"\n跳过报告: {report_path}")

    print(f"\n输出目录: {OUT_DIR}")


if __name__ == "__main__":
    main()
