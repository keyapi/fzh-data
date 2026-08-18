from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .candidates_v2 import CandidateScore
from .report import FEEDBACK_VALUES
from .v2 import V2Decision


REVIEW_COLUMNS = [
    "店铺", "站点", "MSKU", "ASIN", "标题", "Amazon主图", "对象类型", "识别原因",
    *[column for index in range(1, 4) for column in (
        f"Top{index} SKU", f"Top{index} 名称", f"Top{index} 分数",
        f"Top{index} 依据", f"Top{index} 冲突"
    )],
    "人工结论", "正确SKU另填", "审核人", "审核时间",
]


def _candidate_columns(row: dict, candidates: tuple[CandidateScore, ...]) -> dict:
    for index in range(3):
        candidate = candidates[index] if index < len(candidates) else None
        row[f"Top{index + 1} SKU"] = candidate.sku if candidate else ""
        row[f"Top{index + 1} 名称"] = candidate.name if candidate else ""
        row[f"Top{index + 1} 分数"] = candidate.score if candidate else 0.0
        row[f"Top{index + 1} 依据"] = " | ".join(candidate.evidence) if candidate else ""
        row[f"Top{index + 1} 冲突"] = candidate.hard_conflicts if candidate else 0
    return row


def _review_row(listing: dict, decision: V2Decision) -> dict:
    row = _candidate_columns(
        {
            "店铺": listing.get("shopId", ""),
            "站点": listing.get("marketplaceId", ""),
            "MSKU": listing.get("sku", ""),
            "ASIN": listing.get("asin", ""),
            "标题": listing.get("title", ""),
            "Amazon主图": listing.get("mainImage", ""),
            "对象类型": decision.object_type,
            "识别原因": " | ".join(decision.object_reasons),
        },
        decision.candidates,
    )
    row.update({"人工结论": "", "正确SKU另填": "", "审核人": "", "审核时间": ""})
    return row


def _dataframe(rows: list[dict], columns: list[str] | None = None) -> pd.DataFrame:
    if columns is None:
        columns = REVIEW_COLUMNS
    return pd.DataFrame(rows, columns=columns)


def write_v2_workbook(
    output: Path,
    records: list[tuple[dict, V2Decision]],
    summary: dict[str, object],
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    buckets: dict[str, list[tuple[dict, V2Decision]]] = {
        "strong": [], "candidate": [], "low": [], "conflict": [], "special": [], "no_candidate": []
    }
    for listing, decision in records:
        if decision.bucket == "strong_single":
            buckets["strong"].append((listing, decision))
        elif decision.bucket == "candidate":
            buckets["candidate"].append((listing, decision))
        elif decision.bucket == "low_candidate":
            buckets["low"].append((listing, decision))
        elif decision.bucket == "conflict":
            buckets["conflict"].append((listing, decision))
        elif decision.bucket.startswith("special"):
            buckets["special"].append((listing, decision))
        else:
            buckets["no_candidate"].append((listing, decision))

    strong = _dataframe([_review_row(l, d) for l, d in buckets["strong"]])
    top = _dataframe([_review_row(l, d) for l, d in buckets["candidate"]])
    low = _dataframe([_review_row(l, d) for l, d in buckets["low"]])
    conflict = _dataframe([_review_row(l, d) for l, d in buckets["conflict"]])
    special_rows = [
        {
            "店铺": listing.get("shopId", ""),
            "站点": listing.get("marketplaceId", ""),
            "MSKU": listing.get("sku", ""),
            "ASIN": listing.get("asin", ""),
            "标题": listing.get("title", ""),
            "Amazon主图": listing.get("mainImage", ""),
            "对象类型": decision.object_type,
            "识别原因": " | ".join(decision.object_reasons),
            "Top1 SKU": decision.candidates[0].sku if decision.candidates else "",
            "Top1 名称": decision.candidates[0].name if decision.candidates else "",
            "Top1 分数": decision.candidates[0].score if decision.candidates else 0.0,
            "依据": " | ".join(decision.candidates[0].evidence) if decision.candidates else "",
            "冲突": decision.candidates[0].hard_conflicts if decision.candidates else 0,
            "建议": "组合/皮壳/海绵/未知对象需人工走对应工作流",
        }
        for listing, decision in buckets["special"]
    ]
    no_rows = [
        {
            "店铺": listing.get("shopId", ""),
            "站点": listing.get("marketplaceId", ""),
            "MSKU": listing.get("sku", ""),
            "ASIN": listing.get("asin", ""),
            "标题": listing.get("title", ""),
            "原因": "无可靠证据候选",
        }
        for listing, decision in buckets["no_candidate"]
    ]
    audit_rows = [
        {
            "店铺": listing.get("shopId", ""),
            "站点": listing.get("marketplaceId", ""),
            "MSKU": listing.get("sku", ""),
            "ASIN": listing.get("asin", ""),
            "对象类型": decision.object_type,
            "分流": decision.bucket,
            "证据": " | ".join(decision.evidence_sources),
            "候选SKU": " | ".join(row.sku for row in decision.candidates),
        }
        for listing, decision in records
    ]
    sheets = {
        "运行汇总": pd.DataFrame(summary.items(), columns=["指标", "值"]),
        "强证据建议": strong,
        "Top候选审核": top,
        "低证据候选": low,
        "冲突候选审核": conflict,
        "对象专项": pd.DataFrame(special_rows),
        "无候选": pd.DataFrame(no_rows),
        "证据审计": pd.DataFrame(audit_rows),
        "反馈说明": pd.DataFrame(
            [(value, "固定反馈枚举") for value in FEEDBACK_VALUES],
            columns=["可选结论", "说明"],
        ),
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)

    workbook = load_workbook(output)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in sheet.columns:
            values = [str(cell.value or "") for cell in list(column)[:100]]
            width = min(max(max(map(len, values), default=8) + 2, 10), 45)
            sheet.column_dimensions[column[0].column_letter].width = width

    for sheet_name in ("强证据建议", "Top候选审核", "低证据候选", "冲突候选审核"):
        review = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in review[1]}
        feedback_column = review.cell(1, headers["人工结论"]).column_letter
        validation = DataValidation(
            type="list", formula1='"' + ",".join(FEEDBACK_VALUES) + '"', allow_blank=True
        )
        review.add_data_validation(validation)
        validation.add(f"{feedback_column}2:{feedback_column}{max(review.max_row, 2)}")
        for label in ("Top1 分数", "Top2 分数", "Top3 分数"):
            column = review.cell(1, headers[label]).column_letter
            review.conditional_formatting.add(
                f"{column}2:{column}{max(review.max_row, 2)}",
                ColorScaleRule(start_type="min", start_color="F8696B", end_type="max", end_color="63BE7B"),
            )
    workbook.save(output)
