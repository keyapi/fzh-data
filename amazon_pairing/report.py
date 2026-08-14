from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


FEEDBACK_VALUES = (
    "确认Top1",
    "确认Top2",
    "确认Top3",
    "正确SKU另填",
    "无对应商品",
    "对象类型错误",
    "暂不确定",
)
REVIEW_COLUMNS = [
    "店铺", "站点", "MSKU", "ASIN", "标题", "Amazon主图", "对象类型", "识别属性", "警告",
    *[column for index in range(1, 4) for column in (f"Top{index} SKU", f"Top{index} 名称", f"Top{index} 分数", f"Top{index} 依据/冲突")],
    "人工结论", "正确SKU另填", "审核人", "审核时间",
]


@dataclass(frozen=True)
class ReviewRecord:
    shop: str
    marketplace: str
    msku: str
    asin: str
    title: str
    image_url: str
    route: str
    candidates: tuple[tuple[str, str, float, str], ...] = ()
    recognized_attributes: str = ""
    warnings: str = ""


def _review_rows(records: list[ReviewRecord]) -> list[dict]:
    rows = []
    for record in records:
        row = {
            "店铺": record.shop,
            "站点": record.marketplace,
            "MSKU": record.msku,
            "ASIN": record.asin,
            "标题": record.title,
            "Amazon主图": record.image_url,
            "对象类型": record.route,
            "识别属性": record.recognized_attributes,
            "警告": record.warnings,
        }
        for index in range(3):
            candidate = record.candidates[index] if index < len(record.candidates) else ("", "", 0.0, "")
            row[f"Top{index + 1} SKU"] = candidate[0]
            row[f"Top{index + 1} 名称"] = candidate[1]
            row[f"Top{index + 1} 分数"] = candidate[2]
            row[f"Top{index + 1} 依据/冲突"] = candidate[3]
        row["人工结论"] = ""
        row["正确SKU另填"] = ""
        row["审核人"] = ""
        row["审核时间"] = ""
        rows.append(row)
    return rows


def _records_frame(records) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["MSKU", "ASIN", "标题", "原因"])
    if isinstance(records[0], ReviewRecord):
        return pd.DataFrame(_review_rows(records))
    return pd.DataFrame(records)


def write_review_workbook(
    output: Path,
    records: list[ReviewRecord],
    summary: dict[str, object],
    special_records: list,
    no_candidate_records: list,
    quarantined_records: list,
    metrics: dict[str, float],
    exact_records: list[ReviewRecord] | None = None,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    sheets = {
        "运行汇总": pd.DataFrame(summary.items(), columns=["指标", "值"]),
        "高可信精确证据": pd.DataFrame(_review_rows(exact_records or []), columns=REVIEW_COLUMNS),
        "智能候选审核": pd.DataFrame(_review_rows(records), columns=REVIEW_COLUMNS),
        "特殊对象暂缓": _records_frame(special_records),
        "无可靠候选": _records_frame(no_candidate_records),
        "隔离历史数据": _records_frame(quarantined_records),
        "模型评估": pd.DataFrame(metrics.items(), columns=["指标", "值"]),
        "反馈说明": pd.DataFrame(
            [(value, "固定反馈枚举") for value in FEEDBACK_VALUES], columns=["可选结论", "说明"]
        ),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)

    workbook = load_workbook(output)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in sheet.columns:
            values = [str(cell.value or "") for cell in list(column)[:100]]
            width = min(max(max(map(len, values), default=8) + 2, 10), 45)
            sheet.column_dimensions[column[0].column_letter].width = width

    for sheet_name in ("高可信精确证据", "智能候选审核"):
        review = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in review[1]}
        feedback_column = review.cell(1, headers["人工结论"]).column_letter
        validation = DataValidation(
            type="list", formula1='"' + ",".join(FEEDBACK_VALUES) + '"', allow_blank=True
        )
        review.add_data_validation(validation)
        validation.add(f"{feedback_column}2:{feedback_column}{max(review.max_row, 2)}")
        for label in ("Top1 分数", "Top2 分数", "Top3 分数"):
            column = review.cell(1, headers[label]).column_letter
            review.conditional_formatting.add(
                f"{column}2:{column}{max(review.max_row, 2)}",
                ColorScaleRule(start_type="min", start_color="F8696B", end_type="max", end_color="63BE7B"),
            )
    workbook.save(output)
