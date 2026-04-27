# -*- coding: utf-8 -*-
"""
从 EN「产品BOM成本列表」xlsx 计算「绍兴发货成本」；用赛狐「商品导出」的 SKU 列作白名单，工作表
「商品」仅含两边交集（与赛狐导入模板一致，不可使用默认名 Sheet1）。赛狐侧「采购成本」不接受 0：缺数或计算结果为 0 时均导出为空白。有有效采购成本时同写「采购备注」
（前缀 EN绍兴发货成本- 加该行「绍兴发货方式」），避免被误认为完整采购成本。
若 BOM 未维护导致绍兴发货成本缺失，可按下划线分段：产品编号以「-」分节，对至少 4 节者取前 3
节为「品类-面料-尺寸」键，在表内用同键下首次出现的**非 0 有效**成本作借用；明细列「成本借用自」标记来源行。

默认 BOM：en_bom_cost_list 下最新 .xlsx；默认赛狐：edit_item/商品导出 x2215...xlsx；默认输出：item_cost_sx/out/。

源目录有多个 .xlsx 时取修改时间最新（排除 ~$ 锁文件）。

使用：
  python bom_cost_to_saihu_item_cost.py
  python bom_cost_to_saihu_item_cost.py --out-dir out
  python bom_cost_to_saihu_item_cost.py --saihu-commodities "path\\商品导出.xlsx"
  python bom_cost_to_saihu_item_cost.py --skip-saihu-match
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

_DIR = Path(__file__).resolve().parent
_ROOT = _DIR.parent
_DEFAULT_BOM_DIR = _ROOT / "en_bom_cost_list"
_DEFAULT_OUT_DIR = _DIR / "out"
_DEFAULT_SAIHU_EXPORT = _ROOT / "edit_item" / "商品导出 x2215 Commodities2026_04_23(1).xlsx"

COL_SKU = "产品编号"
COL_SAI = "赛狐存在"
COL_MODE = "绍兴发货方式"
COL_PKE = "皮壳成本"
COL_BCP = "绍兴包装半成品成本"
COL_CCP = "绍兴包装成品成本"
COL_TOT = "绍兴总成本"
COL_XS = "绍兴发货成本"
COL_BOR = "成本借用自(产品编号)"
COL_ISSUE = "问题说明"

OUT_DETAIL_COLS = [
    COL_SKU,
    COL_SAI,
    COL_MODE,
    COL_PKE,
    COL_BCP,
    COL_CCP,
    COL_TOT,
    COL_XS,
    COL_BOR,
    COL_ISSUE,
]

SAI_HU_SHEET = "商品"
# 赛狐「商品」表可导入列：*SKU、采购成本(CNY)、采购备注
COL_SAIHU_REMARK = "采购备注"
SAIHU_REMARK_PREFIX = "EN绍兴发货成本-"
DETAIL_SHEET = "BOM处理明细"
ISSUE_SHEET = "问题汇总"
SHEET_EN_ONLY = "对账_仅EN有"
SHEET_SAI_ONLY = "对账_仅赛狐有"


def _norm_str(x: Any) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    return s if s and s.lower() != "nan" else ""


def _to_float(x: Any) -> float | None:
    if x is None:
        return None
    if isinstance(x, (float, int)) and bool(pd.isna(x)):
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _round_cost(v: float | None) -> float | None:
    if v is None:
        return None
    return float(round(v, 8))


def _value_present(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, float) and (math.isnan(v) or (isinstance(v, float) and pd.isna(v))):
        return False
    return True


def _saihu_purchase_cost_export(v: Any) -> Any:
    """
    写入赛狐 Sheet「采购成本(CNY)」的最终值：缺数、NaN、或数值 0 均输出 None（导入后为空）。
    赛狐导入不接受 0 作为有效采购成本，与「缺失」同等处理。
    """
    if not _value_present(v):
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return v
    if f == 0.0:
        return None
    return v


def _sku_borrow_key(sku: str) -> str | None:
    """
    与「KS0001-QDKTR-140-RUSTYORANGE」类似：4 节及以上时取前 3 节为 品类-面料-尺寸 键；仅 3 节时整串为键。
    少于此无法按规则借用（品名中误用「-」时可能分节不准，仅作简单规则）。
    """
    s = _norm_str(sku)
    if not s:
        return None
    parts = s.split("-")
    n = len(parts)
    if n >= 4:
        return "-".join(parts[0:3])
    if n == 3:
        return s
    return None


def _lender_value_ok(v: Any) -> bool:
    """可作为出借方的绍兴发货成本：有数值且非 0。"""
    if not _value_present(v):
        return False
    try:
        return float(v) != 0.0
    except (TypeError, ValueError):
        return False


def _missing_shaoxing_cost(v: Any) -> bool:
    """未算出（None/NaN）。"""
    if v is None:
        return True
    if isinstance(v, float) and (math.isnan(v) or pd.isna(v) or (v != v)):
        return True
    return False


def _needs_borrow_cost(v: Any) -> bool:
    """
    是否应用同前缀借用：无有效成本（未算出 **或** 算出为 0）。
    皮壳等规则下部分颜色行为 0，但同「品类-面料-尺寸」下其它颜色有非 0 成本，应能借；与赛狐侧「0 当空」一致。
    """
    if _missing_shaoxing_cost(v):
        return True
    try:
        return float(v) == 0.0
    except (TypeError, ValueError):
        return False


def _apply_sku_borrow(detail: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    在已有「绍兴发货成本」列上，对「无有效成本（未算出或 0）」的行，用同 borrow_key 下**文档顺序中第一个**
    有非 0 有效成本的行来填充；仅一轮，只从初算有值行出借。
    返回 (新 detail, 借用行数)。
    """
    d = detail.copy()
    d[COL_BOR] = ""

    first: dict[str, tuple[str, float]] = {}
    for _, row in d.iterrows():
        sku = _norm_str(row[COL_SKU])
        k = _sku_borrow_key(sku)
        if not k or not _lender_value_ok(row[COL_XS]):
            continue
        if k not in first:
            rc = _round_cost(float(row[COL_XS]))
            if rc is not None:
                first[k] = (sku, float(rc))
    n_bor = 0
    for i in d.index:
        v = d.at[i, COL_XS]
        if not _needs_borrow_cost(v):
            continue
        sku = _norm_str(d.at[i, COL_SKU])
        k = _sku_borrow_key(sku)
        if not k or k not in first:
            continue
        lender_sku, cost = first[k]
        if lender_sku == sku:
            continue
        d.at[i, COL_XS] = _round_cost(cost)
        d.at[i, COL_BOR] = lender_sku
        n_bor += 1
    return d, n_bor


def _latest_xlsx(folder: Path) -> Path:
    if not folder.is_dir():
        raise FileNotFoundError(f"目录不存在: {folder}")
    cands = [
        p
        for p in folder.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]
    if not cands:
        raise FileNotFoundError(f"{folder} 下无可用 .xlsx（已排除 ~$ 锁文件）")
    return max(cands, key=lambda x: x.stat().st_mtime)


def _read_saihu_sku_set(path: Path) -> set[str]:
    if not path.is_file():
        raise FileNotFoundError(f"未找到赛狐商品导出: {path}")
    xls = pd.ExcelFile(path)
    sh = "商品" if "商品" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sh, header=0)
    if "SKU" not in df.columns:
        raise ValueError(f"赛狐导出缺列 'SKU'，工作表 {sh!r}，有: {list(df.columns)[:20]}")
    s = { _norm_str(x) for x in df["SKU"].values if _norm_str(x) }
    return s


def _compute_ship_cost(
    mode: str, pk: float | None, bcp: float | None, tot: float | None
) -> tuple[float | None, str]:
    m = _norm_str(mode)
    if not m:
        return None, "绍兴发货方式缺失，无法确定计价规则，绍兴发货成本留空"
    if m == "皮壳":
        if pk is None:
            return None, "皮壳：皮壳成本缺失"
        return pk, ""
    if m == "半成品":
        if pk is None or bcp is None:
            return None, "半成品：皮壳成本或绍兴包装半成品成本缺失"
        return pk + bcp, ""
    if m == "成品":
        if tot is None:
            return None, "成品：绍兴总成本缺失"
        return tot, ""
    return None, f"绍兴发货方式非预期值({m!r})，需为 皮壳/半成品/成品 之一"


@dataclass
class ProcessResult:
    """BOM 去重+计算，不含赛狐过滤。"""
    detail: pd.DataFrame
    issues: list[dict[str, Any]] = field(default_factory=list)


def _process_bom_dataframe(
    raw: pd.DataFrame, _file_label: str
) -> ProcessResult:
    need = {COL_SKU, COL_MODE, COL_PKE, COL_BCP, COL_CCP, COL_TOT}
    miss = [c for c in need if c not in raw.columns]
    if miss:
        raise ValueError(f"缺少列 {miss!r}，当前: {list(raw.columns)[:25]}…")

    issues: list[dict[str, Any]] = []
    d0 = len(raw)
    raw2 = raw[raw[COL_SKU].apply(lambda v: bool(_norm_str(v)))].copy()
    n_empty_sku = d0 - len(raw2)
    if n_empty_sku:
        issues.append(
            {
                "类型": "源表空产品编号行",
                "产品编号": "",
                "说明": f"已跳过 {n_empty_sku} 行无「产品编号」",
            }
        )

    dup = raw2[raw2.duplicated(subset=[COL_SKU], keep=False)]
    if not dup.empty:
        for sku, g in dup.groupby(COL_SKU, sort=False):
            idxs = list(g.index)
            if len(idxs) <= 1:
                continue
            last_i = idxs[-1]
            for i in idxs[:-1]:
                issues.append(
                    {
                        "类型": "产品编号重复(保留末行)",
                        "产品编号": str(sku),
                        "说明": f"多行同编号，保留行索引 {last_i}，丢弃本行 {i}",
                    }
                )
    n_before = len(raw2)
    df = raw2.drop_duplicates(subset=[COL_SKU], keep="last").copy()
    n_merged = n_before - len(df)
    if n_merged:
        issues.append(
            {
                "类型": "去重统计",
                "产品编号": "",
                "说明": f"按「{COL_SKU}」去重(保留末行)合并了 {n_merged} 行",
            }
        )

    xs: list[Any] = []
    iss_text: list[str] = []
    for _, row in df.iterrows():
        mode = _norm_str(row.get(COL_MODE))
        val_p = _to_float(row.get(COL_PKE))
        val_b = _to_float(row.get(COL_BCP))
        val_t = _to_float(row.get(COL_TOT))
        w, rmsg = _compute_ship_cost(mode, val_p, val_b, val_t)
        w = _round_cost(w)
        xs.append(w)
        iss_text.append(rmsg)
        if rmsg:
            issues.append(
                {
                    "类型": "绍兴发货成本未填",
                    "产品编号": _norm_str(row[COL_SKU]),
                    "说明": rmsg,
                }
            )

    df[COL_XS] = xs
    df[COL_ISSUE] = iss_text
    n_empty_cost = int(sum(1 for v in xs if v is None))
    issues.insert(
        0,
        {
            "类型": "汇总",
            "产品编号": "",
            "说明": (
                f"行数(去重后)={len(df)}；"
                f"绍兴发货成本可算出数值的={len(df) - n_empty_cost}；"
                f"因规则/缺数需留空={n_empty_cost}"
            ),
        },
    )

    _base = [
        COL_SKU,
        COL_MODE,
        COL_PKE,
        COL_BCP,
        COL_CCP,
        COL_TOT,
        COL_XS,
        COL_ISSUE,
    ]
    detail = df[_base].copy()
    return ProcessResult(detail=detail, issues=issues)


def _read_bom_excel(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    name = "Query Report" if "Query Report" in xls.sheet_names else xls.sheet_names[0]
    return pd.read_excel(path, sheet_name=name, header=0)


def _saihu_import_frame(
    detail: pd.DataFrame, saihu_set: set[str] | None
) -> pd.DataFrame:
    """列 *SKU、采购成本(CNY)、采购备注；仅赛狐中存在的行；缺数/0 用 None，写盘后再清 NaN。"""
    sk = detail[COL_SKU].map(_norm_str)
    if saihu_set is not None:
        m = sk.isin(saihu_set)
        sub = detail.loc[m]
    else:
        sub = detail
    n = len(sub)
    sku_col: list[str] = []
    cost_col: list[Any] = []
    remark_col: list[Any] = []
    for i in range(n):
        sku_col.append(_norm_str(sub[COL_SKU].iloc[i]))
        v = sub[COL_XS].iloc[i]
        exported = _saihu_purchase_cost_export(v)
        cost_col.append(exported)
        if exported is None:
            remark_col.append(None)
        else:
            mode = _norm_str(sub[COL_MODE].iloc[i]) if COL_MODE in sub.columns else ""
            remark_col.append(f"{SAIHU_REMARK_PREFIX}{mode}")
    return pd.DataFrame(
        {
            "*SKU": sku_col,
            "采购成本(CNY)": cost_col,
            COL_SAIHU_REMARK: remark_col,
        }
    )


def _detail_with_sai_col(
    detail: pd.DataFrame, saihu_set: set[str] | None, skip: bool
) -> pd.DataFrame:
    d = detail.copy()
    if skip or saihu_set is None:
        d[COL_SAI] = "（未做赛狐匹配）"
    else:
        d[COL_SAI] = d[COL_SKU].map(
            lambda x: "是" if _norm_str(x) in saihu_set else "否"
        )
    return d[[c for c in OUT_DETAIL_COLS if c in d.columns]].copy()


def _openpyxl_clear_column_nan(
    xlsx_path: Path, sheet: str, header_name: str, *, clear_zero: bool = False
) -> None:
    """将表头为 header_name 的列中 NaN 写为无值；可选将数值 0 也清空。"""
    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet]
    col_i: int | None = None
    for c in range(1, ws.max_column + 1):
        if _norm_str(ws.cell(1, c).value) == _norm_str(header_name):
            col_i = c
            break
    if col_i is None:
        wb.save(xlsx_path)
        wb.close()
        return
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_i).value
        if v is None:
            continue
        if isinstance(v, float) and (math.isnan(v) or (isinstance(v, float) and v != v)):
            ws.cell(r, col_i).value = None
            continue
        if clear_zero and isinstance(v, (int, float)) and not isinstance(v, bool):
            try:
                if float(v) == 0.0:
                    ws.cell(r, col_i).value = None
            except (TypeError, ValueError):
                pass
    wb.save(xlsx_path)
    wb.close()


def _write_result(
    out_path: Path,
    saihu: pd.DataFrame,
    detail: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        saihu.to_excel(w, sheet_name=SAI_HU_SHEET, index=False)
        detail.to_excel(w, sheet_name=DETAIL_SHEET, index=False)
    _openpyxl_clear_column_nan(
        out_path, SAI_HU_SHEET, "采购成本(CNY)", clear_zero=True
    )
    _openpyxl_clear_column_nan(out_path, DETAIL_SHEET, COL_XS)


def _write_issues(
    path: Path,
    issues: list[dict[str, Any]],
    source: Path,
    saihu_src: str,
    en_only: list[str],
    saihu_only: list[str],
) -> None:
    rows = [dict(x, **{"ERP_BOM": source.name, "赛狐源": saihu_src}) for x in issues]
    dfi = (
        pd.DataFrame(rows)
        if rows
        else pd.DataFrame(
            columns=["类型", "产品编号", "说明", "ERP_BOM", "赛狐源"]
        )
    )
    d_en = (
        pd.DataFrame({COL_SKU: sorted(en_only)})
        if en_only
        else pd.DataFrame({COL_SKU: []})
    )
    d_sh = (
        pd.DataFrame({"SKU": sorted(saihu_only)})
        if saihu_only
        else pd.DataFrame({"SKU": []})
    )
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        dfi.to_excel(w, sheet_name=ISSUE_SHEET, index=False)
        d_en.to_excel(w, sheet_name=SHEET_EN_ONLY, index=False)
        d_sh.to_excel(w, sheet_name=SHEET_SAI_ONLY, index=False)


def _print_summary(
    n_detail: int,
    n_bom_cost: int,
    n_sheet1: int,
    n_en_only: int,
    n_sai_only: int,
    out_main: Path,
    out_issue: Path,
    src: Path,
    n_borrow: int = 0,
) -> None:
    print("=" * 60)
    print("EN BOM 成本 -> 赛狐 采购成本 处理完成")
    print("=" * 60)
    print(f"ERP BOM(最新):  {src}")
    print(f"主结果:         {out_main}")
    print(f"问题/对账:      {out_issue}")
    print(
        f"BOM 去重行: {n_detail}；绍兴发货成本可填(非 0、含同前缀借用): {n_bom_cost}；"
        f"需留空采购成本(导入时): {n_detail - n_bom_cost}"
    )
    if n_borrow:
        print(f"其中同「品类-面料-尺寸」前缀借用成本行数: {n_borrow}")
    print(
        f"工作表「{SAI_HU_SHEET}」可导入行(赛狐有该 SKU): {n_sheet1}；"
        f"仅EN有: {n_en_only}；仅赛狐有: {n_sai_only}"
    )
    print("=" * 60)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="从 en_bom_cost_list 读最新 BOM 成本，按赛狐 SKU 交集输出采购成本"
    )
    ap.add_argument("--bom-dir", type=Path, default=_DEFAULT_BOM_DIR, help="BOM 成本 xlsx 所在目录")
    ap.add_argument("--out-dir", type=Path, default=_DEFAULT_OUT_DIR, help="输出目录(默认 item_cost_sx/out)")
    ap.add_argument(
        "--saihu-commodities",
        type=Path,
        default=_DEFAULT_SAIHU_EXPORT,
        help="赛狐「商品导出」xlsx 路径(读 SKU 列，工作表 商品)",
    )
    ap.add_argument(
        "--skip-saihu-match",
        action="store_true",
        help=f"不做赛狐交集，工作表「{SAI_HU_SHEET}」=全部 BOM 行（慎用）",
    )
    ap.add_argument("--source", type=Path, default=None, help="指定单个 BOM xlsx，默认取目录中最新")
    args = ap.parse_args()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir: Path = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    src = args.source
    if src is None:
        src = _latest_xlsx(args.bom_dir)
    elif not src.is_file():
        raise FileNotFoundError(f"未找到: {src}")

    raw = _read_bom_excel(src)
    pr = _process_bom_dataframe(raw, str(src))
    det, n_borrow = _apply_sku_borrow(pr.detail)
    if n_borrow and pr.issues and str(pr.issues[0].get("类型")) == "汇总":
        pr.issues[0]["说明"] = (
            str(pr.issues[0].get("说明", ""))
            + f"；同「品类-面料-尺寸」前缀借用成本行数={n_borrow}"
        )
    if n_borrow:
        pr.issues.append(
            {
                "类型": "同前缀成本借用",
                "产品编号": "",
                "说明": (
                    f"对绍兴发货成本缺失、或初算为 0 的行，按「-」分节(≥4 节取前 3 节为键)，"
                    f"用同键下首次出现的非 0 成本作参照，明细见「{COL_BOR}」"
                ),
            }
        )
    n_bom_filled = int(sum(1 for v in det[COL_XS] if _lender_value_ok(v)))

    saihu_set: set[str] | None = None
    saihu_path_str = "（--skip-saihu-match）"
    if not args.skip_saihu_match:
        saihu_set = _read_saihu_sku_set(args.saihu_commodities)
        saihu_path_str = args.saihu_commodities.name
    en_skus = { _norm_str(x) for x in det[COL_SKU].values if _norm_str(x) }
    if saihu_set is not None:
        en_only = sorted(en_skus - saihu_set)
        saihu_only = sorted(saihu_set - en_skus)
    else:
        en_only, saihu_only = [], []
    d_final = _detail_with_sai_col(det, saihu_set, args.skip_saihu_match)
    sai_df = _saihu_import_frame(det, saihu_set if not args.skip_saihu_match else None)
    n_inter = len(sai_df)
    if pr.issues and str(pr.issues[0].get("类型")) == "汇总":
        s0 = str(pr.issues[0].get("说明", ""))
        if "；赛狐可导入行" not in s0:
            pr.issues[0]["说明"] = (
                s0
                + f"；赛狐可导入行(工作表「{SAI_HU_SHEET}」)={n_inter}；"
                f"仅EN有={len(en_only)}；仅赛狐有={len(saihu_only)}"
            )

    out_main = out_dir / f"赛狐_采购成本导入_{ts}.xlsx"
    out_issue = out_dir / f"BOM成本处理_问题报告_{ts}.xlsx"
    _write_result(out_main, sai_df, d_final)
    _write_issues(
        out_issue,
        pr.issues,
        src,
        saihu_path_str,
        en_only,
        saihu_only,
    )

    _print_summary(
        len(d_final),
        n_bom_filled,
        n_inter,
        len(en_only),
        len(saihu_only),
        out_main,
        out_issue,
        src,
        n_borrow=n_borrow,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
