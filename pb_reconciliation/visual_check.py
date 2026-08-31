#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 渲染视觉自查 — 用项目视觉模型（qwen-vl-plus / OpenRouter VL）描述 sheet 渲染图。

用途：模型非多模态（如 deepseek-v4-flash）时，把 xlsx 指定 sheet 渲染成 PNG，
再用视觉模型描述背景色/字体/换行等，无需人工逐条描述。

用法：
    set DASHSCOPE_API_KEY=sk-...        # 或 AI_API_KEY（sk-or-v1- 前缀走 OpenRouter）
    python visual_check.py <xlsx> [sheet名] ["自定义提示"]

前置：openai 包（pip install openai）、LibreOffice（soffice）、PyMuPDF（pip install pymupdf）。
"""

import base64
import os
import shutil
import subprocess
import sys
import tempfile
from copy import copy
from pathlib import Path

import openpyxl

DEFAULT_PROMPT = (
    "Describe this Excel sheet image precisely: 1) background colors of header cells and "
    "key cells (yellow/red/green/purple), 2) whether any text is cut off or misaligned, "
    "3) any formula errors (#REF!/#NAME?). List per row/section."
)


def render_sheet_png(xlsx, sheet, tmpdir):
    """提取指定 sheet 到临时 xlsx（fit-to-page 单页），用 LibreOffice 转 PDF，PyMuPDF 转 PNG。"""
    wb = openpyxl.load_workbook(xlsx)
    src = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = src.title
    for row in src.iter_rows():
        for cell in row:
            if cell.value is not None:
                c = ws.cell(cell.row, cell.column, cell.value)
                c.font = copy(cell.font)
                c.fill = copy(cell.fill)
                c.border = copy(cell.border)
                c.number_format = cell.number_format
                c.alignment = copy(cell.alignment)
    for k, v in src.column_dimensions.items():
        if v.width:
            ws.column_dimensions[k].width = v.width
    for k, v in src.row_dimensions.items():
        if v.height:
            ws.row_dimensions[k].height = v.height
    ws.print_area = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    solo = Path(tmpdir) / "solo.xlsx"
    out.save(solo)

    pdf = Path(tmpdir) / "solo.pdf"
    soffice = (os.environ.get("SOFFICE")
               or (shutil.which("soffice") if shutil.which("soffice") else "")
               or r"C:\Program Files\LibreOffice\program\soffice.exe"
               or r"C:\Program Files (x86)\LibreOffice\program\soffice.exe")
    if not Path(soffice).exists():
        raise FileNotFoundError("未找到 LibreOffice soffice，请设置 SOFFICE 环境变量")
    subprocess.run([soffice, "--headless", "--convert-to", "pdf", "--outdir", str(tmpdir), str(solo)],
                   check=True, capture_output=True)
    import fitz
    doc = fitz.open(str(pdf))
    page = doc[0]
    pix = page.get_pixmap(dpi=150)
    png = Path(tmpdir) / "solo.png"
    pix.save(str(png))
    return png


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    xlsx = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    prompt = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_PROMPT

    key = os.environ.get("AI_API_KEY") or os.environ.get("DASHSCOPE_API_KEY")
    if not key:
        print("需要 AI_API_KEY 或 DASHSCOPE_API_KEY 环境变量")
        return 1

    from openai import OpenAI
    base = "https://openrouter.ai/api/v1" if key.startswith("sk-or-v1-") else "https://dashscope.aliyuncs.com/compatible-mode/v1"
    model = os.environ.get("AI_MODEL", "nvidia/nemotron-nano-12b-v2-vl:free" if key.startswith("sk-or-v1-") else "qwen-vl-plus")
    client = OpenAI(base_url=base, api_key=key)

    with tempfile.TemporaryDirectory() as td:
        wb = openpyxl.load_workbook(xlsx)
        sn = sheet if sheet and sheet in wb.sheetnames else wb.sheetnames[0]
        png = render_sheet_png(xlsx, sn, td)
        b64 = base64.b64encode(png.read_bytes()).decode()
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                {"type": "text", "text": prompt}]}],
            max_tokens=900)
        print(resp.choices[0].message.content)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
