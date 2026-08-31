#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""PB 佣金结算表（To Tracy Miller）生成。

从给财务对账表过滤账期（每月 19 号 - 下月 18 号）内的付款 + 按天截止的发票，
生成 TM 佣金表（全英文 Notes、5% 佣金）。可复用每月跑。

用法：
    python tm_commission.py --dry-run   # 只读+报告+校验，不写文件
    python tm_commission.py --write     # 生成 To Tracy Miller 目录下的账期文件

下月复用：改顶部 FINANCE_FILE / PERIODS / EXPECTED / PREV_SOURCE 后重跑。
"""

import datetime
import os
import sys
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill

# ================= 本月参数（下月复用只改这里） =================
FINANCE_FILE = r"D:\Work\美国\Tracy Miller\PB orders\payment advice\给财务\PB Remittance Advice Payment Date 20240430-20260813_差5单未付 20260814_171350.xlsx"
OUT_DIR = r"D:\Work\美国\Tracy Miller\PB orders\payment advice\To Tracy Miller"
# 账期列表：(start, end) 格式 YYYYMMDD。默认每月两个独立账期；
# 如需一次合并结算（如 2026-08 付 05/19-07/18 两期），可临时改为 [("20260519","20260718")]
PERIODS = [("20260519", "20260618"), ("20260619", "20260718")]
# 各账期预计付款总额（硬校验，来自财务确认）；如需一次合并结算可加 ("20260519","20260718") -> 23028.46
EXPECTED = {"20260519-20260618": 14185.71, "20260619-20260718": 8842.75}
# 上轮账期 TM 文件（供 P1 的"上轮未付本轮已付"）；未列出的账期自动衔接上一期的未付清单
PREV_SOURCE = {
    "20260519": r"D:\Work\美国\Tracy Miller\PB orders\payment advice\To Tracy Miller\PB Remittance Advice Payment Date 20260419-20260518.xlsx",
}
COMMISSION_RATE = 0.05
K2_NOTE = (
    "Pottery Barn (PB) Commission To Tracy Miller. \n"
    "1) PB payment date: 30 days after shipment. \n"
    "2) Settlement cycle: PB Payment Start Date 19th of the previous month - PB Payment End Date 18th of the current month, US central time. \n"
    "3) Commission payout time: 14 days after a Settlement cycle.\n"
    "4) Worksheet \"Invoice to PB\" downloaded from PB, for invoice and order item details.\n"
    "In column X \"Record Type\": \n"
    "Choose \"H\" (Header) for order level data e.g. Date, Invoice Total etc.;\n"
    "Choose \"D\" (Details) for order item level details, Invoice Total with \"Record Type\":\"D\" are duplicates and should not be counted. \n"
    "e.g. Cell:G2 uses formula =SUMIF('Invoice to PB'!X:X, \"H\", 'Invoice to PB'!CA:CA) to pick only \"Record Type\":\"H\" and accumulates their Invoice Total"
)
# ==============================================================

PB_SHEET = "PB Remittance Advice"
INV_SHEET = "Invoice to PB"
NOTES_SHEET = "Notes"

CA_COL = 79
CG_COL = 85
CH_COL = 86
X_COL = 24  # Record Type


def to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        return datetime.datetime.strptime(v.strip(), "%m/%d/%Y").date()
    return None


def copy_style(dst, src):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.number_format = src.number_format
    dst.alignment = copy(src.alignment)


def read_prev_unpaid(file):
    """从上一账期 TM 文件读取 'Unpaid in this period' 发票清单。"""
    wb = openpyxl.load_workbook(file, data_only=True)
    nws = wb[NOTES_SHEET]
    invs = {}
    in_section = False
    for r in range(1, nws.max_row + 1):
        v = nws.cell(r, 11).value
        if v == "Unpaid in this period":
            in_section = True
            continue
        if in_section and isinstance(v, str) and v.startswith("INV"):
            invs[str(v).strip()] = (nws.cell(r, 12).value, nws.cell(r, 13).value)
        elif in_section and v == "Total:":
            break
    return invs


def build_notes(nws, period_start, period_end, inv_start, inv_end, pay_start, pay_end,
                actual_pay_start, actual_pay_end, unpaid_last_paid, unpaid_this):
    """写 Notes sheet（结构/样式对照示例 20260419-20260518）。"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    f10 = Font(name="Arial", size=10)
    f10b = Font(name="Arial", size=10, bold=True)
    f11 = Font(name="Arial", size=11)
    f11b = Font(name="Arial", size=11, bold=True)
    fill_purple = PatternFill("solid", start_color="FFE5DFEC", end_color="FFE5DFEC")
    fill_yellow = PatternFill("solid", start_color="FFFFFF00", end_color="FFFFFF00")
    fill_green = PatternFill("solid", start_color="FF92D050", end_color="FF92D050")
    fill_red = PatternFill("solid", start_color="FFFF0000", end_color="FFFF0000")
    no_fill = PatternFill(fill_type=None)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    date_fmt = "m/d/yyyy;@"
    amt_fmt = "\\$#,##0.00;\\-\\$#,##0.00"
    wrap = Alignment(wrap_text=True, vertical="center")
    center = Alignment(wrap_text=True, horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # 列宽
    for col, w in {"A": 10.5, "F": 12.875, "G": 15.0, "I": 11.125, "J": 16.375,
                   "K": 67.875, "L": 12.875, "M": 13.375, "N": 15.625}.items():
        nws.column_dimensions[col].width = w
    # 行高
    nws.row_dimensions[1].height = 39
    nws.row_dimensions[2].height = 153.75
    nws.row_dimensions[3].height = 51
    nws.row_dimensions[4].height = 13.5

    def put(r, c, val, font=f11, fill=None, numfmt="General", align=None):
        cell = nws.cell(r, c)
        cell.value = val
        cell.font = copy(font)
        cell.fill = fill if fill else no_fill
        cell.border = border
        cell.number_format = numfmt if numfmt else "General"
        cell.alignment = align if align else (wrap if c == 11 else Alignment())
        return cell

    # Row 1 headers（A1-F1 浅紫底，居中）
    for c, h in [(1, "Invoice To PB Start Date"), (2, "Invoice To PB End Date"),
                 (3, "PB Invoice Start Date"), (4, "PB Invoice End Date"),
                 (5, "PB Payment Start Date"), (6, "PB Payment End Date"),
                 (7, "Invoice Amount"), (8, "Payment Amount"), (9, "Commission Rate"),
                 (10, "Commission Amount"), (11, "Notes")]:
        put(1, c, h, f10b, fill_purple if c <= 6 else None, align=center)

    # Row 2 dates + formulas（A2-F2 无背景色居中；G2 黄底、H2 红底，金额 $ 格式；I2 5%）
    def dt(v):
        return datetime.datetime(v.year, v.month, v.day)
    put(2, 1, dt(inv_start), f11, None, date_fmt, center)
    put(2, 2, dt(inv_end), f11, None, date_fmt, center)
    put(2, 3, dt(inv_start), f11, None, date_fmt, center)
    put(2, 4, dt(inv_end), f11, None, date_fmt, center)
    put(2, 5, dt(pay_start), f11, None, date_fmt, center)
    put(2, 6, dt(pay_end), f11, None, date_fmt, center)
    put(2, 7, f"=SUMIF('{INV_SHEET}'!X:X,\"H\",'{INV_SHEET}'!CA:CA)", f10, fill_yellow, amt_fmt, center)
    put(2, 8, f"=SUM('{PB_SHEET}'!I:I)", f10, fill_red, amt_fmt, center)
    put(2, 9, COMMISSION_RATE, f10, None, "0%", center)
    put(2, 10, "=H2*I2", f10, None, amt_fmt, center)
    put(2, 11, K2_NOTE, f10, None, None, wrap)

    # Row 3 actual dates（实际首末付款日，非账期边界）
    put(3, 5, f"Actual PB Payment Start Date: {actual_pay_start.month}/{actual_pay_start.day}/{actual_pay_start.year}", f10, None, None, wrap)
    put(3, 6, f"Actual PB Payment End Date: {actual_pay_end.month}/{actual_pay_end.day}/{actual_pay_end.year}", f10, None, None, wrap)

    # Unpaid in last period, paid in this period（绿底）
    r = 5
    nws.row_dimensions[r].height = 15
    put(r, 11, "Unpaid in last period, paid in this period", f11b)
    put(r, 12, "Invoice Date", f11b)
    put(r, 13, "Invoice Total", f11b)
    r += 1
    for inv, (d, amt) in sorted(unpaid_last_paid.items()):
        nws.row_dimensions[r].height = 14.25
        put(r, 11, inv, f11, fill_green)
        if d:
            put(r, 12, d, f11, fill_green)
        if amt is not None:
            put(r, 13, amt, f11, fill_green, amt_fmt)
        r += 1
    last_total_row = r
    nws.row_dimensions[r].height = 15
    put(last_total_row, 11, "Total:", f11b, fill_green)
    if unpaid_last_paid:
        put(last_total_row, 13, f"=SUM(M6:M{last_total_row - 1})", f11b, fill_green, amt_fmt)
    else:
        put(last_total_row, 13, 0, f11b, fill_green, amt_fmt)

    # Unpaid in this period（黄底）
    hu = last_total_row + 3  # 隔 2 行
    nws.row_dimensions[hu].height = 15
    nws.row_dimensions[hu + 1].height = 14.25 if unpaid_this else 15
    put(hu, 11, "Unpaid in this period", f11b)
    put(hu, 12, "Invoice Date", f11b)
    put(hu, 13, "Invoice Total", f11b)
    r = hu + 1
    for inv, (d, amt) in sorted(unpaid_this.items()):
        nws.row_dimensions[r].height = 14.25
        put(r, 11, inv, f11, fill_yellow)
        if d:
            put(r, 12, d, f11, fill_yellow)
        if amt is not None:
            put(r, 13, amt, f11, fill_yellow, amt_fmt)
        r += 1
    this_total_row = r
    nws.row_dimensions[r].height = 15.75
    put(this_total_row, 11, "Total:", f11b, fill_yellow)
    if unpaid_this:
        put(this_total_row, 13, f"=SUM(M{hu + 1}:M{this_total_row - 1})", f11b, fill_yellow, amt_fmt)
    else:
        put(this_total_row, 13, 0, f11b, fill_yellow, amt_fmt)

    # Difference
    put(this_total_row, 7, "Difference", f11b)
    put(this_total_row, 8, "=G2-H2", f11b, None, amt_fmt)
    return this_total_row


def build_tm_file(fin_wb_values, fin_wb_styles, period, prev_unpaid):
    start_s, end_s = period
    start = datetime.datetime.strptime(start_s, "%Y%m%d").date()
    end = datetime.datetime.strptime(end_s, "%Y%m%d").date()

    fp = fin_wb_values[PB_SHEET]
    fi = fin_wb_values[INV_SHEET]

    # 1) 过滤付款
    pay_rows = []
    paid_inv = set()
    for r in range(2, fp.max_row + 1):
        b = fp.cell(r, 2).value
        if b and start <= to_date(b) <= end:
            vals = [fp.cell(r, c).value for c in range(1, 11)]  # A..J
            pay_rows.append(vals)
            paid_inv.add(str(vals[2]).strip())
    pay_dates = sorted({to_date(p[1]) for p in pay_rows})

    # 2) 已结算（本账期开始前已付款）与已付集合
    settled_before = set()
    for r in range(2, fp.max_row + 1):
        b = fp.cell(r, 2).value
        c = fp.cell(r, 3).value
        if b and c and to_date(b) < start:
            settled_before.add(str(c).strip())

    # 3) 发票日范围（结转模型）：ledger = [min(上轮未付结转日, 本账期首个付款日), 本账期最后付款日]
    inv_date = {}
    for r in range(2, fi.max_row + 1):
        if fi.cell(r, X_COL).value == "H" and fi.cell(r, 1).value and fi.cell(r, 2).value:
            inv_date[str(fi.cell(r, 1).value).strip()] = to_date(fi.cell(r, 2).value)
    paid_days = {inv_date[i] for i in paid_inv if i in inv_date}
    carry_days = {inv_date[i] for i in prev_unpaid if i in inv_date}
    ledger_start = min(paid_days)
    if carry_days:
        ledger_start = min(ledger_start, min(carry_days))
    ledger_end = max(paid_days)
    inv_start, inv_end = ledger_start, ledger_end

    # 4) 纳入发票 = 范围内 H 行发票号 - 已结算；上轮未付结转必须全保留
    included_inv = {i for i in inv_date if ledger_start <= inv_date[i] <= ledger_end and i not in settled_before}
    included_inv |= {i for i in prev_unpaid if i in inv_date}  # 结转安全网

    inv_rows = []
    for r in range(2, fi.max_row + 1):
        a = fi.cell(r, 1).value
        if a and str(a).strip() in included_inv:
            vals = [fi.cell(r, c).value for c in range(1, 85)]  # A..CF
            inv_rows.append(vals)

    # 5) 未付清单
    unpaid_this = {}
    for inv in sorted(included_inv - paid_inv):
        d = amt = None
        for v in inv_rows:
            if str(v[0]).strip() == inv and v[X_COL - 1] == "H":
                d, amt = v[1], v[CA_COL - 1]
                break
        unpaid_this[inv] = (d, amt)
    unpaid_last_paid = {inv: info for inv, info in prev_unpaid.items() if inv in paid_inv}

    # ---- 构建 workbook（sheet 顺序：Notes / PB Remittance Advice / Invoice to PB） ----
    tpl_p = fin_wb_styles[PB_SHEET]
    tpl_i = fin_wb_styles[INV_SHEET]
    wb = openpyxl.Workbook()
    ws_n = wb.active
    ws_n.title = NOTES_SHEET
    ws_p = wb.create_sheet(PB_SHEET)
    ws_i = wb.create_sheet(INV_SHEET)

    # PB Remittance Advice: header + filtered rows
    for c in range(1, 12):
        copy_style(ws_p.cell(1, c), tpl_p.cell(1, c))
        ws_p.cell(1, c).value = tpl_p.cell(1, c).value
    for idx, vals in enumerate(pay_rows):
        r = 2 + idx
        for c in range(1, 11):
            cell = ws_p.cell(r, c)
            cell.value = vals[c - 1]
            copy_style(cell, tpl_p.cell(2, c))
        cell = ws_p.cell(r, 11)
        cell.value = f"=VLOOKUP(C{r},'{INV_SHEET}'!A:H,2,FALSE)"
        copy_style(cell, tpl_p.cell(2, 11))

    # Invoice to PB: header + filtered rows
    for c in range(1, 87):
        copy_style(ws_i.cell(1, c), tpl_i.cell(1, c))
        ws_i.cell(1, c).value = tpl_i.cell(1, c).value
    for idx, vals in enumerate(inv_rows):
        r = 2 + idx
        for c, v in enumerate(vals, start=1):
            if v not in (None, ""):
                cell = ws_i.cell(r, c)
                cell.value = v
                copy_style(cell, tpl_i.cell(2, c))
        cell = ws_i.cell(r, CG_COL)
        cell.value = f"=_xlfn.IFNA(VLOOKUP(A{r},'{PB_SHEET}'!C:I,7,FALSE),0)"
        copy_style(cell, tpl_i.cell(2, CG_COL))
        cell = ws_i.cell(r, CH_COL)
        cell.value = f"=CA{r}-CG{r}"
        copy_style(cell, tpl_i.cell(2, CH_COL))

    # 未付发票 H 头行黄底标记
    from openpyxl.styles import PatternFill as _PF
    fill_yellow = _PF("solid", start_color="FFFFFF00", end_color="FFFFFF00")
    FILL_COLS = [1, 2, 3, 5, 7, 8, 24, 25, 28, 31, 33, 34, 52, 53, 55, 56, 57, 58, 60, 61, 63, 64, 65, 66, 79, 85, 86]
    for inv in unpaid_this:
        for r in range(2, ws_i.max_row + 1):
            if str(ws_i.cell(r, 1).value).strip() == inv and ws_i.cell(r, X_COL).value == "H":
                for c in FILL_COLS:
                    ws_i.cell(r, c).fill = fill_yellow

    # 首行加筛选，Record Type(X列) 过滤为 "H"（隐藏 D 重复行，方便查 Invoice Total / Check Payment Amount）
    ws_i.auto_filter.ref = f"A1:CH{ws_i.max_row}"
    from openpyxl.worksheet.filters import FilterColumn, CustomFilters, CustomFilter
    _fc = FilterColumn(colId=X_COL - 1)
    _fc.customFilters = CustomFilters(customFilter=[CustomFilter(val="H")])
    ws_i.auto_filter.filterColumn.append(_fc)

    # Notes
    build_notes(ws_n, start, end, inv_start, inv_end, start, end,
                pay_dates[0], pay_dates[-1], unpaid_last_paid, unpaid_this)

    wb.calculation.fullCalcOnLoad = True
    return wb, pay_rows, unpaid_this, inv_start, inv_end


def main():
    mode = "--write" if "--write" in sys.argv else "--dry-run"
    fin_values = openpyxl.load_workbook(FINANCE_FILE, data_only=True)
    fin_styles = openpyxl.load_workbook(FINANCE_FILE, data_only=False)

    prev_unpaid = {}
    for i, period in enumerate(PERIODS):
        start_s = period[0]
        # 上轮未付：P1 用 PREV_SOURCE 文件；后续用上一期输出
        if start_s in PREV_SOURCE:
            prev_unpaid = read_prev_unpaid(PREV_SOURCE[start_s])
            print(f"[{period[0]}-{period[1]}] 上轮未付来自 {os.path.basename(PREV_SOURCE[start_s])}: {len(prev_unpaid)} 张")
        elif i > 0:
            prev_unpaid = prev_unpaid_this
            print(f"[{period[0]}-{period[1]}] 上轮未付衔接上一期: {len(prev_unpaid)} 张")

        wb, pay_rows, unpaid_this, inv_start, inv_end = build_tm_file(fin_values, fin_styles, period, prev_unpaid)
        pay_total = round(sum(p[8] for p in pay_rows), 2)
        key = f"{period[0]}-{period[1]}"
        print(f"[{key}] 付款 {len(pay_rows)} 行, 总额 {pay_total}")
        print(f"   发票范围 {inv_start}..{inv_end}，未付本账期 {len(unpaid_this)} 张")

        # 硬校验
        exp = EXPECTED.get(key)
        if exp is not None and abs(pay_total - exp) > 0.01:
            print(f"   !! 付款总额 {pay_total} 与预期 {exp} 不符，跳过")
            return 1

        if mode == "--write":
            out = os.path.join(OUT_DIR, f"PB Remittance Advice Payment Date {key}.xlsx")
            wb.save(out)
            print(f"   已保存: {out}")
        prev_unpaid_this = unpaid_this

    if mode == "--dry-run":
        print("\n[dry-run] 未写文件。确认后用 --write 输出。")
    else:
        print("\n完成。请在 Excel 打开确认公式已重算。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
