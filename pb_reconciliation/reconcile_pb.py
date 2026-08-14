#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""PB (Pottery Barn) 对账表月度更新。

从 PB 邮件付款批次 + 发票 CSV 更新给财务的对账表：
1) 追加付款到 "PB Remittance Advice" 表（校验不重不漏）
2) 追加发票到 "Invoice to PB" 表（截至首个 0 付款的文件夹，校验不重不漏）
3) 更新 Notes 汇总日期 + 重写"上轮未付本轮已付/本轮未付"区块 + 差额说明
4) 颜色标记：本轮未付发票黄底，之前未付本轮已付发票绿底
5) 双开票映射：批次里的废用发票号改写为 CSV 留用号，便于公式 VLOOKUP

用法：
    python reconcile_pb.py --dry-run   # 只读+校验+打印报告，不写文件
    python reconcile_pb.py --write     # 写入新文件（时间戳后缀，不覆盖源）

下月复用：改顶部常量（FINANCE_FILE / EMAIL_FILE / SCAN_FOLDERS / REMAP）后重跑即可。
"""

import argparse
import csv
import datetime
import glob
import os
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill

# ================= 本月参数（下月复用只改这里） =================
BASE_DIR = r"D:\Work\美国\Tracy Miller\PB orders"
FINANCE_FILE = os.path.join(
    BASE_DIR, r"payment advice\给财务\PB Remittance Advice Payment Date 20240430-20260813.xlsx"
)
EMAIL_FILE = os.path.join(
    BASE_DIR,
    r"payment advice\来自Email Payment Remittance Advice_PaymentDate 20260604-20260813_CheckDate 2026-08-14_14-17-30.xlsx",
)
# 待扫描发票文件夹：可混合"月份文件夹"(202605) 与根目录"每日文件夹"(20260803)
SCAN_FOLDERS = ["202605", "202606", "202607"]
# 双开票映射：批次里的发票号 -> CSV 留用的发票号
REMAP = {"INV0580626000011541": "INV0580626000011530"}
REMAP_NOTE = "PB重复 弃用1541 留用1530"
# 差额说明模板（{} 填本轮未付合计）；-195/-32.5 为历史多付常数，见 Notes 相关区块
DIFF_NOTE = "多付的 -195  -  多付的32.5 = -227.50 + 未付{}"
# ==============================================================

# 颜色标记：黄 = 本轮未付；绿 = 之前未付本轮已付（浅绿）
YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
GREEN = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
# 与现有黄色标记一致的填充列（H 头行信息 + 校验列）
FILL_COLS = [1, 2, 3, 5, 7, 8, 24, 25, 28, 31, 33, 34, 52, 53, 55, 56, 57, 58, 60, 61, 63, 64, 65, 66, 79, 85, 86]

PB_SHEET = "PB Remittance Advice"
INV_SHEET = "Invoice to PB"
NOTES_SHEET = "Notes"

EMAIL_COLS = 10  # 邮件 A..J 直接复制到 PB 表 A..J
CSV_MAX_COLS = 84  # CSV 列 i -> Excel 列 i+1（A..CF）
CA_COL = 79  # Invoice to PB: CA = Invoice Total
CG_COL = 85  # Invoice to PB: CG = Check Payment Amount（公式）
CH_COL = 86  # Invoice to PB: CH = Check If Same（公式）

# 现有表里按数值存储的列（CSV 读到的是文本，写入时转成数字，否则 SUMIF/CH 公式失效）
NUMERIC_COLS = {3, 6, 7, 12, 13, 15, 16, 22, 23, 28, 31, 57, 65, 79, 80, 81, 83, 84}


def to_number(val):
    """把纯数字字符串转成 int/float；无法转换返回原值。"""
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s == "":
        return val
    try:
        if s.replace(".", "", 1).isdigit():
            return float(s) if "." in s else int(s)
        return val
    except ValueError:
        return val


def normalize_inv(v):
    return str(v).strip() if v is not None else ""


def remap_inv(inv):
    return REMAP.get(inv, inv)


def load_email_batch():
    """读邮件付款批次，返回 [{col: value}]，C 列已做双开票映射。"""
    wb = openpyxl.load_workbook(EMAIL_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, EMAIL_COLS + 1)]
        if vals[0] is None and vals[2] is None:
            continue
        rows.append(vals)
    return rows


def read_invoice_csv(path):
    """读发票 CSV，返回数据行列表（跳过表头）。"""
    with open(path, encoding="utf-8-sig", errors="replace") as fh:
        rows = list(csv.reader(fh))
    return [r for r in rows[1:] if r and r[0].strip()]


def collect_day_folders(scan_folders):
    """扫描文件夹，返回按日期排序的 [(day_label, [csv_path,...])]。

    排除 NotUsed 子文件夹；按"日文件夹"分组（月文件夹下的 YYYYMMDD，或根目录的每日文件夹），
    使截止判断能精确到天。"""
    days = []
    for folder in scan_folders:
        root = folder if os.path.isabs(folder) else os.path.join(BASE_DIR, folder)
        if not os.path.isdir(root):
            continue
        csvs = [
            f
            for f in glob.glob(os.path.join(root, "**", "invoice*.csv"), recursive=True)
            if "NotUsed" not in f
        ]
        by_day = {}
        for f in csvs:
            day = os.path.basename(os.path.dirname(os.path.dirname(f)))  # <day>\invoice\*.csv
            by_day.setdefault(day, []).append(f)
        for day in sorted(by_day):
            days.append((day, sorted(by_day[day])))
    return days


def detect_cutoff(folders_with_csv, batch_set):
    """按日期序返回要纳入的文件夹；首个 0 付款的文件夹及其后全部停止。"""
    included = []
    for label, files in folders_with_csv:
        invs = set()
        for fp in files:
            for row in read_invoice_csv(fp):
                invs.add(row[0].strip())
        paid = invs & batch_set
        if paid:
            included.append((label, files))
        else:
            break  # 第一个 0 付款文件夹即停止
    return included


def build_report(batch_rows, include_folders, wb, invoice_rows, added_set, existing_pay_invs, old_sheet_set, green_invs, yellow_invs):
    """汇总报告 + 不重不漏校验。返回 (report_lines, errors)。"""
    lines = []
    errors = []

    batch_invs_raw = [normalize_inv(v[2]) for v in batch_rows]
    batch_invs = [remap_inv(i) for i in batch_invs_raw]
    batch_set = set(batch_invs)

    # 1) 付款不重：批次发票号不得已在现付款表
    dup_pay = batch_set & existing_pay_invs
    lines.append(f"[付款] 批次 {len(batch_set)} 张发票；与现付款表重叠 {len(dup_pay)} 张")
    if dup_pay:
        errors.append(f"付款重复：{sorted(dup_pay)[:10]}")

    # 2) 付款不漏：批次每张发票都必须在 Invoice to PB 命中
    missing = batch_set - (old_sheet_set | added_set)
    lines.append(f"[付款] 批次每张发票在 Invoice to PB 命中：缺少 {len(missing)} 张")
    if missing:
        errors.append(f"付款无对应发票：{sorted(missing)[:10]}")

    # 3) 发票不重：CSV 发票不得与表内重复
    dup_inv = added_set & old_sheet_set
    lines.append(f"[发票] 新增 {len(added_set)} 张；与表内重叠 {len(dup_inv)} 张")
    if dup_inv:
        errors.append(f"发票与表内重复：{sorted(dup_inv)[:10]}")

    # 4) 发票不重：同一发票出现在多个 CSV 文件才算重复（H 头行 + D 明细行属正常）
    inv_files = {}
    for _, files in include_folders:
        for fp in files:
            for row in read_invoice_csv(fp):
                inv = normalize_inv(row[0])
                inv_files.setdefault(inv, set()).add(os.path.basename(fp))
    dup_inside = {inv: sorted(fs) for inv, fs in inv_files.items() if len(fs) > 1}
    if dup_inside:
        errors.append(f"发票跨文件重复：{ {k: v for k, v in list(dup_inside.items())[:5]} }")

    # 5) 发票不漏：全文件夹纳入、数据行数统计
    total_files = sum(len(f) for _, f in include_folders)
    lines.append(f"[发票] 纳入文件夹 {len(include_folders)} 个，CSV 文件 {total_files} 个，数据行 {len(invoice_rows)} 行")
    lines.append(f"[发票] 新增后 Invoice to PB 预计总发票数 {len(old_sheet_set | added_set)}")

    # 6) 未付清单（新增发票中批次未命中的）+ 颜色
    lines.append(f"[对账] 本轮未付（黄底）{len(yellow_invs)} 张：{yellow_invs}")
    lines.append(f"[对账] 之前未付本轮已付（绿底）{len(green_invs)} 张")

    # 7) 双开票映射记录
    mapped = [(a, b) for a, b in zip(batch_invs_raw, batch_invs) if a != b]
    lines.append(f"[双开票] 映射 {len(mapped)} 行：{mapped}")
    return lines, errors


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="只读+校验+打印报告，不写文件")
    ap.add_argument("--write", action="store_true", help="写入新文件（时间戳后缀）")
    args = ap.parse_args()
    if args.dry_run == args.write:
        print("请指定 --dry-run 或 --write 之一")
        return 1

    # ---- 读数据 ----
    print(f"[1/5] 读付款批次: {os.path.basename(EMAIL_FILE)}")
    batch_rows = load_email_batch()
    batch_set = {remap_inv(normalize_inv(v[2])) for v in batch_rows if normalize_inv(v[2])}
    print(f"      批次 {len(batch_rows)} 行 / {len(batch_set)} 张发票")

    print("[2/5] 扫描发票 CSV 文件夹")
    folders_with_csv = collect_day_folders(SCAN_FOLDERS)
    for label, files in folders_with_csv:
        print(f"      {label}: {len(files)} 个 CSV")
    include_folders = detect_cutoff(folders_with_csv, batch_set)
    print(f"      自动截止：纳入 {[l for l, _ in include_folders]}；"
          f"{[l for l, _ in folders_with_csv[len(include_folders):]] or '（无后续）'} 停止")

    invoice_rows = []
    added_set = set()
    for label, files in include_folders:
        for fp in files:
            for row in read_invoice_csv(fp):
                invoice_rows.append(row)
                added_set.add(normalize_inv(row[0]))

    # ---- 读取基础文件 ----
    print(f"[3/5] 读取基础文件: {os.path.basename(FINANCE_FILE)}")
    wb = openpyxl.load_workbook(FINANCE_FILE, data_only=False)
    pws = wb[PB_SHEET]
    existing_pay_invs = set()
    for r in range(2, pws.max_row + 1):
        v = pws.cell(r, 3).value
        if v:
            existing_pay_invs.add(normalize_inv(v))
    iws_base = wb[INV_SHEET]
    old_sheet_set = set()
    for r in range(2, iws_base.max_row + 1):
        v = iws_base.cell(r, 1).value
        if v:
            old_sheet_set.add(normalize_inv(v))

    # 颜色集合：绿=之前未付本轮已付（批次 ∩ 旧表），黄=本轮未付（新加未付）
    green_invs = sorted(batch_set & old_sheet_set)
    yellow_invs = sorted(added_set - batch_set)

    # ---- 校验 + 报告 ----
    lines, errors = build_report(
        batch_rows, include_folders, wb, invoice_rows, added_set, existing_pay_invs,
        old_sheet_set, green_invs, yellow_invs,
    )
    print("[4/5] 对账报告:")
    for ln in lines:
        print("      " + ln)

    if errors:
        print("\n校验失败，不写入：")
        for e in errors:
            print("  - " + e)
        return 1
    print("      校验全部通过 ✔")

    if args.dry_run:
        print("\n[dry-run] 未写任何文件。确认无误后用 --write 输出。")
        return 0

    # ---- 写入 ----
    print("[5/5] 写入新文件 ...")
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.dirname(FINANCE_FILE)
    out_name = os.path.basename(FINANCE_FILE).replace(".xlsx", f"_{stamp}.xlsx")
    out_path = os.path.join(out_dir, out_name)

    # 追加付款
    pws = wb[PB_SHEET]
    tpl = [copy(pws.cell(pws.max_row, c)._style) for c in range(1, pws.max_column + 1)]
    pay_start = pws.max_row + 1
    for idx, vals in enumerate(batch_rows):
        r = pay_start + idx
        for c in range(1, EMAIL_COLS + 1):
            cell = pws.cell(r, c)
            cell.value = vals[c - 1]
        # 双开票映射 + 备注
        if normalize_inv(vals[2]) in REMAP:
            pws.cell(r, 3).value = remap_inv(normalize_inv(vals[2]))
            pws.cell(r, 12).value = REMAP_NOTE
        pws.cell(r, 11).value = f"=VLOOKUP(C{r},'{INV_SHEET}'!A:H,2,FALSE)"
        for c in range(1, pws.max_column + 1):
            pws.cell(r, c)._style = copy(tpl[c - 1])
    print(f"      付款追加 {len(batch_rows)} 行 (PB Remittance Advice R{pay_start}-{pay_start + len(batch_rows) - 1})")

    # 追加发票
    iws = wb[INV_SHEET]
    itpl = [copy(iws.cell(iws.max_row, c)._style) for c in range(1, iws.max_column + 1)]
    inv_start = iws.max_row + 1
    for idx, row in enumerate(invoice_rows):
        r = inv_start + idx
        for i, val in enumerate(row[:CSV_MAX_COLS]):
            if val not in (None, ""):
                col = i + 1
                iws.cell(r, col).value = to_number(val) if col in NUMERIC_COLS else val
        iws.cell(r, CG_COL).value = f"=_xlfn.IFNA(VLOOKUP(A{r},'{PB_SHEET}'!C:I,7,FALSE),0)"
        iws.cell(r, CH_COL).value = f"=CA{r}-CG{r}"
        for c in range(1, iws.max_column + 1):
            iws.cell(r, c)._style = copy(itpl[c - 1])
    print(f"      发票追加 {len(invoice_rows)} 行 (Invoice to PB R{inv_start}-{inv_start + len(invoice_rows) - 1})")

    # 颜色标记：黄=本轮未付，绿=之前未付本轮已付（只标 H 头行）
    def inv_h_rows(inv):
        return [r for r in range(2, iws.max_row + 1) if iws.cell(r, 1).value == inv and iws.cell(r, 24).value == "H"]

    for inv in green_invs:
        for r in inv_h_rows(inv):
            for c in FILL_COLS:
                iws.cell(r, c).fill = GREEN
    for inv in yellow_invs:
        for r in inv_h_rows(inv):
            for c in FILL_COLS:
                iws.cell(r, c).fill = YELLOW
    print(f"      颜色：黄(未付) {len(yellow_invs)} 张，绿(本轮已付) {len(green_invs)} 张")

    # Notes 汇总日期
    nws = wb[NOTES_SHEET]
    # B2 = 新加发票最大日期（CSV 文本 MM/DD/YYYY）
    max_inv_date = None
    for row in invoice_rows:
        d = row[1].strip() if len(row) > 1 else ""
        if d:
            try:
                dt = datetime.datetime.strptime(d, "%m/%d/%Y").date()
                if max_inv_date is None or dt > max_inv_date:
                    max_inv_date = dt
            except ValueError:
                pass
    # D2 = 批次最大发票日期（PB 侧，E 列 datetime），F2 = 批次最大付款日期（B 列）
    max_pb_inv = max_pay = None
    for v in batch_rows:
        b, e = v[1], v[4]
        if isinstance(b, datetime.datetime):
            max_pay = b.date() if max_pay is None or b.date() > max_pay else max_pay
        if isinstance(e, datetime.datetime):
            max_pb_inv = e.date() if max_pb_inv is None or e.date() > max_pb_inv else max_pb_inv
    # 与现有值取较大者
    for col, new in ((2, max_inv_date), (4, max_pb_inv), (6, max_pay)):
        cur = nws.cell(2, col).value
        if isinstance(cur, datetime.datetime) and new is not None:
            if cur.date() > new:
                new = cur.date()
        if new is not None:
            nws.cell(2, col).value = datetime.datetime(new.year, new.month, new.day)
            print(f"      Notes {nws.cell(2, col).coordinate} -> {new}")

    # ---- Notes 区块重写（R46-86）：上轮未付本轮已付 + 本轮未付 + 异常 + 差额 ----
    def inv_detail(inv):
        for r in range(2, iws.max_row + 1):
            if iws.cell(r, 1).value == inv and iws.cell(r, 24).value == "H":
                return iws.cell(r, 2).value, iws.cell(r, 79).value
        return None, None

    # 读取基础文件里的"异常"区块（历史数据，保留原样）
    abn_header = "异常 已加入Invoice to PB"
    abn_start = next((r for r in range(44, 87) if nws.cell(r, 11).value == abn_header), None)
    abn_rows = []
    if abn_start:
        r = abn_start + 1
        while r <= 86 and nws.cell(r, 11).value not in (None, ""):
            abn_rows.append((nws.cell(r, 11).value, nws.cell(r, 12).value, nws.cell(r, 13).value))
            r += 1

    # 捕获样式（头行 + 数据行），再清空
    hdr_k, hdr_l, hdr_m = nws["K46"], nws["L46"], nws["M46"]
    data_k = nws["K52"] if nws["K52"].value else hdr_k
    data_m = nws["M52"] if nws["M52"].value else hdr_m
    no_fill = PatternFill(fill_type=None)
    for r in range(47, 87):
        for c in range(1, 16):
            cell = nws.cell(r, c)
            cell.value = None
            cell.fill = no_fill

    def put(r, col, val, style):
        cell = nws.cell(r, col)
        cell.value = val
        cell.font = copy(style.font)
        cell.fill = copy(style.fill)
        cell.border = copy(style.border)
        cell.number_format = style.number_format
        cell.alignment = copy(style.alignment)
        return cell

    # 上轮未付 本轮已付
    r = 47
    for inv in green_invs:
        d, amt = inv_detail(inv)
        put(r, 11, inv, data_k)
        if d:
            put(r, 12, d, hdr_l)
        if amt is not None:
            put(r, 13, amt, data_m)
        r += 1
    paid_total_row = r
    put(paid_total_row, 11, "金额合计", data_k)
    put(paid_total_row, 13, f"=SUM(M47:M{paid_total_row - 1})", data_m)

    # 本轮未付
    hu = paid_total_row + 2
    put(hu, 11, "本轮未付账单号", hdr_k)
    put(hu, 12, "订单日期", hdr_l)
    put(hu, 13, "账单金额", hdr_m)
    r = hu + 1
    for inv in yellow_invs:
        d, amt = inv_detail(inv)
        put(r, 11, inv, data_k)
        if d:
            put(r, 12, d, hdr_l)
        if amt is not None:
            put(r, 13, amt, data_m)
        r += 1
    unpaid_total_row = r
    put(unpaid_total_row, 11, "金额合计", data_k)
    put(unpaid_total_row, 13, f"=SUM(M{hu + 1}:M{unpaid_total_row - 1})", data_m)

    # 异常（保留历史）
    he = unpaid_total_row + 2
    put(he, 11, abn_header, hdr_k)
    put(he, 12, "订单日期", hdr_l)
    put(he, 13, "账单金额", hdr_m)
    for i, (k, l, m) in enumerate(abn_rows):
        put(he + 1 + i, 11, k, data_k)
        if l:
            put(he + 1 + i, 12, l, hdr_l)
        if m is not None:
            put(he + 1 + i, 13, m, data_m)

    # 差额
    unpaid_total = round(sum((inv_detail(i)[1] or 0) for i in yellow_invs), 2)
    green_total = round(sum((inv_detail(i)[1] or 0) for i in green_invs), 2)
    put(86, 8, "=G2-H2", nws["H86"])
    put(86, 9, "差额", nws["I86"])
    put(86, 11, DIFF_NOTE.format(unpaid_total), nws["K86"])
    print(f"      Notes：绿(已付) {len(green_invs)} 张 合计 {green_total}，黄(未付) {len(yellow_invs)} 张 合计 {unpaid_total}")

    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)
    print(f"      已保存: {out_path}")
    print("\n完成。请在 Excel 打开确认公式已重算（K/CG/CH 列、Notes G2/H2 总额）。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
