"""
Amazon Ads Negative Keyword Bulksheet Generator.
Reads search term analysis and outputs an .xlsx file ready for upload to
Amazon Ads Console > Bulk Operations.

Format: Amazon bulksheet v2 (Product | Entity | Operation triplet).
Research: advertise/docs/research/amazon-bulk-negative-keyword-format.md

Usage: python -m advertise.generate_negatives [--level campaign|adgroup]
"""
import json, os, argparse, sys
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
from advertise import load_data

SCRIPT_DIR = os.path.dirname(__file__)
OUT_DIR = os.path.join(SCRIPT_DIR, "out")


def load_j(name):
    with open(os.path.join(OUT_DIR, name), encoding="utf-8") as f:
        return json.load(f)


def generate(level="campaign", account="BJRYECLTD-US", period="2026-06"):
    """Generate negative keyword bulksheet from search term analysis.

    Args:
        level: 'campaign' for campaign-level negatives (safer, broader),
               'adgroup' for ad group-level negatives (more targeted)
    """
    # Load analysis data
    search = load_j("search_term_analysis.json")
    cross = load_j("cross_analysis.json")

    # Load raw data to get campaign/ad group mappings for each search term
    reports = load_data()
    st_df = reports.get("search_term")
    if st_df is None:
        print("错误: 未找到搜索词报告数据")
        sys.exit(1)

    # Filter negative candidates (waste terms)
    negative_terms = search.get("negative_candidates", [])
    if not negative_terms:
        print("没有找到否定候选词")
        return None

    print(f"找到 {len(negative_terms)} 个否定候选词")

    # Build mapping: search_term → [(campaign_id, campaign_name, ad_group_id, ad_group_name)]
    df = st_df.copy()
    # Ensure we have the right columns
    needed = ["search_term", "campaign_id", "campaign_name", "ad_group_id", "ad_group_name"]
    available = [c for c in needed if c in df.columns]
    if "campaign_id" not in df.columns or "search_term" not in df.columns:
        print(f"警告: 缺少必要列 (有: {list(df.columns)[:10]}...). 将使用通用格式。")
        has_ids = False
    else:
        has_ids = True

    # Get valid campaign/group IDs from targeting data (more reliable for campaign mapping)
    targeting_df = reports.get("targeting")
    campaign_id_map = {}  # campaign_name → campaign_id
    if targeting_df is not None and "campaign_id" in targeting_df.columns:
        for _, row in targeting_df.drop_duplicates(subset=["campaign_name"]).iterrows():
            campaign_id_map[str(row["campaign_name"])] = str(row["campaign_id"])

    # Also from search term data
    if has_ids:
        for _, row in df.drop_duplicates(subset=["search_term", "campaign_name"]).iterrows():
            cname = str(row["campaign_name"])
            if cname not in campaign_id_map:
                campaign_id_map[cname] = str(row["campaign_id"])

    # ── Build bulksheet ──────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Negative Keywords"

    # Header row
    headers = ["Product", "Entity", "Operation",
               "Campaign ID", "Campaign Name",
               "Ad Group ID", "Ad Group Name",
               "Keyword ID", "Keyword Text", "Match Type", "State"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Entity and state based on level
    if level == "campaign":
        entity = "Campaign Negative Keyword"
        default_state = "enabled"
    else:
        entity = "Negative Keyword"
        default_state = "enabled"

    # ── Match type logic ──────────────────────────────────────
    # Different negation strategies based on WHY the term was flagged:
    #
    #   不相关词 (irrelevant): "cheap", "free shipping", "used" — categorically junk
    #     → phrase match: block ALL searches containing this phrase
    #
    #   品牌词 (brand): competitor brand names — do NOT negate
    #     → WARNING only: competitor conquesting may be intentional strategy
    #
    #   品类词/长尾词/其他: category terms with high clicks but 0 orders
    #     → exact match (conservative): might convert with different bid/placement
    #
    #   竞品词: competitor product searches
    #     → exact match: block this specific term only

    def get_match_type(term_data):
        cat = (term_data.get("term_category") or "").strip()
        if cat == "不相关词":
            return "phrase", "irrelevant"
        elif cat in ("品类词", "长尾词", "竞品词"):
            return "exact", "low_performance"
        elif cat == "品牌词":
            return None, "brand_protection"  # DO NOT NEGATE
        else:
            return "exact", "low_performance"

    row = 2
    written = 0
    skipped_no_campaign = 0
    skipped_brand_protection = 0

    for term_data in negative_terms:
        search_term = str(term_data.get("search_term", "")).strip()
        if not search_term or len(search_term) > 100:
            continue

        spend = term_data.get("spend", 0) or 0
        clicks = term_data.get("clicks", 0) or 0

        # Determine match type and reason for negation
        match_type, neg_reason = get_match_type(term_data)

        # Brand terms: DO NOT NEGATE, flag for manual review
        if match_type is None:
            skipped_brand_protection += 1
            print(f"  [保护] 品牌词不否定: '{search_term}' — ${spend:.2f}, {clicks} 点击, "
                  f"可能是有意竞品投放，请人工审核")
            continue

        # Find which campaigns this term appeared in
        if has_ids:
            term_rows = df[df["search_term"] == search_term]
            campaigns = term_rows[["campaign_id", "campaign_name"]].drop_duplicates()
        else:
            campaigns = None

        if campaigns is not None and len(campaigns) > 0:
            for _, crow in campaigns.iterrows():
                cid = str(crow["campaign_id"])
                cname = str(crow["campaign_name"])
                ws.cell(row=row, column=1, value="Sponsored Products")
                ws.cell(row=row, column=2, value=entity)
                ws.cell(row=row, column=3, value="create")
                ws.cell(row=row, column=4, value=cid)
                ws.cell(row=row, column=5, value=cname[:150])
                if level == "adgroup":
                    # Add ad group details if available
                    ag_rows = term_rows[term_rows["campaign_id"] == crow["campaign_id"]]
                    if "ad_group_id" in ag_rows.columns:
                        ag = ag_rows[["ad_group_id", "ad_group_name"]].drop_duplicates().iloc[0]
                        ws.cell(row=row, column=6, value=str(ag["ad_group_id"]))
                        ws.cell(row=row, column=7, value=str(ag["ad_group_name"])[:150])
                ws.cell(row=row, column=9, value=search_term)
                ws.cell(row=row, column=10, value=match_type)
                ws.cell(row=row, column=11, value=default_state)

                # Add context about why this term was flagged
                cat_label = (term_data.get("term_category") or "未知")
                ws.cell(row=row, column=9).comment = openpyxl.comments.Comment(
                    f"${spend:.2f} 花费, {clicks} 点击, 0 订单 | "
                    f"分类: {cat_label} | 策略: {neg_reason} ({match_type} match) | "
                    f"自动识别于 {period}",
                    "auto-generator")

                row += 1
                written += 1
        else:
            # No campaign ID available — add as a note row for manual review
            skipped_no_campaign += 1

    if written == 0:
        print("警告: 没有生成任何否定词行")
        return None

    # Auto-width
    for c in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        max_len = 10
        for r in range(1, row):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Save
    fname = f"{account}-否定词bulksheet-{period}.xlsx"
    out = os.path.join(OUT_DIR, fname)
    wb.save(out)
    print(f"[OK] 否定词 bulksheet: {out}")
    print(f"  级别: {level}-level")
    print(f"  候选词总数: {len(negative_terms)}")
    print(f"  已生成否定: {written} 行")
    print(f"  品牌保护(不否定): {skipped_brand_protection} 个 — 需人工审核")
    print(f"  跳过(无活动ID): {skipped_no_campaign}")
    if skipped_brand_protection > 0:
        print(f"\n  ⚠ 有 {skipped_brand_protection} 个品牌词被保护未加入否定 — "
              f"请确认竞品投放策略后手动处理")
    print(f"\n  下一步: 打开 Amazon Ads Console → Bulk Operations → Upload spreadsheet → 选择此文件")
    return out


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--level", default="campaign", choices=["campaign", "adgroup"],
                        help="Negation level: campaign (default) or adgroup")
    parser.add_argument("--account", default="BJRYECLTD-US")
    parser.add_argument("--period", default="2026-06")
    args = parser.parse_args()
    generate(args.level, args.account, args.period)
