"""
Full report builder — reads all 7 analysis JSONs, produces a comprehensive Excel report.
Usage: python -m advertise.build_full_report [--account bjryecltd-us] [--period 2026-06]
"""
import json, os, argparse, sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(__file__)
OUT_DIR = os.path.join(SCRIPT_DIR, "out")

# ── Styles ───────────────────────────────────────────────────────────────
HDR_FONT = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
SUB_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="2F5496")
GOOD = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BORDER = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
NUM_FONT = Font(name="Consolas", size=10)
CN_FONT = Font(name="Microsoft YaHei", size=10)


def load_json(filename):
    with open(os.path.join(OUT_DIR, filename), encoding="utf-8") as f:
        return json.load(f)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def style_data(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CN_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")


def auto_width(ws, ncols, max_width=40):
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            for val in row:
                if val:
                    max_len = max(max_len, min(len(str(val)), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def add_title(ws, title, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")
    return row + 2


def fmt_pct(val):
    if val is None: return "-"
    return f"{float(val)*100:.1f}%"


def fmt_usd(val):
    if val is None: return "-"
    return f"${float(val):,.2f}"


def fmt_num(val, decimals=0):
    if val is None: return "-"
    return f"{float(val):,.{decimals}f}"


# ═══════════════════════════════════════════════════════════════════════════
# Main builder
# ═══════════════════════════════════════════════════════════════════════════

def build_report(account="BJRYECLTD-US", period="2026-06"):
    out = os.path.join(OUT_DIR, f"{account}-广告分析报告-{period}.xlsx")
    wb = openpyxl.Workbook()

    # Load data
    campaign = load_json("campaign_analysis.json")
    targeting = load_json("targeting_analysis.json")
    search = load_json("search_term_analysis.json")
    placement = load_json("placement_analysis.json")
    ad_group = load_json("ad_group_analysis.json")
    ad_product = load_json("advertised_product_analysis.json")
    purchased = load_json("purchased_item_analysis.json")
    cross = load_json("cross_analysis.json")

    ws_names = ["总览", "跨报告集成", "广告活动", "投放表现", "搜索词", "广告位",
                "广告组结构", "ASIN效率", "品牌光环", "行动建议"]
    for i, name in enumerate(ws_names):
        if i == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)

    # ── Sheet 1: Overview ────────────────────────────────────
    ws = wb["总览"]
    add_title(ws, f"Amazon 广告分析报告 — {account} — {period}", 3)

    # Key metrics card
    cs = campaign["summary"]
    ap = ad_product["summary"]
    pi = purchased["summary"]
    sr = search["summary"]

    cards = [
        ("核心指标", 2, [
            ("总花费", fmt_usd(ap["total_spend"])),
            ("直接销售额", fmt_usd(ap["total_sales"])),
            ("光环销售额", fmt_usd(pi["total_purchased_sales"])),
            ("总销售额(含光环)", fmt_usd(ap["total_sales"] + pi["total_purchased_sales"])),
            ("直接ACOS", fmt_pct(cs["overall_acos"])),
            ("混合ACOS(含光环)", fmt_pct(ap["blended_acos_with_halo"])),
            ("ROAS", fmt_num(cs["overall_roas"], 2)),
            ("总订单", fmt_num(cs["total_orders"])),
        ]),
        ("搜索词概览", 4, [
            ("独立搜索词", fmt_num(sr["unique_search_terms"])),
            ("Harvest 词", fmt_num(len(search.get("harvest_keywords", [])))),
            ("否定候选词", fmt_num(len(search.get("negative_candidates", [])))),
            ("观察中词", fmt_num(len(search.get("monitor_list", [])))),
            ("总花费", fmt_usd(sr.get("total_spend", 0))),
            ("总销售额", fmt_usd(sr.get("total_sales", 0))),
        ]),
        ("结构概览", 6, [
            ("广告活动数", fmt_num(cs["campaign_count"])),
            ("广告组数", fmt_num(ad_group["summary"]["group_count"])),
            ("推广ASIN数", fmt_num(ap["asin_count"])),
            ("Gateway ASIN", fmt_num(pi["gateway_candidates"])),
            ("结构问题", fmt_num(ad_group["summary"]["structural_issues"])),
        ]),
    ]

    for label, col, items in cards:
        r = 3
        ws.cell(row=r, column=col, value=label).font = SUB_FONT
        r += 1
        for k, v in items:
            ws.cell(row=r, column=col, value=k).font = CN_FONT
            ws.cell(row=r, column=col + 1, value=v).font = NUM_FONT
            r += 1

    # Placement summary
    r = 3
    ws.cell(row=r, column=8, value="广告位效率").font = SUB_FONT
    r += 1
    for p in placement["placements"]:
        ws.cell(row=r, column=8, value=p["placement"]).font = CN_FONT
        ws.cell(row=r, column=9, value=fmt_usd(p["spend"])).font = NUM_FONT
        ws.cell(row=r, column=10, value=fmt_usd(p["sales"])).font = NUM_FONT
        ws.cell(row=r, column=11, value=fmt_pct(p["acos"])).font = NUM_FONT
        acos_val = p.get("acos")
        if acos_val is not None and acos_val < 0.25:
            for c in range(8, 12):
                ws.cell(row=r, column=c).fill = GOOD
        elif acos_val is not None and acos_val > 0.50:
            for c in range(8, 12):
                ws.cell(row=r, column=c).fill = BAD
        r += 1

    auto_width(ws, 12)

    # ── Sheet 2: Cross-Report Integration ─────────────────────
    ws = wb["跨报告集成"]
    cs2 = cross["summary"]
    h2 = cs2["account_health"]
    add_title(ws, f"跨报告集成分析 — 健康度 {h2['score']}/100 ({h2['grade']}级)", 6)

    # Account health card
    ws.cell(row=3, column=1, value="账户健康度").font = SUB_FONT
    ws.cell(row=4, column=1, value=f"评分: {h2['score']}/100 — {h2['grade']}级").font = TITLE_FONT
    r = 5
    ws.cell(row=r, column=1, value="优势:").font = SUB_FONT
    for s in h2["key_strengths"]:
        r += 1; ws.cell(row=r, column=1, value=f"  + {s}").font = CN_FONT
        ws.cell(row=r, column=1).fill = GOOD
    r += 1
    ws.cell(row=r, column=1, value="劣势:").font = SUB_FONT
    for w in h2["key_weaknesses"]:
        r += 1; ws.cell(row=r, column=1, value=f"  - {w}").font = CN_FONT
        ws.cell(row=r, column=1).fill = BAD

    # Blended ACOS table
    r = 5
    ws.cell(row=r, column=3, value="正确ACOS (含光环)").font = SUB_FONT
    bh = ["活动", "花费", "直接销售", "光环销售", "总销售", "直接ACOS", "混合ACOS", "光环增幅"]
    for c, h in enumerate(bh, 3):
        ws.cell(row=6, column=c, value=h)
    style_header(ws, 6, len(bh) + 2)
    for i, bc in enumerate(cross["blended_campaign_acos"][:20], 7):
        ws.cell(row=i, column=3, value=str(bc.get("campaign_name", ""))[:35])
        ws.cell(row=i, column=4, value=fmt_usd(bc.get("spend")))
        ws.cell(row=i, column=5, value=fmt_usd(bc.get("direct_sales")))
        ws.cell(row=i, column=6, value=fmt_usd(bc.get("other_sku_sales")))
        ws.cell(row=i, column=7, value=fmt_usd(bc.get("total_sales")))
        ws.cell(row=i, column=8, value=fmt_pct(bc.get("direct_acos")))
        ws.cell(row=i, column=9, value=fmt_pct(bc.get("blended_acos")))
        ws.cell(row=i, column=10, value=f"{bc.get('halo_boost_pct',0):.0f}%")
        if bc.get("blended_acos") is not None and bc["blended_acos"] < 0.25:
            for c2 in range(3, 11):
                ws.cell(row=i, column=c2).fill = GOOD
    style_data(ws, 7, 6 + min(len(cross["blended_campaign_acos"]), 20), 10)
    auto_width(ws, 11)

    # Gateway ASIN final
    rg = 28
    ws.cell(row=rg, column=1, value="Gateway ASIN 最终判定").font = SUB_FONT
    rg += 1
    for a in cross["gateway_asin_final"]:
        if a["is_gateway"]:
            ws.cell(row=rg, column=1, value=f"{a['asin']} {a['sku'][:30]}: {a['action']}").font = CN_FONT
            ws.cell(row=rg, column=2, value=a["gateway_reason"]).font = CN_FONT
            ws.cell(row=rg, column=1).fill = GOOD
            rg += 1

    # ── Sheet 3: Campaign Ranking ─────────────────────────────
    ws = wb["广告活动"]
    rank = campaign["ranking"]
    headers = ["活动名称", "状态", "花费", "销售额", "ACOS", "ROAS", "订单", "点击", "展示", "CTR", "CPC", "标记"]
    keys = ["campaign_name", "status", "spend", "sales", "acos", "roas", "orders", "clicks", "impressions", "ctr", "cpc", "flag"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, r in enumerate(rank, 2):
        for j, k in enumerate(keys, 1):
            val = r.get(k, "")
            if k in ("acos", "ctr"):
                val = fmt_pct(val)
            elif k in ("spend", "sales", "cpc"):
                val = fmt_usd(val) if k != "cpc" else f"${float(r.get(k,0) or 0):.2f}"
            else:
                val = str(val) if val is not None else ""
            ws.cell(row=i, column=j, value=val)
        flag = r.get("flag", "")
        if "高风险" in str(flag) or "问题" in str(flag):
            for c in range(1, len(headers) + 1):
                ws.cell(row=i, column=c).fill = BAD
        elif "优胜" in str(flag):
            for c in range(1, len(headers) + 1):
                ws.cell(row=i, column=c).fill = GOOD
    style_data(ws, 2, len(rank) + 1, len(headers))
    auto_width(ws, len(headers))

    # ── Sheet 3: Targeting ───────────────────────────────────
    ws = wb["投放表现"]
    ws.cell(row=1, column=1, value="匹配类型").font = SUB_FONT
    mt_headers = ["类型", "花费", "销售额", "ACOS", "ROAS", "CTR", "CVR", "CPC", "订单", "点击", "展示"]
    mt_keys = ["match_type", "spend", "sales", "acos", "roas", "ctr", "cvr", "cpc", "orders", "clicks", "impressions"]
    for c, h in enumerate(mt_headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(mt_headers))
    for i, r in enumerate(targeting["match_type"], 3):
        for j, k in enumerate(mt_keys, 1):
            val = r.get(k, "")
            if k in ("acos", "ctr", "cvr"):
                val = fmt_pct(val)
            elif k in ("spend", "sales", "cpc"):
                val = fmt_usd(val) if k != "cpc" else f"${float(r.get(k,0) or 0):.2f}"
            ws.cell(row=i, column=j, value=str(val))

    # Halo summary
    h = targeting["halo_effect"]
    halo_text = (f"光环效应: SameSKU ${h['same_sku_sales']:,.2f} + "
                 f"OtherSKU ${h['other_sku_sales']:,.2f} = "
                 f"${h['total_attributed']:,.2f} (光环比 {h['halo_ratio']:.2f}x)")
    ws.cell(row=len(targeting["match_type"]) + 5, column=1, value=halo_text).font = SUB_FONT

    style_data(ws, 3, len(targeting["match_type"]) + 2, len(mt_headers))
    auto_width(ws, len(mt_headers))

    # ── Sheet 4: Search Terms ─────────────────────────────────
    ws = wb["搜索词"]
    sb = search["summary"]
    hc = len(search.get("harvest_keywords", []))
    nc = len(search.get("negative_candidates", []))
    mc = len(search.get("monitor_list", [])) if search.get("monitor_list") else 0
    pc = len(search.get("protect_list", [])) if search.get("protect_list") else 0
    ic = sb.get("ignore_count", 0) if "ignore_count" in sb else (sb["unique_search_terms"] - hc - nc - mc - pc)
    ws.cell(row=1, column=1, value=f"搜索词分类: Harvest {hc} | "
            f"Negate {nc} | Monitor {mc} | "
            f"Ignore {ic} | Protect {pc}").font = SUB_FONT

    bucket_map = {
        "Harvest": search.get("harvest_keywords", []),
        "Negate": search.get("negative_candidates", []),
        "Monitor": search.get("monitor_list", []),
    }
    for idx, (bucket_name, terms) in enumerate(bucket_map.items()):
        if not terms:
            continue
        col_start = 1 + idx * 4
        ws.cell(row=3, column=col_start, value=f"--- {bucket_name} ({len(terms)}词) ---").font = SUB_FONT
        hdrs = ["搜索词", "花费", "销售额", "ACOS"]
        for c, h in enumerate(hdrs):
            ws.cell(row=4, column=col_start + c, value=h)
            ws.cell(row=4, column=col_start + c).font = HDR_FONT
            ws.cell(row=4, column=col_start + c).fill = HDR_FILL
        for i, t in enumerate(terms[:30], 5):
            ws.cell(row=i, column=col_start, value=str(t.get("search_term", ""))[:50])
            ws.cell(row=i, column=col_start + 1, value=fmt_usd(t.get("spend")))
            ws.cell(row=i, column=col_start + 2, value=fmt_usd(t.get("sales")))
            ws.cell(row=i, column=col_start + 3, value=fmt_pct(t.get("acos")))

    # Category distribution
    if "category_distribution" in search and search["category_distribution"]:
        cat_dist = search["category_distribution"]
        cat_start = 13
        ws.cell(row=cat_start, column=1, value="搜索词分类分布").font = SUB_FONT
        if isinstance(cat_dist, dict):
            for i, (cat, count) in enumerate(cat_dist.items(), cat_start + 1):
                ws.cell(row=i, column=1, value=str(cat)).font = CN_FONT
                ws.cell(row=i, column=2, value=count).font = NUM_FONT
        elif isinstance(cat_dist, list):
            for i, item in enumerate(cat_dist, cat_start + 1):
                if isinstance(item, dict):
                    ws.cell(row=i, column=1, value=str(item.get("category", ""))).font = CN_FONT
                    ws.cell(row=i, column=2, value=item.get("count", "")).font = NUM_FONT
                else:
                    ws.cell(row=i, column=1, value=str(item)).font = CN_FONT

    auto_width(ws, 15)

    # ── Sheet 5: Placement ───────────────────────────────────
    ws = wb["广告位"]
    add_title(ws, "广告位效率对比", 10)
    ph = ["广告位", "花费", "销售额", "ACOS", "ROAS", "CTR", "CVR", "CPC", "订单", "建议"]
    for c, h in enumerate(ph, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(ph))
    for i, p in enumerate(placement["placements"], 4):
        vals = [p["placement"], fmt_usd(p["spend"]), fmt_usd(p["sales"]),
                fmt_pct(p["acos"]), fmt_num(p.get("roas"), 2), fmt_pct(p.get("ctr")),
                fmt_pct(p.get("cvr")), f"${p.get('cpc', 0) or 0:.2f}",
                fmt_num(p.get("orders"))]
        for j, v in enumerate(vals, 1):
            ws.cell(row=i, column=j, value=v)
        if p.get("acos") is not None and p["acos"] < 0.25:
            for c2 in range(1, len(ph) + 1):
                ws.cell(row=i, column=c2).fill = GOOD
        elif p.get("acos") is not None and p["acos"] > 0.50:
            for c2 in range(1, len(ph) + 1):
                ws.cell(row=i, column=c2).fill = BAD

    # Recommendations
    r2 = i + 2
    ws.cell(row=r2, column=1, value="出价建议").font = SUB_FONT
    r2 += 1
    for rec in placement["recommendations"]:
        ws.cell(row=r2, column=1, value=f"{rec['placement']}: {rec['action']} — {rec['detail']}")
        r2 += 1

    style_data(ws, 4, len(placement["placements"]) + 3, len(ph))
    auto_width(ws, len(ph))

    # ── Sheet 6: Ad Group Structure ───────────────────────────
    ws = wb["广告组结构"]
    ags = ad_group["summary"]
    add_title(ws, f"广告组结构 — {ags['group_count']}组 分布在 {ags['campaign_count']}个活动 | "
              f"ACOS {fmt_pct(ags['overall_acos'])}", 10)

    # Structural issues
    ws.cell(row=3, column=1, value="结构诊断").font = SUB_FONT
    diag_h = ["活动名称", "问题", "详情", "组数", "Top组份额"]
    for c, h in enumerate(diag_h, 1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(diag_h))
    for i, d in enumerate(ad_group["structural_diagnostics"][:30], 5):
        ws.cell(row=i, column=1, value=str(d.get("campaign_name", ""))[:40])
        ws.cell(row=i, column=2, value=d.get("issue", ""))
        ws.cell(row=i, column=3, value=d.get("detail", ""))
        ws.cell(row=i, column=4, value=d.get("group_count", ""))
        ws.cell(row=i, column=5, value=fmt_pct(d.get("top_group_share")))
    style_data(ws, 5, 4 + min(len(ad_group["structural_diagnostics"]), 30), len(diag_h))
    auto_width(ws, len(diag_h))

    # Duplicate names
    r3 = i + 2
    if ad_group["duplicate_names"]:
        ws.cell(row=r3, column=1, value=f"跨活动同名组 ({len(ad_group['duplicate_names'])}个): " +
                ", ".join(ad_group["duplicate_names"])).font = CN_FONT

    # ── Sheet 7: ASIN Efficiency ──────────────────────────────
    ws = wb["ASIN效率"]
    aps = ad_product["summary"]
    add_title(ws, f"ASIN 广告效率 — {aps['asin_count']}个ASIN | "
              f"直接ACOS {fmt_pct(aps['overall_acos'])} | "
              f"混合ACOS(含光环) {fmt_pct(aps['blended_acos_with_halo'])}", 12)

    ah = ["ASIN", "SKU", "花费", "销售额", "ACOS", "ROAS", "CTR", "CVR", "直接销售", "光环销售", "混合ACOS", "光环比"]
    akeys = ["asin", "sku", "spend", "sales", "acos", "roas", "ctr", "cvr", "same_sku_sales", "other_sku_sales", "blended_acos", "halo_ratio"]
    for c, h in enumerate(ah, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(ah))
    for i, r in enumerate(ad_product["ranking"], 4):
        vals = [r.get("asin", ""), str(r.get("sku", ""))[:35],
                fmt_usd(r["spend"]), fmt_usd(r["sales"]), fmt_pct(r["acos"]),
                fmt_num(r.get("roas"), 2), fmt_pct(r.get("ctr")), fmt_pct(r.get("cvr")),
                fmt_usd(r.get("same_sku_sales")), fmt_usd(r.get("other_sku_sales")),
                fmt_pct(r.get("blended_acos")), fmt_num(r.get("halo_ratio"), 2)]
        for j, v in enumerate(vals, 1):
            ws.cell(row=i, column=j, value=v)
        if r.get("acos") is not None and r["acos"] < 0.26:
            for c2 in range(1, len(ah) + 1):
                ws.cell(row=i, column=c2).fill = GOOD
        elif r.get("acos") is not None and r["acos"] > 0.66:
            for c2 in range(1, len(ah) + 1):
                ws.cell(row=i, column=c2).fill = BAD
    style_data(ws, 4, 3 + len(ad_product["ranking"]), len(ah))
    auto_width(ws, len(ah))

    # Zero-sale high-spend ASINs
    rz = i + 2
    if ad_product["zero_sale_high_spend"]:
        ws.cell(row=rz, column=1, value="高花费零销售ASIN (建议暂停或修复Listing):").font = SUB_FONT
        rz += 1
        for z in ad_product["zero_sale_high_spend"]:
            ws.cell(row=rz, column=1, value=f"{z['asin']} {z['sku'][:30]} — ${z['spend']:,.2f}")
            rz += 1

    # ── Sheet 8: Purchased Item (Brand Halo) ──────────────────
    ws = wb["品牌光环"]
    pi_s = purchased["summary"]
    add_title(ws, f"品牌光环效应 — 广告ASIN {pi_s['unique_advertised_asins']}个 → "
              f"购买ASIN {pi_s['unique_purchased_asins']}个 | 光环销售额 ${pi_s['total_purchased_sales']:,.2f}", 8)

    ws.cell(row=3, column=1, value="Gateway ASIN (光环驱动者)").font = SUB_FONT
    gh = ["广告SKU", "广告ASIN", "拉动销量(件)", "拉动销售额", "拉动其他产品种数"]
    for c, h in enumerate(gh, 1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(gh))
    for i, g in enumerate(purchased["gateway_asins"], 5):
        ws.cell(row=i, column=1, value=str(g.get("advertised_sku", ""))[:40])
        ws.cell(row=i, column=2, value=str(g.get("advertised_asin", "")))
        ws.cell(row=i, column=3, value=g.get("purchased_units", ""))
        ws.cell(row=i, column=4, value=fmt_usd(g.get("purchased_sales")))
        ws.cell(row=i, column=5, value=g.get("purchase_events", ""))

    # Cross-sell map
    rc = i + 2
    ws.cell(row=rc, column=1, value="交叉销售明细").font = SUB_FONT
    rc += 1
    ch = ["广告ASIN", "购买ASIN", "活动", "投放/匹配", "销量", "销售额"]
    for c, h in enumerate(ch, 1):
        ws.cell(row=rc, column=c, value=h)
    style_header(ws, rc, len(ch))
    rc += 1
    for cs in purchased["cross_sell_map"][:30]:
        ws.cell(row=rc, column=1, value=str(cs.get("advertised_asin", "")))
        ws.cell(row=rc, column=2, value=str(cs.get("purchased_asin", "")))
        ws.cell(row=rc, column=3, value=str(cs.get("campaign_name", ""))[:25])
        ws.cell(row=rc, column=4, value=f"{cs.get('targeting','')} / {cs.get('match_type','')}")
        ws.cell(row=rc, column=5, value=cs.get("units", ""))
        ws.cell(row=rc, column=6, value=fmt_usd(cs.get("sales")))
        rc += 1

    style_data(ws, 5, rc - 1, len(ch))
    auto_width(ws, len(ch))

    # ── Sheet 9: Action Recommendations ───────────────────────
    ws = wb["行动建议"]
    add_title(ws, f"行动建议 — {account} — {period}", 3)

    actions = []

    # From placement
    for rec in placement["recommendations"]:
        if rec["action"] != "maintain":
            actions.append(("广告位", rec["action"], f"{rec['placement']}: {rec['detail']}"))

    # From AdGroup
    for d in ad_group["structural_diagnostics"]:
        actions.append(("广告组结构", d["issue"], f"{d['campaign_name']}: {d['detail']}"))

    # From AdProduct
    for z in ad_product["zero_sale_high_spend"]:
        actions.append(("ASIN暂停", "pause_review",
            f"ASIN {z['asin']} ({z['sku'][:30]}) — 花费${z['spend']:,.2f} 零销售, 建议暂停并修复Listing"))

    # From PurchasedItem — Gateway protection
    for g in purchased["gateway_asins"]:
        actions.append(("光环保护", "do_not_pause",
            f"Gateway ASIN {g['advertised_sku'][:30]}: 拉动{g['purchased_units']}件${g['purchased_sales']:,.2f}其他产品销售 — 即使自身ACOS高也绝不暂停"))

    # From SearchTerm negate candidates
    negates = search.get("negative_candidates", [])
    for n in negates[:10]:
        actions.append(("否定候选", "negate",
            f"搜索词 '{n.get('search_term','')}' — ${n.get('spend',0):.2f} 花费, {n.get('clicks',0)} 点击, 0订单"))

    # From SearchTerm harvest candidates
    harvests = search.get("harvest_keywords", [])
    for h in harvests[:5]:
        actions.append(("关键词收割", "harvest_add_exact",
            f"搜索词 '{h.get('search_term','')}' — ${h.get('spend',0):.2f} → ${h.get('sales',0):.2f}, 建议加入精准匹配"))

    ah = ["类别", "操作", "详情"]
    for c, h in enumerate(ah, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(ah))
    for i, (cat, act, detail) in enumerate(actions[:50], 4):
        ws.cell(row=i, column=1, value=cat).font = CN_FONT
        ws.cell(row=i, column=2, value=act).font = CN_FONT
        ws.cell(row=i, column=3, value=detail).font = CN_FONT
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        if act in ("lower_bid", "negate", "pause_review"):
            ws.cell(row=i, column=1).fill = BAD
        elif act in ("raise_bid", "harvest_add_exact", "do_not_pause"):
            ws.cell(row=i, column=1).fill = GOOD
    style_data(ws, 4, 3 + min(len(actions), 50), len(ah))
    auto_width(ws, len(ah), max_width=60)
    ws.column_dimensions["C"].width = 80

    # ── Save ──────────────────────────────────────────────────
    wb.save(out)
    print(f"[OK] Report saved: {out}")
    print(f"  Sheets: {ws_names}")
    print(f"  Actions: {len(actions)} recommendations")
    return out


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--account", default="BJRYECLTD-US", help="Account name")
    parser.add_argument("--period", default="2026-06", help="Report period")
    args = parser.parse_args()
    build_report(args.account, args.period)
