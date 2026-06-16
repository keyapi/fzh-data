"""
汇总报告生成器 — 读取 4 个中间 JSON，输出多 sheet Excel + openpyxl 图表。

输出: out/如森US-广告分析报告.xlsx
"""
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from copy import copy

OUT_DIR = os.path.join(os.path.dirname(__file__), "out")

# ── 样式常量 ──────────────────────────────────────────
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="微软雅黑", bold=True, size=12, color="2F5496")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NUM_FONT = Font(name="Consolas", size=10)
CN_FONT = Font(name="微软雅黑", size=10)

METRIC_NAMES = {
    "spend": "花费", "sales_7d": "销售额(7天)", "orders_7d": "订单(7天)",
    "clicks": "点击量", "impressions": "展示量", "acos": "ACOS",
    "roas": "ROAS", "ctr": "CTR", "cpc": "CPC", "cvr": "转化率",
    "budget": "日预算", "budget_utilization": "预算利用率",
    "campaign_name": "广告活动名称", "portfolio_name": "广告组合",
    "status": "状态", "match_type": "匹配类型", "targeting": "投放对象",
    "search_term": "搜索词", "placement": "广告位", "placement_category": "广告位类别",
    "spend_share": "花费占比", "term_category": "搜索词分类",
    "halo_ratio": "光环比率(其他/广告SKU)",
    "advertised_sku_sales": "广告SKU销售额", "other_sku_sales": "其他SKU销售额",
    "advertised_sku_units": "广告SKU销量", "other_sku_units": "其他SKU销量",
    "conversion_rate_7d": "转化率", "top_search_is": "顶部搜索份额",
    "campaign_type": "广告活动类型", "targeting_type": "定位类型",
    "bidding_strategy": "竞价策略", "country": "国家", "retailer": "零售商",
    "currency": "货币", "flag": "标记",
}


def _load_json(filename):
    with open(os.path.join(OUT_DIR, filename), "r", encoding="utf-8") as f:
        return json.load(f)


def _cn(name):
    return METRIC_NAMES.get(name, name)


def _style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(cell, is_money=False, is_pct=False):
    cell.font = NUM_FONT if (is_money or is_pct) else CN_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if is_money and isinstance(cell.value, (int, float)):
        cell.number_format = '#,##0.00'
    elif is_pct and isinstance(cell.value, (float,)):
        cell.number_format = '0.00%'


def _write_table(ws, start_row, headers, rows, money_cols=None, pct_cols=None):
    """在指定 worksheet 写入表头和行数据，返回下一可用行。"""
    money_cols = set(money_cols or [])
    pct_cols = set(pct_cols or [])
    # Header
    for ci, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=ci, value=_cn(h))
    _style_header(ws, start_row, len(headers))
    # Data
    for ri, row in enumerate(rows):
        for ci, h in enumerate(headers):
            val = row.get(h)
            cell = ws.cell(row=start_row + 1 + ri, column=ci + 1, value=val)
            _style_data_cell(cell, is_money=(h in money_cols), is_pct=(h in pct_cols))
    return start_row + 1 + len(rows)


def _add_bar_chart(ws, title, data_start_row, data_end_row, cat_col, val_cols, position, width=18, height=12):
    """添加分组柱状图。"""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height

    n_rows = data_end_row - data_start_row + 1
    n_cols = len(val_cols) + 1  # +cat col
    cat_ref = Reference(ws, min_col=cat_col, min_row=data_start_row, max_row=data_end_row)
    data_ref = Reference(ws, min_col=val_cols[0], min_row=data_start_row - 1,
                         max_col=val_cols[-1], max_row=data_end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.legend.position = "b"
    ws.add_chart(chart, position)


def build():
    campaign = _load_json("campaign_analysis.json")
    targeting = _load_json("targeting_analysis.json")
    search_term = _load_json("search_term_analysis.json")
    placement = _load_json("placement_analysis.json")

    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    # ====== Sheet 1: 总览 ======
    ws1 = wb.create_sheet("总览")
    _build_overview(ws1, campaign, targeting, search_term, placement)

    # ====== Sheet 2: 广告活动 ======
    ws2 = wb.create_sheet("广告活动")
    _build_campaign_sheet(ws2, campaign)

    # ====== Sheet 3: 投放表现 ======
    ws3 = wb.create_sheet("投放表现")
    _build_targeting_sheet(ws3, targeting)

    # ====== Sheet 4: 搜索词洞察 ======
    ws4 = wb.create_sheet("搜索词洞察")
    _build_search_term_sheet(ws4, search_term)

    # ====== Sheet 5: 广告位效率 ======
    ws5 = wb.create_sheet("广告位效率")
    _build_placement_sheet(ws5, placement)

    # ====== Sheet 6: 行动建议 ======
    ws6 = wb.create_sheet("行动建议")
    _build_action_sheet(ws6, campaign, targeting, search_term, placement)

    # 保存
    out_path = os.path.join(OUT_DIR, "如森US-广告分析报告.xlsx")
    wb.save(out_path)
    print(f"报告已生成: {out_path}")
    return out_path


def _build_overview(ws, campaign, targeting, search_term, placement):
    """总览 sheet — 关键数字卡片 + 数据一致性校验。"""
    ws.sheet_properties.tabColor = "2F5496"

    s = campaign.get("summary", {})
    st = targeting.get("summary", {})
    ss = search_term.get("summary", {})
    sp = placement.get("summary", {})

    # 标题
    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value="Amazon 广告分析报告 — 如森US 近30天").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"数据周期: 2026-05-17 ~ 2026-06-15 | 账户: A2.如森跨境电商").font = CN_FONT

    # 关键数字卡片
    cards = [
        ("总花费", f"${s.get('total_spend', 0):,.2f}", "花费总计"),
        ("总销售额(7d)", f"${s.get('total_sales_7d', 0):,.2f}", "7天归因销售额"),
        ("总订单(7d)", f"{s.get('total_orders_7d', 0)}", "7天归因订单"),
        ("整体ACOS", f"{s.get('overall_acos', 0):.2%}", "花费÷销售额"),
        ("整体ROAS", f"{s.get('overall_roas', 0):.2f}x", "销售额÷花费"),
        ("整体CTR", f"{s.get('overall_ctr', 0):.2%}", "点击÷展示"),
        ("广告活动数", f"{s.get('campaign_count', 0)}", "总活动数"),
        ("搜索词数", f"{ss.get('search_term_count', 0):,}", "客户实际搜索词"),
    ]

    for i, (label, value, desc) in enumerate(cards):
        row = 5 + i // 4 * 3
        col = 1 + (i % 4) * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(name="微软雅黑", size=10, color="666666")
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=row + 1, column=col, value=value).font = Font(name="Consolas", bold=True, size=16, color="2F5496")
        ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        dc = ws.cell(row=row + 2, column=col, value=desc)
        dc.font = Font(name="微软雅黑", size=9, color="999999")
        dc.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)

    # 数据一致性校验
    check_row = 14
    ws.cell(row=check_row, column=1, value="数据一致性校验").font = SUBTITLE_FONT
    checks = []
    if s.get("total_spend") and ss.get("total_spend"):
        diff = abs(s["total_spend"] - ss["total_spend"])
        checks.append((f"活动花费 ${s['total_spend']:,.2f} vs 搜索词花费 ${ss['total_spend']:,.2f}", "一致" if diff < 10 else "有差异"))
    if s.get("total_spend") and sp.get("total_spend"):
        diff = abs(s["total_spend"] - sp["total_spend"])
        checks.append((f"活动花费 vs 广告位花费 ${sp['total_spend']:,.2f}", "一致" if diff < 10 else "有差异"))
    for i, (text, status) in enumerate(checks):
        c = ws.cell(row=check_row + 1 + i, column=1, value=text)
        c.font = CN_FONT
        sc = ws.cell(row=check_row + 1 + i, column=2, value=status)
        sc.font = Font(name="微软雅黑", bold=True, color="006100" if status == "一致" else "9C0006")
        sc.fill = GOOD_FILL if status == "一致" else BAD_FILL

    # 整体结论
    conclusion_row = 20
    ws.cell(row=conclusion_row, column=1, value="核心发现").font = SUBTITLE_FONT
    findings = []
    acos_val = s.get("overall_acos", 0)
    roas_val = s.get("overall_roas", 0)
    if acos_val and acos_val < 0.25:
        findings.append(f"整体 ACOS {acos_val:.1%}，处于健康区间（<25%），广告投放整体盈利。")
    elif acos_val and acos_val <= 0.35:
        findings.append(f"整体 ACOS {acos_val:.1%}，处于可接受区间（25-35%），建议优化高 ACOS 活动。")
    else:
        findings.append(f"整体 ACOS {acos_val:.1%}，偏高，需重点审视花费结构。")
    if roas_val and roas_val > 3:
        findings.append(f"ROAS {roas_val:.2f}x，投入产出表现良好。")
    if ss.get("negative_candidate_count", 0) > 10:
        findings.append(f"发现 {ss['negative_candidate_count']} 个否定词候选，屏蔽后可节省约 ${ss['negative_wasted_spend']:,.2f}。")
    if ss.get("harvest_count", 0) > 0:
        findings.append(f"发现 {ss['harvest_count']} 个关键词收割机会，建议加入精准匹配活动。")
    for i, f_text in enumerate(findings):
        ws.cell(row=conclusion_row + 1 + i, column=1, value=f"• {f_text}").font = CN_FONT

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    for ci in range(5, 9):
        ws.column_dimensions[get_column_letter(ci)].width = 18


def _build_campaign_sheet(ws, campaign):
    """广告活动 sheet — 排行表 + ACOS vs ROAS 散点图。"""
    ws.sheet_properties.tabColor = "4472C4"

    ws.cell(row=1, column=1, value="广告活动排行").font = SUBTITLE_FONT
    headers = ["campaign_name", "status", "portfolio_name", "spend", "sales_7d",
               "acos", "roas", "orders_7d", "clicks", "impressions", "ctr", "cpc",
               "budget", "budget_utilization", "flag"]
    ranking = campaign.get("ranking", [])
    money_cols = {"spend", "sales_7d", "cpc", "budget"}
    pct_cols = {"acos", "roas", "ctr", "budget_utilization"}
    end_row = _write_table(ws, 3, headers, ranking, money_cols=money_cols, pct_cols=pct_cols)

    # 条件格式: flag 列
    flag_col = len(headers)  # flag 是最后一列
    for ri in range(len(ranking)):
        row_idx = 4 + ri
        flag_cell = ws.cell(row=row_idx, column=flag_col)
        if flag_cell.value == "优胜":
            for ci in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=ci).fill = GOOD_FILL
        elif flag_cell.value in ("高风险", "问题"):
            for ci in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=ci).fill = BAD_FILL

    # 散点图: ACOS vs ROAS
    chart_row = end_row + 2
    ws.cell(row=chart_row, column=1, value="ACOS vs ROAS 散点图").font = SUBTITLE_FONT

    # 临时写入图表数据
    data_start = chart_row + 1
    ws.cell(row=data_start, column=1, value="ACOS")
    ws.cell(row=data_start, column=2, value="ROAS")
    ws.cell(row=data_start, column=3, value="活动名")
    for i, r in enumerate(ranking[:37]):
        acos_v = r.get("acos")
        roas_v = r.get("roas")
        if acos_v is not None and roas_v is not None and roas_v > 0:
            ws.cell(row=data_start + 1 + i, column=1, value=float(acos_v) if acos_v else 0)
            ws.cell(row=data_start + 1 + i, column=2, value=float(roas_v) if roas_v else 0)
            ws.cell(row=data_start + 1 + i, column=3, value=str(r.get("campaign_name", "")))

    chart = ScatterChart()
    chart.title = "ACOS vs ROAS"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    data_ref = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(ranking))
    cat_ref = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(ranking))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.x_axis.title = "ACOS"
    chart.y_axis.title = "ROAS"
    ws.add_chart(chart, f"D{chart_row}")

    # 列宽
    ws.column_dimensions['A'].width = 35
    for ci in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _build_targeting_sheet(ws, targeting):
    """投放表现 sheet — 匹配类型 + 光环效应 + TOP/BOTTOM。"""
    ws.sheet_properties.tabColor = "ED7D31"

    # 匹配类型
    ws.cell(row=1, column=1, value="按匹配类型表现").font = SUBTITLE_FONT
    match_headers = ["match_type", "spend", "sales_7d", "acos", "roas", "orders_7d",
                      "clicks", "impressions", "ctr_calc", "cpc_avg"]
    match_data = targeting.get("by_match_type", [])
    _write_table(ws, 3, match_headers, match_data,
                 money_cols={"spend", "sales_7d", "cpc_avg"},
                 pct_cols={"acos", "roas", "ctr_calc"})

    # 柱状图
    if match_data:
        _add_bar_chart(ws, "匹配类型 ACOS 对比", 4, 3 + len(match_data), 1, [4],
                       "I3", width=16, height=10)

    # 光环效应
    halo_row = 3 + len(match_data) + 2
    halo = targeting.get("halo_effect", {})
    if halo:
        ws.cell(row=halo_row, column=1, value="光环效应（广告SKU → 其他SKU）").font = SUBTITLE_FONT
        halo_items = [
            ("广告SKU销售额", f"${halo.get('advertised_sku_sales', 0):,.2f}"),
            ("其他SKU销售额", f"${halo.get('other_sku_sales', 0):,.2f}"),
            ("光环比率", f"{halo.get('halo_ratio', 0):.2f}x"),
            ("广告SKU销量", f"{halo.get('advertised_sku_units', 0)}"),
            ("其他SKU销量", f"{halo.get('other_sku_units', 0)}"),
        ]
        for i, (label, val) in enumerate(halo_items):
            ws.cell(row=halo_row + 1 + i, column=1, value=label).font = CN_FONT
            ws.cell(row=halo_row + 1 + i, column=2, value=val).font = NUM_FONT

    # 零转化投放 — 在光环下面
    waste_row = halo_row + 2 + len(halo_items)
    bottom = targeting.get("bottom_targets", [])
    if bottom:
        ws.cell(row=waste_row, column=1, value=f"零转化投放 TOP{len(bottom)}（有花费无订单）").font = SUBTITLE_FONT
        bt_headers = ["targeting", "match_type", "campaign_name", "spend", "clicks", "impressions"]
        _write_table(ws, waste_row + 1, bt_headers, bottom, money_cols={"spend"})

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 40
    for ci in range(4, 12):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _build_search_term_sheet(ws, search_term):
    """搜索词洞察 sheet — 关键词收割 + 否定词 + 搜索词分类。"""
    ws.sheet_properties.tabColor = "70AD47"

    # 关键词收割
    ws.cell(row=1, column=1, value="关键词收割清单（建议加入精准匹配）").font = SUBTITLE_FONT
    harvest = search_term.get("harvest_keywords", [])
    h_headers = ["search_term", "campaign_name", "match_type", "clicks", "orders_7d",
                  "sales_7d", "spend", "acos", "roas", "ctr", "cpc"]
    next_row = _write_table(ws, 3, h_headers, harvest,
                            money_cols={"spend", "sales_7d", "cpc"},
                            pct_cols={"acos", "roas", "ctr"})

    # 否定词
    neg_row = next_row + 2
    negatives = search_term.get("negative_candidates", [])
    ws.cell(row=neg_row, column=1, value=f"否定词候选（点击≥10、零订单，建议屏蔽）— 共{len(negatives)}个，浪费${search_term['summary'].get('negative_wasted_spend',0):,.2f}").font = SUBTITLE_FONT
    n_headers = ["search_term", "campaign_name", "match_type", "clicks", "spend", "impressions"]
    end_neg = _write_table(ws, neg_row + 1, n_headers, negatives, money_cols={"spend"})

    # 搜索词分类
    cat_row = end_neg + 2
    cats = search_term.get("category_distribution", [])
    if cats:
        ws.cell(row=cat_row, column=1, value="搜索词分类分布").font = SUBTITLE_FONT
        cat_headers = ["term_category", "spend", "sales_7d", "orders_7d", "clicks",
                        "impressions", "acos", "roas", "count"]
        cat_end = _write_table(ws, cat_row + 1, cat_headers, cats,
                               money_cols={"spend", "sales_7d"},
                               pct_cols={"acos", "roas"})

        # 饼图
        pie = PieChart()
        pie.title = "搜索词分类-花费占比"
        pie.width = 16
        pie.height = 10
        cat_data_start = cat_row + 2
        data_ref = Reference(ws, min_col=2, min_row=cat_row + 1,
                             max_row=cat_row + 1 + len(cats))
        cat_ref = Reference(ws, min_col=1, min_row=cat_data_start,
                            max_row=cat_data_start - 1 + len(cats))
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cat_ref)
        colors = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47"]
        for i in range(len(cats)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = colors[i % len(colors)]
            pie.series[0].data_points.append(pt)
        ws.add_chart(pie, f"I{cat_row}")

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 35
    for ci in range(3, 12):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _build_placement_sheet(ws, placement):
    """广告位效率 sheet — 四位对比 + 柱状图 + 出价建议。"""
    ws.sheet_properties.tabColor = "FFC000"

    ws.cell(row=1, column=1, value="广告位效率对比").font = SUBTITLE_FONT
    placements = placement.get("placements", [])
    p_headers = ["placement", "spend", "spend_share", "sales_7d", "acos", "roas",
                  "orders_7d", "clicks", "impressions", "ctr", "cpc", "cvr"]
    p_end = _write_table(ws, 3, p_headers, placements,
                         money_cols={"spend", "sales_7d", "cpc"},
                         pct_cols={"spend_share", "acos", "roas", "ctr", "cvr"})

    # 分组柱状图: CPC/CTR/CVR 对比
    if placements:
        chart_row = p_end + 2
        # 临时写入数据
        ws.cell(row=chart_row, column=1, value="广告位")
        ws.cell(row=chart_row, column=2, value="CPC")
        ws.cell(row=chart_row, column=3, value="CTR")
        ws.cell(row=chart_row, column=4, value="CVR")
        ws.cell(row=chart_row, column=5, value="ACOS")
        for i, p in enumerate(placements):
            ws.cell(row=chart_row + 1 + i, column=1, value=p["placement"])
            ws.cell(row=chart_row + 1 + i, column=2, value=p.get("cpc", 0) or 0)
            ws.cell(row=chart_row + 1 + i, column=3, value=p.get("ctr", 0) or 0)
            ws.cell(row=chart_row + 1 + i, column=4, value=p.get("cvr", 0) or 0)
            ws.cell(row=chart_row + 1 + i, column=5, value=p.get("acos", 0) or 0)

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "广告位效率对比 (CPC/CTR/CVR/ACOS)"
        chart.style = 10
        chart.width = 20
        chart.height = 12
        data_ref = Reference(ws, min_col=2, min_row=chart_row,
                             max_col=5, max_row=chart_row + len(placements))
        cat_ref = Reference(ws, min_col=1, min_row=chart_row + 1,
                            max_row=chart_row + len(placements))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        ws.add_chart(chart, f"H{chart_row}")

        # 出价调整建议
        rec_row = chart_row + len(placements) + 2
        recs = placement.get("recommendations", [])
        if recs:
            ws.cell(row=rec_row, column=1, value="广告位出价调整建议").font = SUBTITLE_FONT
            ws.cell(row=rec_row + 1, column=1, value="广告位").font = HEADER_FONT
            ws.cell(row=rec_row + 1, column=1).fill = HEADER_FILL
            ws.cell(row=rec_row + 1, column=2, value="建议操作").font = HEADER_FONT
            ws.cell(row=rec_row + 1, column=2).fill = HEADER_FILL
            for i, rec in enumerate(recs):
                ws.cell(row=rec_row + 2 + i, column=1, value=rec["placement"]).font = CN_FONT
                ws.cell(row=rec_row + 2 + i, column=2, value=rec["action"]).font = CN_FONT

    ws.column_dimensions['A'].width = 20
    for ci in range(2, 13):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _build_action_sheet(ws, campaign, targeting, search_term, placement):
    """行动建议 sheet — 自动生成的操作清单。"""
    ws.sheet_properties.tabColor = "A5A5A5"

    ws.cell(row=1, column=1, value="自动生成行动建议").font = TITLE_FONT
    ws.cell(row=2, column=1, value="基于广告数据分析，建议以下操作（按优先级排序）").font = CN_FONT

    actions = []
    n = 1

    # 1. 否定词添加
    negatives = search_term.get("negative_candidates", [])
    if negatives:
        wasted = search_term["summary"].get("negative_wasted_spend", 0)
        actions.append(("高", f"添加否定关键词（{len(negatives)}个）",
                         f"这些搜索词共花费 ${wasted:,.2f}，零转化。建议在对应广告活动中添加为否定精准匹配。",
                         "具体词参见「搜索词洞察」sheet 否定词候选列表。"))

    # 2. 关键词收割
    harvest = search_term.get("harvest_keywords", [])
    if harvest:
        harvest_spend = sum(h.get("spend", 0) or 0 for h in harvest)
        harvest_sales = sum(h.get("sales_7d", 0) or 0 for h in harvest)
        actions.append(("高", f"关键词收割（{len(harvest)}个）",
                         f"这些搜索词共产生 ${harvest_sales:,.2f} 销售额，花费 ${harvest_spend:,.2f}。建议新建精准匹配活动或加入现有精准组。",
                         "具体词参见「搜索词洞察」sheet 关键词收割清单。"))

    # 3. 广告位优化
    placements = placement.get("placements", [])
    for p in placements:
        acos_val = p.get("acos", 0) or 0
        if acos_val > 0.40 and p.get("spend", 0) > 100:
            actions.append(("中", f"降低 {p['placement']} 出价",
                             f"该广告位 ACOS {acos_val:.1%}，花费 ${p['spend']:,.2f}。建议降低出价 15-30%。",
                             ""))
        elif acos_val < 0.20 and p.get("cvr", 0) and p.get("cvr", 0) > 0.03:
            actions.append(("中", f"提高 {p['placement']} 出价",
                             f"该广告位 ACOS {acos_val:.1%}，CVR {p.get('cvr',0):.2%}，转化效率高。建议提高出价 10-20% 扩展流量。",
                             ""))

    # 4. 问题活动
    problems = campaign.get("problems", [])
    for p in problems[:5]:
        actions.append(("中", f"检查活动: {p.get('campaign_name', '')}",
                         f"ACOS {p.get('acos', 0):.1%}，花费 ${p.get('spend', 0):,.2f}。建议检查投放词相关性和 Listing 转化率。",
                         ""))

    # 5. 光环效应
    halo = targeting.get("halo_effect", {})
    halo_ratio = halo.get("halo_ratio", 0)
    if halo_ratio and halo_ratio > 1:
        actions.append(("低", "关注光环效应",
                         f"其他SKU销售额是广告SKU的 {halo_ratio:.1f}x，广告在拉动非广告商品销售。继续保持商品投放策略。",
                         ""))

    # 写入
    action_row = 4
    ws.cell(row=action_row, column=1, value="优先级").font = HEADER_FONT
    ws.cell(row=action_row, column=1).fill = HEADER_FILL
    ws.cell(row=action_row, column=2, value="操作项").font = HEADER_FONT
    ws.cell(row=action_row, column=2).fill = HEADER_FILL
    ws.cell(row=action_row, column=3, value="原因").font = HEADER_FONT
    ws.cell(row=action_row, column=3).fill = HEADER_FILL
    ws.cell(row=action_row, column=4, value="参考").font = HEADER_FONT
    ws.cell(row=action_row, column=4).fill = HEADER_FILL

    for i, (priority, title, reason, ref) in enumerate(actions):
        r = action_row + 1 + i
        pc = ws.cell(row=r, column=1, value=priority)
        pc.font = Font(name="微软雅黑", bold=True, size=11)
        if priority == "高":
            pc.fill = BAD_FILL
        elif priority == "中":
            pc.fill = WARN_FILL
        else:
            pc.fill = GOOD_FILL
        pc.alignment = Alignment(horizontal="center")
        pc.border = THIN_BORDER

        tc = ws.cell(row=r, column=2, value=f"{i+1}. {title}")
        tc.font = Font(name="微软雅黑", bold=True)
        tc.border = THIN_BORDER

        rc = ws.cell(row=r, column=3, value=reason)
        rc.font = CN_FONT
        rc.border = THIN_BORDER

        refc = ws.cell(row=r, column=4, value=ref)
        refc.font = Font(name="微软雅黑", size=9, color="666666")
        refc.border = THIN_BORDER

    # 周优化建议
    weekly_row = action_row + len(actions) + 3
    ws.cell(row=weekly_row, column=1, value="后续优化节奏").font = SUBTITLE_FONT
    weekly_cadence = [
        "周一: 下载最新搜索词报告 → 跑本分析脚本 → 检查新增否定词候选",
        "周三: 审查自动广告搜索词 → 将有转化的词加入关键词收割清单",
        "周五: 调整出价 → 检查预算利用率 → 确认库存情况",
        "每两周: 下载广告位报告 → 对比广告位效率 → 调整出价系数",
    ]
    for i, text in enumerate(weekly_cadence):
        ws.cell(row=weekly_row + 1 + i, column=1, value=f"• {text}").font = CN_FONT

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 38


if __name__ == "__main__":
    build()
