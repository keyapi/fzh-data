# -*- coding: utf-8 -*-
"""销售出库(DN) → 报关单据导出（4 个固定 sheet，严格对齐海关模板）。

用法:
  python customs_export.py --dn DN-26-00063          # 生产环境
  python customs_export.py --dn DN-26-00063 --test   # 测试环境

设计:
  用 openpyxl 打开 数据源/ZJ26DZJR0403-报关单据.xlsx 作为模板，
  仅覆盖「可变单元格」，保留合并单元格/边框/字体/列宽；并清除模板自带的图片(中基抬头/印章)。
"""
from __future__ import annotations

import argparse
import datetime
import os
import sys
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

import requests

sys.stdout.reconfigure(encoding="utf-8")

_DIR = Path(__file__).resolve().parent
TEMPLATE = _DIR / "数据源" / "ZJ26DZJR0403-报关单据.xlsx"
OUT_DIR = _DIR / "out"

ENV_URLS = {"test": "https://ensh.vilavi.cn", "prod": "https://erpnext.vilavi.cn"}

# ──────────────────────────────────────────────────────────────
# 常量配置
# ──────────────────────────────────────────────────────────────
CONFIG = {
    "shipper_cn": "方州汇国际电子商务（北京）有限公司",
    "shipper_en": "FANGZHOUHUI INTERNATIONAL E-COMMERCE (BEIJING) CO., LTD.",
    "shipper_addr": "Room 1910, 16th Floor, Building 1, Cuijing Beili, Tongzhou District, Beijing, China",
    "currency": "USD",            # 报关币制
    "exchange_rate": 6.8,         # 人民币 → 美元（暂定，后续换汇率表）
    "price_markup": 1.45,         # 报关单价加成系数：报关单价 = BOM成本 ÷ 汇率 × 1.45
    "trade_term": "FOB",
    "trade_term_full": "FOB NINGBO,CHINA",
    "payment_term": "BY T/T 90 DAYS",
    "transport_route": "FROM NINGBO,CHINA TO LONG BEACH,UNITED STATES BY SEA",
    "transport_mode": "BY SEA",
    "port_loading": "NINGBO,CHINA",                # 出境关别/离境口岸
    "port_discharge": "LONG BEACH,UNITED STATES",  # 指运港
    "supervision_mode": "一般贸易",        # 监管方式
    "origin_place": "绍兴市（33069）",      # 境内货源地（常量）
    "declaration_unit": "宁波市鸿欣报关有限公司",   # 申报单位（弹窗可传，未传则保留模板原值）
    "production_unit": "绍兴雪雁针纺有限公司",       # 生产销售单位（固定值，不在弹窗体现）
    # 单位映射：EN 的 uom → (英文单位, 中文单位)
    "uom_map": {
        "个": ("PIECES", "个"),
        "套": ("SETS", "套"),
        "件": ("PIECES", "件"),
        "只": ("PIECES", "只"),
        "条": ("PIECES", "条"),
    },
}

# 目的国（地区）映射：客户 → (英文, 中文)
DEST_COUNTRY = {
    "UNITED STATES(502)": "美国(502)",
    "PL(327)": "波兰(327)",
}
US_CUSTOMERS = {"DANEEY", "CENTRADE"}   # 美国客户（含美国FBA仓，具体客户名待补充）
PL_CUSTOMERS = set()                     # 波兰公司（具体客户名待补充）

# 境外收货人预设（与 EN 弹窗下拉同源；导出时填 C5 境外收货人：名称 + 地址）。
# 新增/修改收货人只改这一处。
CONSIGNEE_DATA = {
    "centrade": {"name": "Centrade Inc", "addr": "389 Route 10 Unit R, East Hanover, NJ 07936 U.S.A."},
    "daneey":   {"name": "Daneey LLC", "addr": "10812 Fallstone Rd, Suite 402, Houston TX 77099 U.S.A."},
    "poland":   {"name": "SHAOXING XUEYAN ZHENGFANG Sp. z o.o.", "addr": "ul. Prosta 2 95-035 Ozorków Poland"},
}
DEFAULT_CONSIGNEE = "centrade"           # 弹窗默认选中项（EN 侧）；本脚本按客户自动映射

# 装箱组合（混装版）：key = DN 单号 → 组合列表。装箱信息完全由用户确认，不用外箱子表。
# 每个组合:
#   name          组合名
#   carton_qty    箱数（该组合几个箱子）
#   gross_kg      每箱毛重 kg
#   net_kg        每箱净重 kg
#   volume_cbm    每箱体积 m³
#   items         每箱内容: [{code: 去色后物料编码, qty_per_carton: 每箱数量}, ...]
CARTON_GROUPS_BY_DN = {
    "DN-26-00056": [
        {"name": "组合1 KS0001-153 皮壳+内胆", "carton_qty": 40, "gross_kg": 18.5, "net_kg": 17.8, "volume_cbm": 0.09024,
         "items": [
             {"code": "PK#KS0001-DM-153", "qty_per_carton": 2},
             {"code": "PK#KS0001-HLR-153", "qty_per_carton": 8},
             {"code": "ND#KS0001-153-CYF", "qty_per_carton": 9},
         ]},
        {"name": "组合2 KS0001-194 皮壳+内胆", "carton_qty": 32, "gross_kg": 17.0, "net_kg": 16.2, "volume_cbm": 0.09024,
         "items": [
             {"code": "PK#KS0001-DM-194", "qty_per_carton": 2},
             {"code": "PK#KS0001-HLR-194", "qty_per_carton": 1},
             {"code": "ND#KS0001-194-CYF", "qty_per_carton": 10},
         ]},
        {"name": "组合3 KS0001-100/140 内胆+皮壳", "carton_qty": 7, "gross_kg": 15.0, "net_kg": 14.2, "volume_cbm": 0.0768,
         "items": [
             {"code": "ND#KS0001-100-CYF", "qty_per_carton": 15},
             {"code": "ND#KS0001-140-CYF", "qty_per_carton": 3},
             {"code": "PK#KS0001-HLR-140", "qty_per_carton": 3},
         ]},
        {"name": "组合4 KS0003/KS0007 60/194", "carton_qty": 5, "gross_kg": 16.0, "net_kg": 15.2, "volume_cbm": 0.0768,
         "items": [
             {"code": "PK#KS0003-DM-60", "qty_per_carton": 4},
             {"code": "ND#KS0003-60-CYF", "qty_per_carton": 4},
             {"code": "PK#KS0007-DM-194", "qty_per_carton": 1},
             {"code": "ND#KS0007-194-CYF", "qty_per_carton": 1},
         ]},
        {"name": "组合5 KS0321+KS0383 床头靠枕", "carton_qty": 8, "gross_kg": 20.0, "net_kg": 19.1, "volume_cbm": 0.09024,
         "items": [
             {"code": "PK#KS0321-HLR-153", "qty_per_carton": 3},
             {"code": "ND#KS0321-153-CYF", "qty_per_carton": 3},
             {"code": "ND#KS0383-153x50x24-CYF", "qty_per_carton": 6},
             {"code": "ND#KS0383-194x50x24-CYF", "qty_per_carton": 2},
         ]},
        {"name": "组合6 KS0248+KS0401 大件", "carton_qty": 3, "gross_kg": 22.0, "net_kg": 21.0, "volume_cbm": 0.09024,
         "items": [
             {"code": "PK#KS0248-DM-194", "qty_per_carton": 2},
             {"code": "PK#KS0401-XRDR-100x100x80", "qty_per_carton": 3},
         ]},
    ],
}


def load_credentials(env: str) -> tuple[str, str]:
    vals: dict[str, str] = {}
    env_file = _DIR / ".env"
    if env_file.is_file():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            vals[k.strip()] = v.strip()
    if env == "test":
        key, sec = vals.get("TEST_ERP_API_KEY", ""), vals.get("TEST_ERP_API_SECRET", "")
    else:
        key, sec = vals.get("PROD_ERP_API_KEY", ""), vals.get("PROD_ERP_API_SECRET", "")
    return key or vals.get("ERP_API_KEY", ""), sec or vals.get("ERP_API_SECRET", "")


def api_get(base: str, key: str, sec: str, path: str) -> dict:
    r = requests.get(f"{base}{path}", headers={"Authorization": f"token {key}:{sec}"}, timeout=60)
    r.raise_for_status()
    return r.json()


def get_doc(base: str, key: str, sec: str, doctype: str, name: str) -> dict:
    path = "/api/resource/{}/{}".format(
        urllib.parse.quote(doctype, safe=""), urllib.parse.quote(name, safe="")
    )
    return api_get(base, key, sec, path).get("data", {})


# ──────────────────────────────────────────────────────────────
# 简单翻译：中文品名 → 英文（复用 DeepSeek API，与 EN 的 AIContentGenerator 同源）
# ──────────────────────────────────────────────────────────────
DEEPSEEK_BASE_URL = "https://api.vilavi.cn/v1"   # 用户 AI 网关（部署到 EN 后改走 AIContentGenerator/PIM Settings）
DEEPSEEK_MODEL = "deepseek-v4-flash"


def load_deepseek_key() -> str:
    key = os.environ.get("DEEPSEEK_API_KEY", "")
    if key:
        return key
    env_file = _DIR / ".env"
    if env_file.is_file():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("DEEPSEEK_API_KEY="):
                return line.partition("=")[2].strip()
    return ""


def translate_zh_to_en(text: str, api_key: str) -> str:
    """中文 → 英文（海关报关品名），失败回退原文。"""
    if not text or not text.strip() or not api_key:
        return text
    url = f"{DEEPSEEK_BASE_URL}/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }
    payload = {
        "model": DEEPSEEK_MODEL,
        "messages": [
            {
                "role": "system",
                "content": "你是专业的海关报关品名翻译助手。把中文品名翻译成简洁、正式的英文品名，"
                           "用于出口报关单。只输出英文品名本身，不要任何解释、引号或前后缀。",
            },
            {"role": "user", "content": text},
        ],
        "temperature": 0.3,
    }
    try:
        for attempt in range(3):
            try:
                r = requests.post(url, json=payload, headers=headers, timeout=60)
                r.raise_for_status()
                return r.json()["choices"][0]["message"]["content"].strip()
            except requests.RequestException as e:
                if attempt == 2:
                    raise
                print(f"  [翻译重试 {attempt+1}] {text!r}: {e}", file=sys.stderr)
                time.sleep(3)
    except Exception as e:
        print(f"  [翻译失败] {text!r}: {e}", file=sys.stderr)
        return text


# ──────────────────────────────────────────────────────────────
# 聚合：去掉 item_code / item_name 最后一个 "-" 段（颜色）
# ──────────────────────────────────────────────────────────────
def drop_color(s: str) -> str:
    s = (s or "").strip()
    return s.rsplit("-", 1)[0] if s else s


def aggregate_items(dn: dict) -> list[dict]:
    groups: dict[str, dict] = {}
    for it in dn.get("items", []):
        code = it.get("item_code", "")
        key = drop_color(code)
        g = groups.setdefault(key, {
            "code_agg": key,
            "name_agg": drop_color(it.get("item_name", "")),
            "qty": 0.0,
            "amount": 0.0,      # 销售金额（发票/合同用）
            "bom_cost": 0.0,    # BOM 成本（报关单用）
            "uom": it.get("uom") or it.get("stock_uom") or "个",
        })
        g["qty"] += float(it.get("qty") or 0)
        g["amount"] += float(it.get("amount") or 0)
        g["bom_cost"] += float(it.get("bom_cost") or 0)
    for g in groups.values():
        g["rate"] = g["amount"] / g["qty"] if g["qty"] else 0.0          # 销售单价
        g["bom_rate"] = g["bom_cost"] / g["qty"] if g["qty"] else 0.0    # BOM 单价
    return list(groups.values())


def num_to_words(n: int) -> str:
    ones = ["", "ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN",
            "EIGHT", "NINE", "TEN", "ELEVEN", "TWELVE", "THIRTEEN",
            "FOURTEEN", "FIFTEEN", "SIXTEEN", "SEVENTEEN", "EIGHTEEN", "NINETEEN"]
    tens = ["", "", "TWENTY", "THIRTY", "FORTY", "FIFTY", "SIXTY",
            "SEVENTY", "EIGHTY", "NINETY"]
    n = int(n)
    if n < 20:
        return ones[n]
    if n < 100:
        return tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")
    if n < 1000:
        return ones[n // 100] + " HUNDRED" + (" AND " + num_to_words(n % 100) if n % 100 else "")
    if n < 1000000:
        return num_to_words(n // 1000) + " THOUSAND" + (" " + num_to_words(n % 1000) if n % 1000 else "")
    return str(n)


def fmt_date_contract(d: str) -> str:
    dt = datetime.date.fromisoformat(d[:10])
    m = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
         "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"][dt.month - 1]
    return f"{dt.day}/{m}/{dt.year}"


def uom_en(uom: str) -> str:
    return CONFIG["uom_map"].get(uom, (uom, uom))[0]


def uom_cn(uom: str) -> str:
    return CONFIG["uom_map"].get(uom, (uom, uom))[1]


def resolve_country(customer: str) -> tuple[str, str]:
    """客户 → (英文国别, 中文国别)"""
    if customer in US_CUSTOMERS:
        return "UNITED STATES(502)", "美国(502)"
    if customer in PL_CUSTOMERS:
        return "PL(327)", "波兰(327)"
    return "", ""


def resolve_consignee(args, customer: str) -> dict | None:
    """境外收货人：优先 CLI --consignee 指定，否则按客户自动映射。

    返回 {"name": ..., "addr": ...}；无法解析返回 None（保留模板原值）。
    """
    choice = (args.consignee or "").strip().lower()
    if choice == "custom":
        if args.consignee_name or args.consignee_addr:
            return {"name": args.consignee_name or "", "addr": args.consignee_addr or ""}
        return None
    if choice in CONSIGNEE_DATA:
        info = dict(CONSIGNEE_DATA[choice])
        if args.consignee_name:
            info["name"] = args.consignee_name
        if args.consignee_addr:
            info["addr"] = args.consignee_addr
        return info
    cust = (customer or "").lower()
    if "daneey" in cust:
        return dict(CONSIGNEE_DATA["daneey"])
    if "centrade" in cust:
        return dict(CONSIGNEE_DATA["centrade"])
    for key in PL_CUSTOMERS:
        if key.lower() in cust:
            return dict(CONSIGNEE_DATA["poland"])
    if args.consignee_name or args.consignee_addr:
        return {"name": args.consignee_name or "", "addr": args.consignee_addr or ""}
    return None


def usd_price(rmb: float) -> float:
    """RMB → USD 报关价：BOM成本 × 加成系数 ÷ 汇率（先涨价、后换汇），保留 3 位小数。"""
    return round(float(rmb) * CONFIG["price_markup"] / CONFIG["exchange_rate"], 3)


def _n3(ws, coord: str, value):
    """写入数值并强制保留 3 位小数（补零对齐）。"""
    ws[coord] = round(float(value), 3)
    ws[coord].number_format = "0.000"


def compute_packing(agg, groups):
    """根据装箱组合计算每个物料的 箱数/毛重/净重/体积（按数量占比分摊）。

    返回:
        {"qty","cartons","gross","net","volume": {code: 值}, "total_cartons","total_gross","total_net","total_volume"}
    """
    from collections import defaultdict
    c_qty = defaultdict(float)
    c_cartons = defaultdict(float)
    c_gross = defaultdict(float)
    c_net = defaultdict(float)
    c_volume = defaultdict(float)
    total_cartons = total_gross = total_net = total_volume = 0.0
    for g in groups:
        c = float(g.get("carton_qty", 0) or 0)
        g_net = float(g.get("net_kg", 0) or 0)
        g_gross = float(g.get("gross_kg", 0) or 0)
        g_vol = float(g.get("volume_cbm", 0) or 0)
        total_cartons += c
        total_gross += c * g_gross
        total_net += c * g_net
        total_volume += c * g_vol
        items = g.get("items", []) or []
        total_in = sum(float(i.get("qty_per_carton", 0) or 0) for i in items) or 1
        for i in items:
            code = i.get("code", "")
            qpc = float(i.get("qty_per_carton", 0) or 0)
            share = qpc / total_in
            c_qty[code] += c * qpc
            c_cartons[code] += c
            c_net[code] += c * g_net * share
            c_gross[code] += c * (g_net * share + (g_gross - g_net) * share)
            c_volume[code] += c * g_vol * share
    return {
        "qty": dict(c_qty),
        "cartons": dict(c_cartons),
        "gross": dict(c_gross),
        "net": dict(c_net),
        "volume": dict(c_volume),
        "total_cartons": round(total_cartons),
        "total_gross": round(total_gross, 3),
        "total_net": round(total_net, 3),
        "total_volume": round(total_volume, 3),
    }


def _copy_row_style(ws, from_row: int, to_row: int, max_col: int):
    """把 from_row 单元格样式（字体/边框/填充/对齐/数字格式）复制到 to_row（insert_rows 不继承样式）。"""
    import copy
    for col in range(1, max_col + 1):
        src = ws.cell(row=from_row, column=col)
        dst = ws.cell(row=to_row, column=col)
        if src.has_style:
            dst._style = copy.copy(src._style)


def _expand_rows(ws, at_row: int, k: int, extend_ranges: list[str], ref_rows: list[int] | None = None):
    """在 at_row 前插入 k 行，修复合并单元格，并复制样式。

    - at_row 及以下的合并整体下移 k 行
    - extend_ranges（如 marks 竖列）下边界 + k
    - 复制模板数据行样式到插入行（insert_rows 不继承样式，否则新行字体/边框缺失）
    - ref_rows：插入行对应的样式参考行（缺省取 at_row-1）；报关单NEW 每物料 2 行交替传 [item行, element行]
    """
    if k <= 0:
        return
    ws.insert_rows(at_row, k)
    to_move = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= at_row:
            to_move.append(CellRange(str(rng)))
            ws.merged_cells.remove(rng)
    for r in to_move:
        ws.merged_cells.add(
            f"{get_column_letter(r.min_col)}{r.min_row + k}:{get_column_letter(r.max_col)}{r.max_row + k}"
        )
    for coord in extend_ranges:
        r = CellRange(coord)
        if str(r) in [str(x) for x in ws.merged_cells.ranges]:
            ws.merged_cells.remove(r)
        ws.merged_cells.add(
            f"{get_column_letter(r.min_col)}{r.min_row}:{get_column_letter(r.max_col)}{r.max_row + k}"
        )
    ref_rows = ref_rows or [at_row - 1]
    max_col = ws.max_column
    for j in range(k):
        _copy_row_style(ws, ref_rows[j % len(ref_rows)], at_row + j, max_col)


def expand_sheets(wb, n_items: int):
    """按聚合物料数扩展 4 个 sheet 的数据行（模板默认 16 行）。"""
    # (sheet名, 插入位置, marks列延长区间, 样式参考行)
    specs = [
        ("报关发票 ", 32, ["B15:B33"], [31]),
        ("装箱单", 33, ["B16:B34"], [32]),
        ("报关合同 ", 32, [], [31]),
    ]
    for name, at_row, ext, ref in specs:
        ws = wb[name]
        k = n_items - 16
        _expand_rows(ws, at_row, k, ext, ref)
    # 报关单NEW：每物料 2 行，插在第 50 行（分隔线）之前；item/element 交替参考 48/49
    ws = wb["报关单NEW "]
    k = 2 * (n_items - 16)
    _expand_rows(ws, 50, k, [], [48, 49])


def _delete_rows_safe(ws, start_row: int, count: int):
    """整行删除数据区末尾的多余行，并正确处理合并单元格。

    - 完全在删除区内 → 移除
    - 跨越删除区边界 → 收缩到边界（保留删除区外部分）
    - 完全在删除区下方 → 整体上移 count 行（openpyxl 的 delete_rows 不会平移合并区）
    - 完全在删除区上方 → 不动
    """
    if count <= 0:
        return
    end = start_row + count - 1
    for rng in list(ws.merged_cells.ranges):
        rmin, rmax = rng.min_row, rng.max_row
        if rmax < start_row:
            continue                       # 完全在上方，不动
        if rmin >= start_row and rmax <= end:
            ws.merged_cells.remove(rng)    # 完全在删除区内，移除
            continue
        keep_min, keep_max = rmin, rmax
        if rmin < start_row <= rmax:
            keep_max = start_row - 1       # 跨越上边界，保留上方部分
        elif rmin <= end < rmax:
            keep_min = end + 1             # 跨越下边界，保留下方部分
        new_min, new_max = keep_min, keep_max
        if keep_min > end:
            new_min, new_max = keep_min - count, keep_max - count   # 删除区下方整体上移
        if new_min > new_max:
            continue
        ws.merged_cells.remove(rng)
        ws.merged_cells.add(
            f"{get_column_letter(rng.min_col)}{new_min}:{get_column_letter(rng.max_col)}{new_max}"
        )
    ws.delete_rows(start_row, count)


def delete_extra_rows(wb, n_items: int):
    """按实际导出的物料数 N 删除数据区末尾的多余整行。

    模板默认 16 行数据；当 N < 16 时删除 N 之后的多余行，合计行/页脚随之上移。
    发票/装箱单/报关合同/报关单NEW 都处理。N >= 16 时由 expand_sheets 扩展，无需删除。
    """
    if n_items >= 16:
        return
    # 报关发票：数据行 16..31（16 行），TOTAL 33
    _delete_rows_safe(wb["报关发票 "], 16 + n_items, 16 - n_items)
    # 装箱单：数据行 17..32，TOTAL 34
    _delete_rows_safe(wb["装箱单"], 17 + n_items, 16 - n_items)
    # 报关合同：数据行 16..31（16 行），TOTAL 33
    _delete_rows_safe(wb["报关合同 "], 16 + n_items, 16 - n_items)
    # 报关单NEW：每物料 2 行（项号+申报要素），数据 18..49，TOTAL 51
    _delete_rows_safe(wb["报关单NEW "], 18 + 2 * n_items, 2 * (16 - n_items))
    # 报关合同：不处理


# ──────────────────────────────────────────────────────────────
# 各 sheet 填充
# ──────────────────────────────────────────────────────────────
def fill_invoice(ws, dn, agg, totals):
    c = CONFIG
    k = max(0, len(agg) - 16)   # 扩展行偏移
    ws["B2"] = c["shipper_cn"]
    ws["B3"] = c["shipper_en"]
    ws["B4"] = c["shipper_addr"]
    ws["B8"] = None                        # 买方(To Messrs) 留空
    ws["G8"] = datetime.datetime.fromisoformat(dn["posting_date"][:10])
    ws["G9"] = None                        # 发票号 留空
    ws["G10"] = dn["name"]                 # 合同号
    ws["C13"] = c["transport_route"]
    ws["F13"] = c["payment_term"]
    for i in range(max(16, len(agg))):
        row = 16 + i
        if i < len(agg):
            it = agg[i]
            ws[f"C{row}"] = it["name_en"]                    # 品名(英文)
            ws[f"D{row}"] = it["qty"]
            ws[f"E{row}"] = uom_en(it["uom"])
            _n3(ws, f"F{row}", usd_price(it["bom_rate"]))    # 单价(USD)
            _n3(ws, f"G{row}", usd_price(it["bom_cost"]))    # 总金额(USD)
        else:
            for col in "CDEFG":
                ws[f"{col}{row}"] = None
    tr = 33 + k
    ws[f"C{tr}"] = "TOTAL:"
    ws[f"D{tr}"] = totals["qty"]
    ws[f"F{tr}"] = c["currency"]
    _n3(ws, f"G{tr}", usd_price(totals["bom_cost"]))
    nr = 35 + k
    ws[f"C{nr}"] = f"TOTAL PACKED IN {num_to_words(totals['cartons'])} CTNS"


def fill_packing(ws, dn, agg, totals):
    c = CONFIG
    k = max(0, len(agg) - 16)   # 扩展行偏移
    ws["B2"] = c["shipper_cn"]
    ws["B3"] = c["shipper_en"]
    ws["B4"] = c["shipper_addr"]
    ws["B8"] = None
    ws["H8"] = datetime.datetime.fromisoformat(dn["posting_date"][:10])
    ws["H9"] = None                        # 发票号 留空
    ws["H10"] = dn["name"]
    ws["C13"] = c["transport_route"]
    for i in range(max(16, len(agg))):
        row = 17 + i
        if i < len(agg):
            it = agg[i]
            ws[f"C{row}"] = it["name_en"]
            ws[f"D{row}"] = f"{int(it['qty'])} {uom_en(it['uom'])}"
            ws[f"E{row}"] = f"{int(it.get('cartons', 0))}CTNS"
            _n3(ws, f"F{row}", it.get("gross", 0.0))
            _n3(ws, f"H{row}", it.get("net", 0.0))
            _n3(ws, f"J{row}", it.get("measrs", 0.0))
        else:
            for col in "CDEFGHIJK":
                ws[f"{col}{row}"] = None
    tr = 34 + k
    ws[f"C{tr}"] = "TOTAL："
    ws[f"E{tr}"] = f"{totals['cartons']}\n CTNS"
    _n3(ws, f"F{tr}", totals["gross"])
    ws[f"G{tr}"] = "KGS"
    _n3(ws, f"H{tr}", totals["net"])
    ws[f"I{tr}"] = "KGS"
    _n3(ws, f"J{tr}", totals["measrs"])
    ws[f"K{tr}"] = "CBM"
    n36 = 36 + k
    ws[f"D{n36}"] = f"TOTAL PACKED IN {num_to_words(totals['cartons'])} CTNS"
    ws[f"D{n36+1}"] = f"TOTAL GROSS WEIGHT {totals['gross']:.3f}KGS"
    ws[f"D{n36+2}"] = f"TOTAL NET WEIGHT {totals['net']:.3f}KGS"
    ws[f"D{n36+3}"] = f"TOTAL MEASUREMENTS {totals['measrs']:.3f}M3"


def fill_contract(ws, dn, agg, totals):
    c = CONFIG
    k = max(0, len(agg) - 16)   # 扩展行偏移
    ws["B2"] = c["shipper_cn"]
    ws["B3"] = c["shipper_en"]
    ws["B4"] = c["shipper_addr"]
    ws["I7"] = dn["name"]
    ws["I8"] = fmt_date_contract(dn["posting_date"])
    ws["B9"] = None                        # 买方 留空
    for i in range(max(16, len(agg))):
        row = 16 + i
        if i < len(agg):
            it = agg[i]
            ws[f"B{row}"] = it["name_en"]
            ws[f"F{row}"] = it["qty"]
            ws[f"G{row}"] = uom_en(it["uom"])
            _n3(ws, f"H{row}", usd_price(it["bom_rate"]))
            ws[f"I{row}"] = f"/{uom_en(it['uom'])}"
            _n3(ws, f"J{row}", usd_price(it["bom_cost"]))
        else:
            for col in "BFGHIJ":
                ws[f"{col}{row}"] = None
    tr = 33 + k
    ws[f"B{tr}"] = "TOTAL:"
    ws[f"F{tr}"] = totals["qty"]
    _n3(ws, f"J{tr}", usd_price(totals["bom_cost"]))
    ws[f"E{39+k}"] = None                   # Time of Shipment(装运期) 留空
    ws[f"E{43+k}"] = None                   # TERMS OF PAYMENT 留空
    ws[f"H{48+k}"] = None                   # 底部 BUYERS 留空


def fill_declaration(ws, dn, agg, totals, country,
                     consignee_name="", consignee_addr="", declaration_unit="", production_unit=""):
    c = CONFIG
    country_en, country_cn = country
    k = 2 * max(0, len(agg) - 16)   # 扩展行偏移（每物料 2 行）
    ws["I2"] = None                        # 发票号 留空
    ws["A4"] = c["shipper_cn"]             # 境内发货人
    ws["G4"] = None                        # 出境关别 留空
    if consignee_name or consignee_addr:
        ws["C5"] = "\n".join(x for x in (consignee_name, consignee_addr) if x)   # 境外收货人：名称 + 地址
        ws["C5"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["G6"] = None                        # 运输方式 留空
    ws["A8"] = production_unit or None     # 生产销售单位（固定值）
    ws["G8"] = c["supervision_mode"]       # 监管方式
    ws["A10"] = dn["name"]                 # 合同协议号
    ws["D10"] = None                       # 贸易国（地区） 留空
    ws["G10"] = None                       # 运抵国（地区） 留空
    ws["J10"] = None                       # 指运港 留空
    ws["L10"] = None                       # 离境口岸 留空
    ws["A12"] = "CTNS"
    ws["D12"] = totals["cartons"]          # 件数
    _n3(ws, "E12", totals["gross"])        # 毛重
    _n3(ws, "G12", totals["net"])          # 净重
    ws["H12"] = c["trade_term"]            # 成交方式
    for i in range(max(16, len(agg))):
        row = 18 + i * 2
        if i < len(agg):
            it = agg[i]
            ws[f"A{row}"] = i + 1
            ws[f"B{row}"] = None                        # 商品编号(HS) 留空
            ws[f"C{row}"] = it["name_agg"]              # 中文名称（保留）
            ws[f"D{row}"] = it["name_en"]               # 英文名称
            ws[f"E{row}"] = f"{int(it['qty'])}{uom_cn(it['uom'])}"
            _n3(ws, f"G{row}", usd_price(it["bom_rate"]))     # 单价 = BOM成本 ÷ 汇率 × 加成
            _n3(ws, f"H{row}", usd_price(it["bom_cost"]))     # 总价 = BOM成本 ÷ 汇率 × 加成
            ws[f"I{row}"] = c["currency"]
            ws[f"J{row}"] = "中国"
            ws[f"K{row}"] = country_cn                  # 最终目的国（地区）
            ws[f"L{row}"] = c["origin_place"]           # 境内货源地
            ws[f"M{row}"] = None                        # 征免
            ws[f"C{row + 1}"] = None                    # 申报要素 留空
            # 申报要素行合并（A:B、C:N），超出模板 16 项时需新建
            for mg in (f"A{row+1}:B{row+1}", f"C{row+1}:N{row+1}"):
                if str(CellRange(mg)) not in [str(x) for x in ws.merged_cells.ranges]:
                    ws.merged_cells.add(mg)
        else:
            for col in "ABCDEFGHIJKLMN":
                ws[f"{col}{row}"] = None
            ws[f"C{row + 1}"] = None
    tr = 51 + k
    ws[f"A{tr}"] = f"TOTAL：{c['currency']} {usd_price(totals['bom_cost']):.3f}"
    if declaration_unit:                    # 申报单位：仅当传入才写入，否则保留模板原值
        ws[f"A{54+k}"] = f"申报单位  {declaration_unit}"


def main():
    ap = argparse.ArgumentParser(description="DN → 报关单据导出")
    ap.add_argument("--dn", required=True, help="DN 单号，如 DN-26-00063")
    ap.add_argument("--test", action="store_true", help="使用测试环境")
    ap.add_argument("--output", "-o", help="输出路径")
    ap.add_argument("--consignee", help="境外收货人预设: centrade/daneey/poland/custom（缺省按客户自动映射）")
    ap.add_argument("--consignee-name", help="境外收货人名称（覆盖预设/自定义）")
    ap.add_argument("--consignee-addr", help="境外收货人地址（覆盖预设/自定义）")
    ap.add_argument("--declaration-unit", help="申报单位（缺省保留模板原值）")
    args = ap.parse_args()

    env = "test" if args.test else "prod"
    key, sec = load_credentials(env)
    if not key or not sec or "your_" in key:
        print("✗ 凭证未配置（检查 .env）", file=sys.stderr)
        sys.exit(1)
    base = ENV_URLS[env]
    print(f"系统: {base}")

    dn = get_doc(base, key, sec, "Delivery Note", args.dn)
    if not dn:
        print(f"✗ 未找到 DN: {args.dn}", file=sys.stderr)
        sys.exit(1)

    agg = aggregate_items(dn)
    # 英文品名：并行翻译中文品名（去色后）
    dskey = load_deepseek_key()
    if dskey:
        def _do(it):
            return it, translate_zh_to_en(it["name_agg"], dskey)
        with ThreadPoolExecutor(max_workers=5) as ex:
            futs = [ex.submit(_do, it) for it in agg]
            for i, f in enumerate(as_completed(futs), 1):
                it, en = f.result()
                it["name_en"] = en
                print(f"  [{i}/{len(agg)}] {it['code_agg']} -> {en}", flush=True)
    else:
        for it in agg:
            it["name_en"] = it["name_agg"]
    print(f"聚合后物料 {len(agg)} 行:")
    for it in agg:
        print(f"  {it['code_agg']}  qty={it['qty']}  rate={it['rate']:.4f}  "
              f"amount={it['amount']:.2f}  bom_rate={it['bom_rate']:.4f}  bom_cost={it['bom_cost']:.2f}")
        print(f"    中文: {it['name_agg']}")
        print(f"    英文: {it['name_en']}")

    # 目的国
    country = resolve_country(dn.get("customer", ""))
    print(f"客户: {dn.get('customer')} ({dn.get('customer_name')}) → 目的国: {country}")

    # 境外收货人（CLI 指定或按客户自动映射）
    consignee = resolve_consignee(args, dn.get("customer", ""))
    if consignee:
        print(f"境外收货人: {consignee['name']} / {consignee['addr']}")
    else:
        print("境外收货人: 未匹配预设，保留模板原值（可用 --consignee 指定）")

    # 装箱数据：仅用用户确认的装箱组合(CARTON_GROUPS_BY_DN)，不用外箱子表(outer_box_summary/item_weight_cats)
    groups = CARTON_GROUPS_BY_DN.get(args.dn, []) or []
    if groups:
        pk = compute_packing(agg, groups)
        for it in agg:
            code = it["code_agg"]
            it["cartons"] = int(pk["cartons"].get(code, 0))
            it["gross"] = pk["gross"].get(code, 0.0)
            it["net"] = pk["net"].get(code, 0.0)
            it["measrs"] = pk["volume"].get(code, 0.0)
        total_cartons = pk["total_cartons"]
        total_gross = pk["total_gross"]
        total_net = pk["total_net"]
        total_volume = pk["total_volume"]
    else:
        for it in agg:
            it["cartons"] = it["gross"] = it["net"] = it["measrs"] = 0
        total_cartons = total_gross = total_net = total_volume = 0

    totals = {
        "qty": round(sum(float(i.get("qty") or 0) for i in dn.get("items", [])), 2),
        "amount": round(sum(float(i.get("amount") or 0) for i in dn.get("items", [])), 2),
        "bom_cost": round(sum(float(i.get("bom_cost") or 0) for i in dn.get("items", [])), 2),
        "cartons": total_cartons,
        "gross": total_gross,
        "net": total_net,
        "measrs": total_volume,
    }
    print(f"合计: qty={totals['qty']}  amount={totals['amount']}  bom_cost={totals['bom_cost']}  "
          f"cartons={total_cartons}  gross={total_gross}kg  net={total_net}kg  vol={total_volume}m³")

    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:      # 清除模板自带图片（中基抬头/印章/签名）
        ws._images = []
    expand_sheets(wb, len(agg))
    fill_invoice(wb["报关发票 "], dn, agg, totals)
    fill_packing(wb["装箱单"], dn, agg, totals)
    fill_contract(wb["报关合同 "], dn, agg, totals)
    fill_declaration(wb["报关单NEW "], dn, agg, totals, country,
                     consignee_name=(consignee or {}).get("name", ""),
                     consignee_addr=(consignee or {}).get("addr", ""),
                     declaration_unit=args.declaration_unit or "",
                     production_unit=CONFIG.get("production_unit", ""))

    # 按实际物料数删除数据区多余整行（发票/装箱单/报关单NEW，报关合同不动）
    delete_extra_rows(wb, len(agg))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = Path(args.output) if args.output else OUT_DIR / f"报关单据_{args.dn}.xlsx"
    wb.save(str(out))
    print(f"OK: {out}")


# ──────────────────────────────────────────────────────────────
# BOM 成本参照实现（与 EN delivery_plan/doc_event.py 的 _get_nd_cost 逻辑一致）
# 用于内胆(ND#) bom_rate 取 BOM内胆成本 cost_nd 列；皮壳(PK#) 取 cost_pk 列。
# ──────────────────────────────────────────────────────────────
BOM_COST_FILE = _DIR / "数据源" / "bom_cost_list_v2_2026-57-25-15-8.json.gz"


def _load_bom_cost_report() -> list[dict]:
    """读取 BOM 成本报表(gzip JSON)，返回 result 行列表（键=item_fg 成品编码）。"""
    import gzip
    import json as _json
    if not BOM_COST_FILE.is_file():
        print(f"✗ 未找到 BOM 成本报表: {BOM_COST_FILE}", file=sys.stderr)
        return []
    with gzip.open(BOM_COST_FILE, "rt", encoding="utf-8") as f:
        return _json.load(f).get("result", [])


def _get_nd_cost_ref(item_nd: str, bom_rows: list[dict]) -> tuple[float | None, dict | None]:
    """按内胆物料编码反向查找 BOM 内胆成本 cost_nd（参照实现）。

    匹配逻辑（membership-based，不依赖颜色词表）：
      1. 去掉 ND# 前缀，按 '-' 切分：段[0]=KS号，段[1]=尺寸；
      2. 遍历 BOM 成品行，item_fg 不去色直接按 '-' 切分，
         匹配条件 = 首段==KS号 且 尺寸出现在任一字段段
         （不用「最后一段=尺寸」，因成品可能带颜色/不带颜色，段位不固定）；
      3. 命中取该行 cost_nd；cost_nd 为 0/缺失则继续找有值的行。
    返回 (cost_nd, matched_row)；未命中返回 (None, None)。
    """
    if not item_nd or not item_nd.startswith("ND#"):
        return None, None
    parts = item_nd[3:].split("-")
    if len(parts) < 2:
        return None, None
    ks, size = parts[0], parts[1]
    for row in bom_rows:
        item_fg = (row.get("item_fg") or "").strip()
        if not item_fg:
            continue
        fg_parts = item_fg.split("-")
        if len(fg_parts) < 2:
            continue
        if fg_parts[0] == ks and size in fg_parts:
            cost_nd = row.get("cost_nd")
            if cost_nd is not None and cost_nd != 0:
                return cost_nd, row
    return None, None


if __name__ == "__main__":
    main()
