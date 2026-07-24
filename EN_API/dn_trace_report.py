#!/usr/bin/env python3
"""销售出库→物料移动追溯报表

数据链:
  Delivery Note (DN, docstatus=1)
    → DN Item.against_sales_order → Sales Order (SO)
      → Work Order.sales_order → Work Order (WO, status: In Process/Completed)
        → Stock Entry.work_order → Stock Entry (SE, purpose: Material Consumption for Manufacture, docstatus=1)
          → SE Item (s_warehouse, item_code, item_name, qty, uom)
  uv run python EN_API/dn_trace_report.py --month 2026-07
  uv run python EN_API/dn_trace_report.py --dn DN-2407-00001,DN-2407-00002
  uv run python EN_API/dn_trace_report.py --month 2026-07 --test
  uv run python EN_API/dn_trace_report.py --dn DN-xxx --output my_report.xlsx
"""
import argparse, json, os, sys
from calendar import monthrange
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths (script at EN_API/dn_trace_report.py → PROJECT_ROOT = repo root) ──
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "EN_API" / "out"

# ── Environments ──────────────────────────────────────
ENV_URLS = {
    "test": "https://ensh.vilavi.cn",
    "prod": "https://erpnext.vilavi.cn",
}

# ── Credentials ──────────────────────────────────────
def load_credentials(env_name: str = "prod") -> tuple[str, str]:
    """Load API credentials for given env. Priority: env vars > .env file."""
    if env_name == "test":
        key = os.environ.get("TEST_ERP_API_KEY", "")
        secret = os.environ.get("TEST_ERP_API_SECRET", "")
    else:
        key = os.environ.get("PROD_ERP_API_KEY", "")
        secret = os.environ.get("PROD_ERP_API_SECRET", "")
    # Fallback to generic
    if not key:
        key = os.environ.get("ERP_API_KEY", "")
    if not secret:
        secret = os.environ.get("ERP_API_SECRET", "")
    if key and secret:
        return key, secret

    env_file = PROJECT_ROOT / "EN_API" / ".env"
    if env_file.is_file():
        with open(env_file, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                _k, _, _v = line.partition("=")
                _k, _v = _k.strip(), _v.strip()
                if env_name == "test":
                    if _k == "TEST_ERP_API_KEY" and not key:
                        key = _v
                    elif _k == "TEST_ERP_API_SECRET" and not secret:
                        secret = _v
                else:
                    if _k == "PROD_ERP_API_KEY" and not key:
                        key = _v
                    elif _k == "PROD_ERP_API_SECRET" and not secret:
                        secret = _v
                # Generic fallback for either env
                if _k == "ERP_API_KEY" and not key:
                    key = _v
                elif _k == "ERP_API_SECRET" and not secret:
                    secret = _v
    return key, secret


# ── API helpers (simple requests, no session/adapter to avoid nginx 417) ──
import urllib.parse as _urlparse


def _build_url(resource: str, filters: list, fields: list[str],
               limit_start: int = 0, limit_page_length: int = 100) -> str:
    """Build ERPNext REST API path."""
    params = _urlparse.urlencode({
        "filters": json.dumps(filters, ensure_ascii=False),
        "fields": json.dumps(fields, ensure_ascii=False),
        "limit_start": str(limit_start),
        "limit_page_length": str(limit_page_length),
    })
    res = _urlparse.quote(resource, safe="")
    return f"/api/resource/{res}?{params}"


def _auth_headers(key: str, secret: str) -> dict:
    return {"Authorization": f"token {key}:{secret}"}


def api_get(path: str, base_url: str, key: str, secret: str) -> dict:
    """Simple GET request (no session, no keep-alive)."""
    url = f"{base_url}{path}"
    try:
        resp = requests.get(url, headers=_auth_headers(key, secret), timeout=60)
        resp.raise_for_status()
        return resp.json()
    except requests.RequestException as e:
        status = getattr(getattr(e, "response", None), "status_code", 0)
        print(f"  HTTP {status}: {url[:120]}", file=sys.stderr)
        return {"data": []}


def paginated_get(resource: str, base_url: str, key: str, secret: str,
                  filters: list, fields: list[str],
                  page_size: int = 100, max_pages: int = 100) -> list[dict]:
    """Fetch all pages from an ERPNext list endpoint."""
    all_data: list[dict] = []
    for page in range(max_pages):
        start = page * page_size
        path = _build_url(resource, filters, fields, start, page_size)
        result = api_get(path, base_url, key, secret)
        data = result.get("data", [])
        if not data:
            break
        all_data.extend(data)
    return all_data


def get_single(resource: str, docname: str, base_url: str,
               key: str, secret: str) -> dict:
    """Fetch a single document by name (includes child tables)."""
    path = "/api/resource/{}/{}".format(
        _urlparse.quote(resource, safe=''),
        _urlparse.quote(docname, safe=''),
    )
    return api_get(path, base_url, key, secret).get("data", {})


# ── Excel styles ──────────────────────────────────────
HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HDR_FONT = Font(bold=True, size=11, color="FFFFFF")
ALT_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")


def border() -> Border:
    return Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )


# ── Column definitions ────────────────────────────────
HEADERS = [
    ("DN单据号", 18), ("DN状态", 14), ("DN操作人", 18), ("DN日期", 14),
    ("SO单据号", 18), ("SO状态", 14), ("SO操作人", 18),
    ("WO单据号", 18), ("WO状态", 14), ("WO操作人", 18),
    ("SE单据号", 18), ("SE状态", 14), ("SE操作人", 18),
    ("发料仓", 14), ("物料编码", 24), ("物料名称", 30), ("数量", 10), ("单位", 8),
    ("是否面料", 10),
]


def safe(val, default=""):
    return val if val is not None else default


def build_row(dn: dict, so: dict | None, wo: dict | None,
              se: dict | None, item: dict | None,
              resolve=None, is_fabric=None) -> list:
    """Build one flat row from the 5-level hierarchy."""
    r = resolve or (lambda x: x or "")
    f = is_fabric or (lambda x: "")
    return [
        safe(dn.get("name")),
        safe(dn.get("status")),
        r(dn.get("owner")),
        safe(dn.get("posting_date", "")[:10]),
        safe(so.get("name") if so else ""),
        safe(so.get("status") if so else ""),
        r(so.get("owner") if so else ""),
        safe(wo.get("name") if wo else ""),
        safe(wo.get("status") if wo else ""),
        r(wo.get("owner") if wo else ""),
        safe(se.get("name") if se else ""),
        safe(se.get("docstatus") if se else ""),
        r(se.get("owner") if se else ""),
        safe(item.get("s_warehouse") if item else ""),
        safe(item.get("item_code") if item else ""),
        safe(item.get("item_name") if item else ""),
        item.get("qty", "") if item else "",
        safe(item.get("uom") if item else ""),
        f(item.get("item_code")) if item else "",
    ]


# ── Main ──────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="销售出库→物料移动追溯报表",
        epilog="示例:\n"
               "  %(prog)s --month 2026-07\n"
               "  %(prog)s --dn DN-2407-00001\n"
               "  %(prog)s --month 2026-07 --test",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--month", help="目标月份 (如 2026-07), 拉取该月所有已提交的 DN")
    group.add_argument("--dn", help="指定 DN 单号 (逗号分隔, 如 DN-xxx,DN-yyy)")
    parser.add_argument("--test", action="store_true", help="使用测试系统 (ensh.vilavi.cn)")
    parser.add_argument("--output", "-o", help="输出路径 (默认 EN_API/out/ 下)")
    args = parser.parse_args()

    # Credentials
    env = "test" if args.test else "prod"
    key, secret = load_credentials(env)
    if not key or not secret or "your_" in key:
        print("✗ API 凭证未配置。请先检查 EN_API/.env 文件", file=sys.stderr)
        sys.exit(1)

    base_url = ENV_URLS[env]
    print(f"系统: {base_url} (env: {env})")

    # ── Step 1: Fetch Delivery Notes ──
    if args.month:
        year_s, mon_s = args.month.split("-")
        year, mon = int(year_s), int(mon_s)
        last_day = monthrange(year, mon)[1]
        start_date = f"{args.month}-01"
        end_date = f"{args.month}-{last_day}"
        dn_filter: list = [
            ["docstatus", "=", "1"],
            ["posting_date", ">=", start_date],
            ["posting_date", "<=", end_date],
        ]
        dn_fields = ["name", "status", "owner", "posting_date"]
        print(f"拉取 DN (posting_date {start_date}~{end_date})...")
        delivery_notes = paginated_get("Delivery Note", base_url, key, secret,
                                        dn_filter, dn_fields)
        print(f"  → {len(delivery_notes)} 条")
    else:
        dn_names = [n.strip() for n in args.dn.split(",") if n.strip()]
        dn_filter = [["name", "in", dn_names]]
        dn_fields = ["name", "status", "owner", "posting_date"]
        print(f"拉取 DN (单号: {', '.join(dn_names)})...")
        delivery_notes = paginated_get("Delivery Note", base_url, key, secret,
                                        dn_filter, dn_fields)
        print(f"  → {len(delivery_notes)} 条")

    if not delivery_notes:
        print("无符合条件的 DN，退出。")
        sys.exit(0)

    dn_names = [d["name"] for d in delivery_notes]

    # ── Step 2: Fetch each DN's full doc to get items (child table) ──
    print(f"拉取 DN 完整明细 (含 items 子表)...")
    dn_so_map: dict[str, set[str]] = {}
    for i, dn in enumerate(delivery_notes):
        dn_name = dn["name"]
        full = get_single("Delivery Note", dn_name, base_url, key, secret)
        items = full.get("items", [])
        for item in items:
            so = item.get("against_sales_order", "")
            if so:
                dn_so_map.setdefault(dn_name, set()).add(so)
        if (i + 1) % 10 == 0:
            print(f"  {i + 1}/{len(delivery_notes)}")
    print(f"  {len(delivery_notes)}/{len(delivery_notes)} 完成")

    # Collect unique SO names
    all_so_names = sorted(set(
        so for sos in dn_so_map.values() for so in sos
    ))
    print(f"  关联 SO: {len(all_so_names)} 个")

    # ── Step 3: Fetch Sales Orders ──
    so_map: dict[str, dict] = {}
    if all_so_names:
        print("拉取 SO 详情...")
        so_list = paginated_get("Sales Order", base_url, key, secret,
            [["name", "in", all_so_names]],
            ["name", "status", "owner"],
        )
        for so in so_list:
            so_map[so["name"]] = so
        print(f"  → {len(so_list)} 条")

    # ── Step 4: Fetch Work Orders linked to SOs ──
    wo_by_so: dict[str, list[dict]] = {}
    if all_so_names:
        print("拉取 Work Order (status In Process/Completed)...")
        wo_list = paginated_get("Work Order", base_url, key, secret,
            [
                ["sales_order", "in", all_so_names],
                ["status", "in", ["In Process", "Completed"]],
            ],
            ["name", "status", "owner", "sales_order", "production_item"],
        )
        for wo in wo_list:
            so_name = wo.get("sales_order", "")
            if so_name:
                wo_by_so.setdefault(so_name, []).append(wo)
        print(f"  → {len(wo_list)} 条工单")

    # Collect unique WO names
    all_wo_names = sorted(set(
        wo["name"] for wos in wo_by_so.values() for wo in wos
    ))

    # ── Step 5: Fetch Stock Entries (Material Consumption for Manufacture) ──
    # No date filter — find ALL SEs of this type, then filter by WO list in Python
    se_by_wo: dict[str, list[dict]] = {}
    if all_wo_names:
        wo_set = set(all_wo_names)
        print(f"拉取所有 Stock Entry (Material Consumption for Manufacture, docstatus<2)...")
        se_list = paginated_get("Stock Entry", base_url, key, secret,
            [
                ["stock_entry_type", "=", "Material Consumption for Manufacture"],
                ["docstatus", "<", "2"],
            ],
            ["name", "owner", "work_order", "docstatus"],
            page_size=200,
        )
        for se in se_list:
            wo_name = se.get("work_order", "")
            if wo_name in wo_set:
                se_by_wo.setdefault(wo_name, []).append(se)
        print(f"  → 系统共 {len(se_list)} 条, 其中 {sum(len(v) for v in se_by_wo.values())} 条关联到当前 WO")

    # Collect unique SE names
    all_se_names = sorted(set(
        se["name"] for ses in se_by_wo.values() for se in ses
    ))

    # ── Step 6: Fetch SE Items via individual SE docs ──
    se_items_map: dict[str, list[dict]] = {}
    if all_se_names:
        print(f"拉取 SE 发料明细 ({len(all_se_names)} 个)...")
        for i, se_name in enumerate(all_se_names):
            full = get_single("Stock Entry", se_name, base_url, key, secret)
            items = full.get("items", [])
            for item in items:
                se_items_map.setdefault(se_name, []).append(item)
            if (i + 1) % 20 == 0:
                print(f"  {i + 1}/{len(all_se_names)}")
        total = sum(len(v) for v in se_items_map.values())
        print(f"  {len(all_se_names)}/{len(all_se_names)} 完成, 共 {total} 条明细")

    # ── Step 7: Detect fabric items ──
    print("判断物料是否为面料...")
    # Collect all unique item_codes from SE items
    se_item_codes: set[str] = set()
    for se_name, items in se_items_map.items():
        for item in items:
            code = item.get("item_code", "")
            if code:
                se_item_codes.add(code)

    # Fetch item → item_group mapping
    item_group_map: dict[str, str] = {}
    if se_item_codes:
        items_data = paginated_get("Item", base_url, key, secret,
            [["name", "in", sorted(se_item_codes)]],
            ["name", "item_group"],
            page_size=200,
        )
        for it in items_data:
            ig = it.get("item_group", "")
            if ig:
                item_group_map[it["name"]] = ig
        print(f"  → {len(item_group_map)} 个物料查到物料组")

    # Fetch ALL Item Groups and build fabric descendant set
    all_groups = paginated_get("Item Group", base_url, key, secret,
        [], ["name", "item_group_name", "parent_item_group", "is_group"],
        page_size=500,
    )
    # Build parent→children map
    group_children: dict[str, list[dict]] = {}
    group_info: dict[str, dict] = {}
    for g in all_groups:
        group_info[g["name"]] = g
        parent = g.get("parent_item_group", "")
        group_children.setdefault(parent, []).append(g)

    # Find 面料 root and collect all descendants
    fabric_groups: set[str] = set()
    for g in all_groups:
        if g.get("item_group_name", "") == "面料":
            # DFS collect all descendants
            stack = [g["name"]]
            while stack:
                cur = stack.pop()
                fabric_groups.add(cur)
                for child in group_children.get(cur, []):
                    stack.append(child["name"])
            break
    print(f"  → 面料组及子组: {len(fabric_groups)} 个")

    def is_fabric(item_code: str) -> str:
        ig = item_group_map.get(item_code, "")
        return "是" if ig in fabric_groups else "否"

    # ── Step 8: Resolve user names ──
    print("解析操作人姓名...")
    all_owners: set[str] = set()
    for dn in delivery_notes:
        if dn.get("owner"): all_owners.add(dn["owner"])
    for so in so_map.values():
        if so.get("owner"): all_owners.add(so["owner"])
    for wos in wo_by_so.values():
        for wo in wos:
            if wo.get("owner"): all_owners.add(wo["owner"])
    for ses in se_by_wo.values():
        for se in ses:
            if se.get("owner"): all_owners.add(se["owner"])

    user_map: dict[str, str] = {}
    if all_owners:
        owner_list = sorted(all_owners)
        users = paginated_get("User", base_url, key, secret,
            [["name", "in", owner_list]],
            ["name", "full_name"],
            page_size=200,
        )
        for u in users:
            name = u.get("full_name") or u.get("name", "")
            if name:
                user_map[u["name"]] = name
        # Fetch any missing owners individually (batch size limit may have missed some)
        for owner in owner_list:
            if owner not in user_map:
                u = get_single("User", owner, base_url, key, secret)
                name = u.get("full_name") or owner
                user_map[owner] = name
        print(f"  → {len(user_map)} 个用户")

    def resolve(owner: str | None) -> str:
        return user_map.get(owner, owner) if owner else ""

    # ── Step 8: Flatten into rows (separate KS 成品 from non-KS 半成品) ──
    print("整理数据...")
    detail_rows: list[list] = []
    fg_rows: list[list] = []
    _BR = lambda *a: build_row(*a, resolve=resolve, is_fabric=is_fabric)
    for dn in delivery_notes:
        dn_name = dn["name"]
        sos = dn_so_map.get(dn_name, set())
        if not sos:
            detail_rows.append(_BR(dn, None, None, None, None))
            continue
        for so_name in sorted(sos):
            so = so_map.get(so_name)
            if not so:
                detail_rows.append(_BR(dn, None, None, None, None))
                continue
            wos = wo_by_so.get(so_name, [])
            if not wos:
                detail_rows.append(_BR(dn, so, None, None, None))
                continue
            for wo in wos:
                is_ks = (wo.get("production_item", "") or "").startswith("KS")
                if is_ks:
                    # 成品工单(KS开头): 不涉及耗用, 记入成品表
                    fg_rows.append(_BR(dn, so, wo, None, None))
                    continue
                ses = se_by_wo.get(wo["name"], [])
                if not ses:
                    detail_rows.append(_BR(dn, so, wo, None, None))
                    continue
                for se in ses:
                    items = se_items_map.get(se["name"], [])
                    if not items:
                        detail_rows.append(_BR(dn, so, wo, se, None))
                        continue
                    for item in items:
                        detail_rows.append(_BR(dn, so, wo, se, item))

    print(f"  → 追溯明细: {len(detail_rows)} 行, 成品工单: {len(fg_rows)} 行")

    # ── Generate Excel ──
    def _write_sheet(wb, title, data, ncols):
        """Write a sheet with header + data rows."""
        ws = wb.create_sheet(title) if title != "追溯明细" else wb.active
        ws.title = title
        for c, (h, w) in enumerate(HEADERS[:ncols], 1):
            cl = ws.cell(row=1, column=c, value=h)
            cl.font, cl.fill = HDR_FONT, HDR_FILL
            cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cl.border = border()
            ws.column_dimensions[get_column_letter(c)].width = w
        alt_idx = 0
        for row_idx, vals in enumerate(data, 2):
            use_alt = (alt_idx % 2 == 1)
            for c in range(1, ncols + 1):
                v = vals[c - 1] if c <= len(vals) else ""
                cl = ws.cell(row=row_idx, column=c, value=v)
                cl.border = border()
                cl.alignment = Alignment(vertical="center", wrap_text=True)
                if use_alt:
                    cl.fill = ALT_FILL
            if row_idx == 2:
                prev_dn = vals[0]
            elif vals[0] != prev_dn:
                alt_idx += 1
                prev_dn = vals[0]
        ws.freeze_panes = "A2"
        return ws

    print("生成 Excel...")
    wb = Workbook()
    # Sheet 1: 追溯明细 (19列, 含SE/SE Item + 是否面料)
    _write_sheet(wb, "追溯明细", detail_rows, 19)
    # Sheet 2: 成品工单 (10列, DN→SO→WO, KS开头不涉及耗用)
    _write_sheet(wb, "成品工单", fg_rows, 10)
    print(f"  Sheet1 追溯明细: {len(detail_rows)} 行, Sheet2 成品工单: {len(fg_rows)} 行")

    # Output path
    if args.output:
        output = Path(args.output)
    else:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        suffix = args.month or "custom"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = DATA_DIR / f"{suffix}_DN追溯报表_{ts}.xlsx"

    wb.save(str(output))
    print(f"OK: {output}")
    print(f"Sheet1 追溯明细: {len(detail_rows)} 行, Sheet2 成品工单: {len(fg_rows)} 行, 共 {len(HEADERS)} 列(含是否面料)")


if __name__ == "__main__":
    main()
