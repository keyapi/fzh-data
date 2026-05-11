# -*- coding: utf-8 -*-
"""
从 EN 重量模板（同事手工填写国外分公司重尺数据）按 SKU 前缀匹配到赛狐全量 SKU，
生成赛狐「商品重尺」导入文件。

匹配规则：重量模板物料编码去除 ZLMB# 前缀 → 3 段键（品类-面料-尺寸）；
赛狐 SKU（4 段，末段颜色）取前 3 段作为键。键相同则匹配。

长宽高三者全部有值才写入；装箱量缺省时若长宽高有值则默认 1。

数据源默认从 ./数据源/ 读取。

使用：
  python build_saihu_weight_import.py
  python build_saihu_weight_import.py --weight-data "path/彭建.xlsx"
  python build_saihu_weight_import.py --saihu-export "path/商品导出.xlsx"
  python build_saihu_weight_import.py --template "path/模板.xlsx"
"""

from __future__ import annotations

import argparse
import math
import os
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

_DIR = Path(__file__).resolve().parent
os.chdir(_DIR)
_ROOT = _DIR.parent

_DEFAULT_WEIGHT_DIR = _DIR / "数据源"
_DEFAULT_OUT_DIR = _DIR / "out"

# ── 彭建重量模板 列名 ──────────────────────────────
COL_WT_CODE = "物料编码"
COL_WT_STYLE = "款式ID"
COL_WT_GROUP = "物料组"
COL_WT_NAME = "物料名称"
COL_WT_GROSS_WEIGHT = "国外分公司成品包装后实重(g)"
COL_WT_PKG_L = "国外分公司成品包装长(cm)"
COL_WT_PKG_W = "国外分公司成品包装宽(cm)"
COL_WT_PKG_H = "国外分公司成品包装高(cm)"
COL_WT_BOX_QTY = "装箱量"

# ── 赛狐商品导出 列名 ──────────────────────────────
COL_SX_SKU = "SKU"
COL_SX_SPU = "spu"

# ── 赛狐模板输出列（工作表名必须为 商品）────────────
SHEET_MAIN = "商品"
COL_OUT_SKU = "*SKU"
COL_OUT_SPEC_L = "商品规格长(cm)"
COL_OUT_SPEC_W = "商品规格宽(cm)"
COL_OUT_SPEC_H = "商品规格高(cm)"
COL_OUT_WEIGHT = "商品重量"
COL_OUT_WEIGHT_UNIT = "商品重量单位"
COL_OUT_BOX_L = "箱规长(cm)"
COL_OUT_BOX_W = "箱规宽(cm)"
COL_OUT_BOX_H = "箱规高(cm)"
COL_OUT_BOX_WEIGHT = "单箱重量(kg)"
COL_OUT_BOX_QTY = "单箱数量(PCS)"
COL_OUT_PKG_L = "商品包装规格长(cm)"
COL_OUT_PKG_W = "商品包装规格宽(cm)"
COL_OUT_PKG_H = "商品包装规格高(cm)"
COL_OUT_PKG_WEIGHT = "商品包装重量"
COL_OUT_PKG_WEIGHT_UNIT = "商品包装重量单位"

# 暂不填充的列
_OUT_SKIP_COLS = frozenset({
    COL_OUT_SPEC_L, COL_OUT_SPEC_W, COL_OUT_SPEC_H,
    COL_OUT_WEIGHT, COL_OUT_WEIGHT_UNIT,
})

# ── 问题报告 sheet 名 ──────────────────────────────
SHEET_SUMMARY = "汇总"
SHEET_UNFILLED_DETAIL = "未填充SKU明细"
SHEET_WT_MISSING_LWH = "重量模板_长宽高不全(去重)"
SHEET_WT_NO_MATCH = "重量模板_未匹配任何赛狐SKU"
SHEET_SX_NO_MATCH = "赛狐SKU_无对应重量模板"
SHEET_BOX_DEFAULTED = "装箱量_已默认1(去重)"

# ── 匹配键 ─────────────────────────────────────────
_WEIGHT_KEY_PREFIX = "ZLMB#"


def _saihu_match_key(sku: Any) -> str | None:
    """赛狐 SKU 前 3 段为匹配键（≥4 段取前 3，=3 段取全串，<3 返回 None）。"""
    s = _norm(sku)
    if not s:
        return None
    parts = s.split("-")
    if len(parts) >= 4:
        return "-".join(parts[:3])
    if len(parts) == 3:
        return s
    return None


def _weight_match_key(code: Any) -> str | None:
    """重量模板物料编码去除 ZLMB# 前缀后为匹配键。"""
    s = _norm(code)
    if not s:
        return None
    if s.upper().startswith(_WEIGHT_KEY_PREFIX.upper()):
        s = s[len(_WEIGHT_KEY_PREFIX):]
    return s if s else None


# ── 工具函数 ───────────────────────────────────────
def _norm(x: Any) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    return s if s and s.lower() != "nan" else ""


def _to_float(x: Any) -> float | None:
    if x is None:
        return None
    if isinstance(x, (float, int)) and bool(pd.isna(x)):
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _value_present(x: Any) -> bool:
    if x is None:
        return False
    if isinstance(x, float) and (math.isnan(x) or pd.isna(x)):
        return False
    return True


def _safe_div(a: float, b: float) -> float:
    if b == 0.0:
        return a
    return a / b


# ── 数据读取 ───────────────────────────────────────
def _read_weight_data(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    sheet = "物料" if "物料" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet, header=0)
    required = {COL_WT_CODE}
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"重量模板缺列 {missing!r}，实际: {list(df.columns)[:30]}")
    return df


def _read_saihu_export(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    sheet = SHEET_MAIN if SHEET_MAIN in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet, header=0)
    for c in (COL_SX_SKU, COL_SX_SPU):
        if c not in df.columns:
            raise ValueError(f"赛狐导出缺列 {c!r}，实际: {list(df.columns)[:20]}")
    return df[[COL_SX_SKU, COL_SX_SPU]].copy()


# ── 字段计算 ───────────────────────────────────────
@dataclass
class MappedRow:
    """一行赛狐 SKU 对应的重尺数据（可能为空，表示未匹配或校验未通过）。"""

    sku: str
    box_l: float | None = None
    box_w: float | None = None
    box_h: float | None = None
    box_weight_kg: float | None = None
    box_qty: int | None = None
    pkg_l: float | None = None
    pkg_w: float | None = None
    pkg_h: float | None = None
    pkg_weight: float | None = None

    def to_template_dict(self) -> dict[str, Any]:
        return {
            COL_OUT_SKU: self.sku,
            COL_OUT_BOX_L: self.box_l,
            COL_OUT_BOX_W: self.box_w,
            COL_OUT_BOX_H: self.box_h,
            COL_OUT_BOX_WEIGHT: self.box_weight_kg,
            COL_OUT_BOX_QTY: self.box_qty,
            COL_OUT_PKG_L: self.pkg_l,
            COL_OUT_PKG_W: self.pkg_w,
            COL_OUT_PKG_H: self.pkg_h,
            COL_OUT_PKG_WEIGHT: self.pkg_weight,
            COL_OUT_PKG_WEIGHT_UNIT: "g" if self.pkg_weight is not None else None,
        }


def _build_mapped_rows(
    saihu_df: pd.DataFrame,
    weight_df: pd.DataFrame,
    issues: list[dict[str, Any]],
) -> tuple[list[MappedRow], dict[str, list[dict[str, Any]]]]:
    """核心匹配与映射。"""

    # 重量模板 → dict[key, row]
    wt_map: dict[str, Any] = {}
    wt_dup_warn: list[str] = []
    for _i, row in weight_df.iterrows():
        key = _weight_match_key(row[COL_WT_CODE])
        if not key:
            continue
        if key in wt_map:
            wt_dup_warn.append(f"重量模板键重复: {key!r}，保留末行")
        wt_map[key] = row
    for w in wt_dup_warn:
        issues.append({"类型": "重量模板键重复", "说明": w})

    # ── 逐行匹配 ──
    wt_matched_keys: set[str] = set()
    # 按重量模板去重的集合：记录每个模板被多少 SKU 引用
    wt_missing_lwh_set: dict[str, dict[str, Any]] = {}  # key -> info
    wt_missing_lwh_sku_count: dict[str, int] = {}        # key -> affected SKU count
    wt_box_defaulted_set: dict[str, dict[str, Any]] = {} # key -> info

    mapped: list[MappedRow] = []
    sx_unmatched: list[dict[str, Any]] = []
    unfilled_detail: list[dict[str, Any]] = []          # 每个未填充 SKU 一行

    for _i, sx_row in saihu_df.iterrows():
        sku = _norm(sx_row[COL_SX_SKU])
        spu = _norm(sx_row.get(COL_SX_SPU, ""))
        key = _saihu_match_key(sku)

        # ── 情况 1: 无法匹配 ──
        if not key or key not in wt_map:
            mapped.append(MappedRow(sku=sku))
            reason = "SKU 少于 3 段，无法提取匹配键" if not key else f"匹配键 {key!r} 在重量模板中不存在"
            unfilled_detail.append({
                "SKU": sku, "spu": spu, "匹配键": key or "(无)",
                "未填充原因": reason,
            })
            if key:
                sx_unmatched.append({
                    "SKU": sku, "spu": spu, "匹配键": key,
                    "说明": "该匹配键在彭建重量模板中不存在，无法获取重尺数据",
                })
            continue

        wt_row = wt_map[key]
        wt_matched_keys.add(key)

        l = _to_float(wt_row.get(COL_WT_PKG_L))
        w = _to_float(wt_row.get(COL_WT_PKG_W))
        h = _to_float(wt_row.get(COL_WT_PKG_H))

        # ── 情况 2: 匹配成功但长宽高不全 ──
        if not (_value_present(l) and _value_present(w) and _value_present(h)):
            missing_cols = []
            if not _value_present(l):
                missing_cols.append(COL_WT_PKG_L)
            if not _value_present(w):
                missing_cols.append(COL_WT_PKG_W)
            if not _value_present(h):
                missing_cols.append(COL_WT_PKG_H)
            reason = f"已匹配重量模板 {_norm(wt_row[COL_WT_CODE])}，但缺: {', '.join(missing_cols)}"
            unfilled_detail.append({
                "SKU": sku, "spu": spu, "匹配键": key,
                "未填充原因": reason,
            })
            if key not in wt_missing_lwh_set:
                wt_missing_lwh_set[key] = {
                    "物料编码": _norm(wt_row[COL_WT_CODE]),
                    "款式ID": _norm(wt_row.get(COL_WT_STYLE, "")),
                    "物料组": _norm(wt_row.get(COL_WT_GROUP, "")),
                    "匹配键": key,
                    "缺失列": ", ".join(missing_cols),
                }
                wt_missing_lwh_sku_count[key] = 0
            wt_missing_lwh_sku_count[key] += 1
            mapped.append(MappedRow(sku=sku))
            continue

        # ── 情况 3: 匹配成功且长宽高齐全 → 正常填充 ──
        gross = _to_float(wt_row.get(COL_WT_GROSS_WEIGHT))
        box_qty_raw = _to_float(wt_row.get(COL_WT_BOX_QTY))

        box_qty: int = 1
        if _value_present(box_qty_raw):
            box_qty = max(1, int(box_qty_raw))
        else:
            if key not in wt_box_defaulted_set:
                wt_box_defaulted_set[key] = {
                    "物料编码": _norm(wt_row[COL_WT_CODE]),
                    "款式ID": _norm(wt_row.get(COL_WT_STYLE, "")),
                    "物料组": _norm(wt_row.get(COL_WT_GROUP, "")),
                    "匹配键": key,
                    "已默认装箱量": 1,
                    "说明": "彭建未填写装箱量，因长宽高有值，按规则自动设为 1",
                }

        box_weight: float | None = None
        if _value_present(gross) and gross is not None:
            box_weight = round(gross * box_qty / 1000.0, 8)

        pkg_weight: float | None = gross
        pkg_h_val = round(_safe_div(h, float(box_qty)), 8) if h is not None else None

        mapped.append(MappedRow(
            sku=sku,
            box_l=round(l, 8) if l is not None else None,
            box_w=round(w, 8) if w is not None else None,
            box_h=round(h, 8) if h is not None else None,
            box_weight_kg=box_weight,
            box_qty=box_qty,
            pkg_l=round(l, 8) if l is not None else None,
            pkg_w=round(w, 8) if w is not None else None,
            pkg_h=pkg_h_val,
            pkg_weight=pkg_weight,
        ))

    # ── 收集重量模板未匹配的（在 wt_map 中但没被任何赛狐 SKU 引用）──
    wt_unmatched: list[dict[str, Any]] = []
    for key, row in wt_map.items():
        if key not in wt_matched_keys:
            wt_unmatched.append({
                "物料编码": _norm(row[COL_WT_CODE]),
                "款式ID": _norm(row.get(COL_WT_STYLE, "")),
                "物料组": _norm(row.get(COL_WT_GROUP, "")),
                "物料名称": _norm(row.get(COL_WT_NAME, "")),
                "匹配键": key,
                "说明": "该重量模板的匹配键在赛狐商品导出中找不到任何 SKU（可能赛狐尚未建档，或键不匹配）",
            })

    # ── 汇总 ──
    n_sx_unmatched = len(sx_unmatched)
    n_lwh_incomplete = len(unfilled_detail) - n_sx_unmatched
    n_filled = sum(1 for m in mapped if m.box_l is not None)
    n_wt_lwh_keys = len(wt_missing_lwh_set)
    n_wt_box_def_keys = len(wt_box_defaulted_set)

    issues.insert(0, {
        "类型": "汇总",
        "赛狐SKU总数": len(saihu_df),
        "已填充重尺": n_filled,
        "未填充合计": len(saihu_df) - n_filled,
        "未填充_无匹配": n_sx_unmatched,
        "未填充_匹配但缺长宽高": n_lwh_incomplete,
        "重量模板总数": len(wt_map),
        "重量模板_长宽高不全(去重模板数)": n_wt_lwh_keys,
        "重量模板_未匹配任何赛狐SKU": len(wt_unmatched),
        "装箱量_已默认1(去重模板数)": n_wt_box_def_keys,
    })

    # 去重后加「影响SKU数」列
    wt_lwh_report: list[dict[str, Any]] = []
    for key, info in wt_missing_lwh_set.items():
        info["影响SKU数"] = wt_missing_lwh_sku_count.get(key, 0)
        wt_lwh_report.append(info)

    return mapped, {
        SHEET_UNFILLED_DETAIL: unfilled_detail,
        SHEET_WT_MISSING_LWH: wt_lwh_report,
        SHEET_WT_NO_MATCH: wt_unmatched,
        SHEET_SX_NO_MATCH: sx_unmatched,
        SHEET_BOX_DEFAULTED: list(wt_box_defaulted_set.values()),
    }


# ── 写入输出 ────────────────────────────────────────
def _write_output(
    template_path: Path,
    out_path: Path,
    rows: list[MappedRow],
) -> None:
    wb = load_workbook(template_path)
    if SHEET_MAIN not in wb.sheetnames:
        raise ValueError(f"模板缺工作表 {SHEET_MAIN!r}: {wb.sheetnames}")
    ws = wb[SHEET_MAIN]

    # 获取模板表头顺序
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # 清空数据行（保留表头）
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for r, row in enumerate(rows, start=2):
        d = row.to_template_dict()
        for c, h in enumerate(headers, start=1):
            v = d.get(h)
            if h in _OUT_SKIP_COLS:
                ws.cell(row=r, column=c, value=None)
            elif v is None or (isinstance(v, float) and pd.isna(v)):
                ws.cell(row=r, column=c, value=None)
            else:
                ws.cell(row=r, column=c, value=v)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()


def _write_issues(out_path: Path, issues: list[dict], reports: dict[str, list[dict]]) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        # 汇总（单行多列）
        if issues:
            pd.DataFrame(issues).to_excel(w, sheet_name=SHEET_SUMMARY, index=False)

        # 各分类报告
        sheets_order = [
            SHEET_UNFILLED_DETAIL,
            SHEET_WT_MISSING_LWH,
            SHEET_WT_NO_MATCH,
            SHEET_SX_NO_MATCH,
            SHEET_BOX_DEFAULTED,
        ]
        for s in sheets_order:
            data = reports.get(s, [])
            if data:
                pd.DataFrame(data).to_excel(w, sheet_name=s, index=False)
            else:
                pd.DataFrame({"说明": ["（无数据）"]}).to_excel(w, sheet_name=s, index=False)


def _print_summary(
    n_saihu: int,
    n_filled: int,
    n_sx_unmatched: int,
    n_lwh_incomplete: int,
    n_wt_lwh_keys: int,
    n_wt_unmatched: int,
    n_box_def: int,
    reports: dict[str, list[dict]],
    out_main: Path,
    out_issues: Path,
) -> None:
    n_unfilled = n_saihu - n_filled
    n_total_unfilled_detail = len(reports.get(SHEET_UNFILLED_DETAIL, []))
    print("=" * 60)
    print("赛狐 商品重尺导入 — 处理完成")
    print("=" * 60)
    print(f"主结果:         {out_main}")
    print(f"问题报告:       {out_issues}")
    print(f"赛狐 SKU 总数:   {n_saihu}")
    print(f"已填充重尺:       {n_filled}")
    print(f"未填充合计:       {n_unfilled}")
    print(f"  其中 未匹配:    {n_sx_unmatched} 行 (赛狐 SKU 找不到对应重量模板)")
    print(f"  其中 缺长宽高:  {n_lwh_incomplete} 行 (匹配到但重量模板长宽高不全)")
    print(f"---")
    print(f"重量模板长宽高不全(去重): {n_wt_lwh_keys} 个模板")
    print(f"重量模板未匹配任何赛狐SKU: {n_wt_unmatched} 个模板")
    print(f"装箱量已默认1(去重):       {n_box_def} 个模板")
    print(f"---")
    print(f"详见 {out_issues.name}:")
    print(f"  「{SHEET_SUMMARY}」- 总览")
    print(f"  「{SHEET_UNFILLED_DETAIL}」- 每个未填充 SKU 的原因 ({n_total_unfilled_detail} 行)")
    print(f"  「{SHEET_WT_MISSING_LWH}」- 长宽高不全的重量模板 (去重, 含影响SKU数)")
    print(f"  「{SHEET_WT_NO_MATCH}」- 未匹配的重量模板")
    print(f"  「{SHEET_SX_NO_MATCH}」- 未匹配的赛狐 SKU")
    print(f"  「{SHEET_BOX_DEFAULTED}」- 装箱量被默认设为 1 的模板")
    print("=" * 60)


# ── 主入口 ─────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(
        description="从 EN 重量模板（手工填写的重尺数据）生成赛狐商品重尺导入文件"
    )
    ap.add_argument(
        "--weight-data",
        type=Path,
        default=None,
        help="彭建已填的重量数据 xlsx（默认在 ./数据源/ 中自动选择含「重尺」的最新文件）",
    )
    ap.add_argument(
        "--saihu-export",
        type=Path,
        default=None,
        help="赛狐商品导出 xlsx（默认在 ./数据源/ 中自动选择含「商品导出」「填重尺」的最新文件）",
    )
    ap.add_argument(
        "--template",
        type=Path,
        default=None,
        help="赛狐商品重尺导入模板 xlsx（默认在 ./数据源/ 中自动选择含「模板」「商品重尺」的最新文件）",
    )
    ap.add_argument("--out-dir", type=Path, default=_DEFAULT_OUT_DIR, help="输出目录")
    args = ap.parse_args()

    data_dir = _DEFAULT_WEIGHT_DIR

    def _find_file(*keywords: str) -> Path:
        if not data_dir.is_dir():
            raise FileNotFoundError(f"数据目录不存在: {data_dir}")
        cands = [
            p for p in data_dir.iterdir()
            if p.is_file() and p.suffix.lower() == ".xlsx"
            and not p.name.startswith("~$")
            and all(kw in p.name for kw in keywords)
        ]
        if not cands:
            raise FileNotFoundError(
                f"{data_dir} 下找不到含 {keywords!r} 的 xlsx（已排除 ~$ 锁文件）"
            )
        return max(cands, key=lambda x: x.stat().st_mtime)

    wt_path = args.weight_data or _find_file("重尺数据")
    saihu_path = args.saihu_export or _find_file("商品导出", "填重尺")
    tpl_path = args.template or _find_file("模板", "商品重尺")
    out_dir: Path = args.out_dir

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_main = out_dir / f"赛狐_商品重尺导入_{ts}.xlsx"
    out_issues = out_dir / f"重尺处理_问题报告_{ts}.xlsx"

    print(f"重量数据: {wt_path}")
    print(f"赛狐导出: {saihu_path}")
    print(f"模板:     {tpl_path}")

    saihu_df = _read_saihu_export(saihu_path)
    weight_df = _read_weight_data(wt_path)

    issues: list[dict[str, Any]] = []
    rows, reports = _build_mapped_rows(saihu_df, weight_df, issues)

    _write_output(tpl_path, out_main, rows)
    _write_issues(out_issues, issues, reports)

    n_filled = sum(1 for r in rows if r.box_l is not None)
    n_sx_unmatched = len(reports.get(SHEET_SX_NO_MATCH, []))
    n_unfilled = len(saihu_df) - n_filled
    n_lwh_incomplete = n_unfilled - n_sx_unmatched
    n_wt_lwh_keys = len(reports.get(SHEET_WT_MISSING_LWH, []))
    n_wt_unmatched = len(reports.get(SHEET_WT_NO_MATCH, []))
    n_box_def = len(reports.get(SHEET_BOX_DEFAULTED, []))
    _print_summary(
        len(saihu_df), n_filled,
        n_sx_unmatched, n_lwh_incomplete,
        n_wt_lwh_keys, n_wt_unmatched, n_box_def,
        reports, out_main, out_issues,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
