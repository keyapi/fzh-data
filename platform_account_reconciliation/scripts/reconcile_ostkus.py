#!/usr/bin/env python3
"""OSTKUS 账期 vs EN/Tongtool Order 费用级对账。

用法:
  uv run python platform_account_reconciliation/scripts/reconcile_ostkus.py \
      --account "D:/Work/尹/OSTKUS-2026-07-01.xlsx" \
      --account "D:/Work/尹/OSTKUS-2026-07-16.xlsx" \
      --out "D:/Work/尹/OSTKUS费用级核对.xlsx"

默认从 EN_API/.env 读取生产 ERPNext 凭证，并通过只读 REST 拉取 Tongtool Order。
凭证缺失或显式传 --no-en 时只生成账期侧明细，不调用 EN。
"""
from __future__ import annotations

import argparse
import json
import math
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
ENV_FILE = ROOT / "EN_API" / ".env"
EN_BASE = "https://erpnext.vilavi.cn"


def base_id(value: object) -> str:
    match = re.search(r"\d+", str(value or ""))
    return match.group(0) if match else str(value or "").strip()


def num(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def uniq_join(values, sep: str = " | ") -> str:
    seen: list[str] = []
    for v in values:
        text = str(v or "").strip()
        if text and text not in seen:
            seen.append(text)
    return sep.join(seen)


def return_category(desc: object) -> str:
    text = str(desc or "")
    if text.startswith("First Cost"):
        return "退货-货值"
    if "Return-Related Customer Service Cost" in text:
        return "退货-客服费"
    if "Supplier to Customer Shipping Cost" in text:
        return "退货-供应商到客户运费"
    return "退货-其他"


def adjustment_category(desc: object) -> str:
    text = str(desc or "")
    if text.startswith("Marketing Allowance"):
        return "调整-营销扣点"
    if "Sponsored Product Ads" in text:
        return "调整-BBB广告"
    if text.startswith("Audit Fee"):
        return "调整-审计费"
    if ("Expected" in text and "Billed" in text) or "Overbill Charge-Back" in text:
        return "调整-运费纠正"
    return "调整-其他"


def line_category(row: pd.Series) -> str:
    line_type = str(row.get("Line Type") or "")
    if line_type == "Sales":
        return "销售"
    if line_type == "Returns":
        return return_category(row.get("Description"))
    if line_type == "Adjustments":
        return adjustment_category(row.get("Description"))
    if line_type == "Supplier Oasis Fees":
        return "费用-Supplier Oasis"
    return "其他"


def parse_payment_summary(path: Path) -> tuple[str, str, list[tuple[str, float | None, float | None]]]:
    df = pd.read_excel(path, sheet_name="Payment Summary", header=None, dtype=str)
    section = ""
    rows: list[tuple[str, float | None, float | None]] = []
    check_no = ""
    check_date = ""
    for _, row in df.iterrows():
        values = [str(x).strip() for x in row.tolist() if pd.notna(x) and str(x).strip()]
        if not values:
            continue
        text = values[0]
        if text == "Sales":
            section = "sales"
            continue
        if text == "Returns":
            section = "returns"
            continue
        if text == "Adjustments":
            section = "adjustments"
            continue
        if text == "Supplier Oasis Fees":
            section = "fees"
            continue
        if text.startswith("Check Number"):
            check_no = text.split(":")[-1].strip()
        if text.startswith("Check Date"):
            check_date = text.split(":")[-1].strip()
        if text == "Total Sales" and section == "sales":
            rows.append(("销售合计", num(values[1]), num(values[2])))
        elif section == "returns":
            if text == "First Cost":
                rows.append(("退货-货值", num(values[1]), num(values[2])))
            elif text == "Others":
                rows.append(("退货-其他", num(values[1]), num(values[2])))
            elif text == "Return-Related Customer Service Cost":
                rows.append(("退货-客服费", num(values[1]), num(values[2])))
            elif text == "Supplier to Customer Shipping Cost":
                rows.append(("退货-供应商到客户运费", num(values[1]), num(values[2])))
            elif text == "Total":
                rows.append(("退货-合计", num(values[1]), num(values[2])))
        elif section == "adjustments":
            if text == "Marketing Allowance 8.25%":
                rows.append(("调整-营销扣点", None, num(values[1])))
            elif text == "Total Adjustments":
                rows.append(("调整-其他合计", None, num(values[1])))
            elif text == "Total":
                rows.append(("调整-合计", None, num(values[1])))
        elif section == "fees":
            if text == "Supplier Oasis Transaction Fees":
                rows.append(("费用-Supplier Oasis", None, num(values[1])))
            elif text == "Total":
                rows.append(("费用-合计", None, num(values[1])))
        if text == "Check Total":
            rows.append(("Check Total", None, num(values[1])))
    return check_no, check_date, rows


def load_account(path: Path) -> pd.DataFrame:
    detail = pd.read_excel(path, sheet_name="Detail", dtype=str, header=0)
    for col in ["OS Order #", "SOFS Order #", "OS SKU", "Supplier SKU"]:
        if col in detail.columns:
            detail[col] = detail[col].astype(str).str.strip()
    detail["base_os"] = detail["OS Order #"].map(base_id)
    detail["Quantity"] = pd.to_numeric(detail["Quantity"], errors="coerce")
    detail["Unit Price"] = pd.to_numeric(detail["Unit Price"], errors="coerce")
    detail["Total"] = pd.to_numeric(detail["Total"], errors="coerce")
    detail["费用分类"] = detail.apply(line_category, axis=1)
    return detail


def load_en_credentials() -> tuple[str, str]:
    values: dict[str, str] = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                key, value = line.split("=", 1)
                values[key.strip()] = value.strip().strip('"').strip("'")
    key = values.get("PROD_ERP_API_KEY") or values.get("ERP_API_KEY") or ""
    secret = values.get("PROD_ERP_API_SECRET") or values.get("ERP_API_SECRET") or ""
    return key, secret


def en_headers(key: str, secret: str) -> dict[str, str]:
    return {"Authorization": f"token {key}:{secret}"}


def discover_en_names(base_ids: set[str], headers: dict[str, str]) -> set[str]:
    prefixes = ["OS-", "OSFD-", "OSC-", "OSF-"]
    suffixes = ["", "-1", "-2", "-3", "-4", "_1", "_2", "_3", "_4", "-01", "-02", "-03", "_01", "_02", "_03"]
    names: set[str] = set()

    def one(base: str) -> list[str]:
        variants = [prefix + base + suffix for prefix in prefixes for suffix in suffixes]
        params = {
            "filters": json.dumps([["Tongtool Order", "name", "in", variants]]),
            "fields": json.dumps(["name"]),
            "limit_page_length": "100",
        }
        try:
            resp = requests.get(
                f"{EN_BASE}/api/resource/Tongtool Order",
                headers=headers,
                params=params,
                timeout=60,
            )
            if resp.status_code == 200:
                return [doc["name"] for doc in resp.json().get("data", [])]
        except requests.RequestException:
            pass
        return []

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = [pool.submit(one, base) for base in sorted(base_ids)]
        for future in as_completed(futures):
            names.update(future.result())
    return names


def stripped_name(name: str) -> str:
    return re.sub(r"^(?:OS|OSFD|OSC|OSF)-", "", name or "")


def is_split_name(name: str) -> bool:
    return bool(re.search(r"[_-]\d+$", stripped_name(name)))


def mark_duplicate_masters(docs: list[dict]) -> None:
    groups: dict[str, list[dict]] = {}
    for doc in docs:
        groups.setdefault(doc.get("base_order", ""), []).append(doc)
    for base, group in groups.items():
        masters = [doc for doc in group if not is_split_name(doc.get("name", ""))]
        splits = [doc for doc in group if is_split_name(doc.get("name", ""))]
        if not masters or not splits:
            for doc in group:
                doc["duplicate_master"] = False
            continue
        master_amount = sum(num(doc.get("order_amount")) for doc in masters)
        split_amount = sum(num(doc.get("order_amount")) for doc in splits)
        duplicate = abs(master_amount - split_amount) <= 0.01
        for doc in masters:
            doc["duplicate_master"] = duplicate
        for doc in splits:
            doc["duplicate_master"] = False


def fetch_en_docs(names: set[str], headers: dict[str, str]) -> list[dict]:
    top_fields = [
        "name", "platform_order_id", "order_id_code", "platform_code", "sale_account",
        "order_status", "order_type", "sale_time", "despatch_complete_time", "warehouse_name",
        "order_amount", "order_amount_currency", "products_total_price", "actual_total_price",
        "platform_fee", "shipping_fee", "gross_profit", "profit_margin",
        "total_item_cost", "first_leg_cost", "match_status", "sync_status",
    ]
    item_fields = [
        "platform_sku", "tongtool_sku", "erp_item_code", "quantity", "transaction_price",
        "cal_price", "item_cost", "shipping_cost", "sx_shipping_cost", "last_leg_fee",
        "allocated_carrier_fee", "fg_billing_weight_kg",
    ]
    package_fields = ["package", "allocated_fee", "currency", "allocated_carrier_fee"]

    def fetch(name: str) -> dict:
        try:
            resp = requests.get(
                f"{EN_BASE}/api/resource/Tongtool Order/{name}",
                headers=headers,
                timeout=60,
            )
        except requests.RequestException as exc:
            return {"name": name, "status": "error", "error": type(exc).__name__}
        if resp.status_code != 200:
            return {"name": name, "status": "error", "error": f"http_{resp.status_code}"}
        doc = resp.json().get("data") or {}
        record = {field: doc.get(field) for field in top_fields}
        record["status"] = "ok"
        record["items"] = [
            {field: item.get(field) for field in item_fields}
            for item in doc.get("order_items") or []
        ]
        record["packages"] = [
            {field: pkg.get(field) for field in package_fields}
            for pkg in doc.get("packages") or []
        ]
        try:
            raw = json.loads(doc.get("raw_data") or "{}")
            platform_goods = (raw.get("goodsInfo") or {}).get("platformGoodsInfoList") or []
            record["raw_platform_items"] = [
                {
                    "sku": str(g.get("webstoreSku") or ""),
                    "quantity": g.get("quantity"),
                    "web_transaction_id": str(g.get("webTransactionId") or ""),
                }
                for g in platform_goods
            ]
        except Exception:
            record["raw_platform_items"] = []
        record["base_order"] = base_id(record.get("platform_order_id") or record.get("name"))
        return record

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = [pool.submit(fetch, name) for name in sorted(names)]
        docs = [future.result() for future in as_completed(futures)]
    mark_duplicate_masters(docs)
    return docs


def en_summary(docs: list[dict]) -> dict[str, float | int]:
    return {
        "docs": len(docs),
        "order_amount": sum(num(doc.get("order_amount")) for doc in docs),
        "products_total": sum(num(doc.get("products_total_price")) for doc in docs),
        "actual_total": sum(num(doc.get("actual_total_price")) for doc in docs),
        "platform_fee": sum(num(doc.get("platform_fee")) for doc in docs),
        "shipping_fee": sum(num(doc.get("shipping_fee")) for doc in docs),
        "gross_profit": sum(num(doc.get("gross_profit")) for doc in docs),
        "item_cost": sum(num(doc.get("total_item_cost")) for doc in docs),
    }


def write_workbook(accounts: dict[str, pd.DataFrame], docs: list[dict], output: Path) -> None:
    effective = [doc for doc in docs if not doc.get("duplicate_master")]
    docs_by_base: dict[str, list[dict]] = {}
    for doc in effective:
        docs_by_base.setdefault(doc.get("base_order", ""), []).append(doc)

    overview_rows: list[dict] = []
    payment_rows: list[dict] = []
    detail_rows: list[dict] = []
    order_rows: list[dict] = []
    for key, detail in accounts.items():
        check_no, check_date, payment = parse_payment_summary(Path(key))
        sales = detail[detail["Line Type"] == "Sales"]
        returns = detail[detail["Line Type"] == "Returns"]
        adjustments = detail[detail["Line Type"] == "Adjustments"]
        fees = detail[detail["Line Type"] == "Supplier Oasis Fees"]
        sales_bases = set(sales["base_os"])
        return_bases = set(returns["base_os"])
        sales_docs = [doc for doc in effective if doc["base_order"] in sales_bases]
        return_docs = [doc for doc in effective if doc["base_order"] in return_bases - sales_bases]
        es = en_summary(sales_docs)
        sales_total = float(sales["Total"].sum())
        marketing_abs = abs(float(adjustments[adjustments["费用分类"] == "调整-营销扣点"]["Total"].sum()))
        overview_rows.append({
            "账期文件": key,
            "Check Number": check_no,
            "Check Date": check_date,
            "Detail行数": len(detail),
            "销售行数": len(sales),
            "销售金额": round(sales_total, 2),
            "退货金额": round(float(returns["Total"].sum()), 2),
            "调整金额": round(float(adjustments["Total"].sum()), 2),
            "费用金额": round(float(fees["Total"].sum()), 2),
            "EN有效销售单据数": es["docs"],
            "EN销售订单金额": round(es["order_amount"], 2),
            "销售金额差异": round(sales_total - es["order_amount"], 2),
            "EN平台费": round(es["platform_fee"], 2),
            "EN运费": round(es["shipping_fee"], 2),
            "EN实际回款": round(es["actual_total"], 2),
            "EN毛利": round(es["gross_profit"], 2),
            "EN总成本": round(es["item_cost"], 2),
            "EN退货原单金额": round(en_summary(return_docs)["order_amount"], 2),
            "账期营销扣点(绝对值)": round(marketing_abs, 2),
            "平台费vs营销扣点差异": round(float(es["platform_fee"]) - marketing_abs, 2),
        })
        for label, qty, amount in payment:
            payment_rows.append({"账期文件": key, "类别": label, "数量": qty, "金额": amount})
        payment_rows.append({"账期文件": key, "类别": "EN对照-销售订单金额", "数量": es["docs"], "金额": round(es["order_amount"], 2)})
        payment_rows.append({"账期文件": key, "类别": "EN对照-平台费", "数量": "", "金额": round(es["platform_fee"], 2)})
        payment_rows.append({"账期文件": key, "类别": "EN对照-退货原单金额", "数量": len(return_docs), "金额": round(en_summary(return_docs)["order_amount"], 2)})

        for _, row in detail.iterrows():
            detail_rows.append({
                "账期文件": key,
                "行类型": row["Line Type"],
                "费用分类": row["费用分类"],
                "描述": row.get("Description"),
                "OS订单号": row.get("OS Order #"),
                "基础OS订单号": row.get("base_os"),
                "SOFS订单号": row.get("SOFS Order #"),
                "供应商SKU": row.get("Supplier SKU"),
                "数量": row["Quantity"] if pd.notna(row["Quantity"]) else "",
                "金额": row["Total"] if pd.notna(row["Total"]) else "",
            })

        order_frame = detail[detail["base_os"].str.fullmatch(r"\d+", na=False)].copy()
        for base, group in order_frame.groupby("base_os"):
            group_sales = group[group["Line Type"] == "Sales"]
            group_returns = group[group["Line Type"] == "Returns"]
            docs_for_order = docs_by_base.get(base, [])
            es_order = en_summary(docs_for_order)
            sales_total = float(group_sales["Total"].sum()) if len(group_sales) else math.nan
            amount_diff = sales_total - float(es_order["order_amount"]) if len(group_sales) else None
            if not docs_for_order:
                status = "未匹配"
            elif len(group_sales) == 0:
                status = "仅退货"
            elif amount_diff is not None and abs(amount_diff) <= 0.01:
                status = "金额一致"
            else:
                status = "金额差异"
            order_rows.append({
                "账期文件": key,
                "基础OS订单号": base,
                "匹配关系": "1:1" if len(docs_for_order) == 1 else f"1:{len(docs_for_order)}拆单",
                "EN有效单据数": len(docs_for_order),
                "账期销售行数": len(group_sales),
                "账期销售金额": round(sales_total, 2) if not math.isnan(sales_total) else "",
                "账期退货金额": round(float(group_returns["Total"].sum()), 2) if len(group_returns) else "",
                "EN订单金额": round(es_order["order_amount"], 2),
                "EN平台费": round(es_order["platform_fee"], 2),
                "EN实际回款": round(es_order["actual_total"], 2),
                "EN毛利": round(es_order["gross_profit"], 2),
                "金额差异": round(amount_diff, 2) if amount_diff is not None else "",
                "核对状态": status,
            })

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(overview_rows).to_excel(writer, sheet_name="核对总览", index=False)
        pd.DataFrame(payment_rows).to_excel(writer, sheet_name="PaymentSummary逐项", index=False)
        pd.DataFrame(detail_rows).to_excel(writer, sheet_name="账期Detail", index=False)
        pd.DataFrame(order_rows).to_excel(writer, sheet_name="订单级费用核对", index=False)
        if docs:
            en_rows = []
            for doc in docs:
                en_rows.append({
                    "Tongtool订单号": doc.get("name"),
                    "基础OS订单号": doc.get("base_order"),
                    "platform_order_id": doc.get("platform_order_id"),
                    "sale_account": doc.get("sale_account"),
                    "是否重复主单": "是" if doc.get("duplicate_master") else "否",
                    "订单状态": doc.get("order_status"),
                    "仓库": doc.get("warehouse_name"),
                    "sale_time": doc.get("sale_time"),
                    "发货完成时间": doc.get("despatch_complete_time"),
                    "order_amount": doc.get("order_amount"),
                    "actual_total_price": doc.get("actual_total_price"),
                    "platform_fee": doc.get("platform_fee"),
                    "shipping_fee": doc.get("shipping_fee"),
                    "gross_profit": doc.get("gross_profit"),
                    "total_item_cost": doc.get("total_item_cost"),
                    "P号": uniq_join(pkg.get("package") for pkg in doc.get("packages") or []),
                })
            pd.DataFrame(en_rows).to_excel(writer, sheet_name="EN订单财务明细", index=False)

    workbook = load_workbook(output)
    fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = bold
            cell.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for idx, column in enumerate(sheet.columns, 1):
            width = min(max([len(str(cell.value or "")) for cell in column[:100]]) + 3, 46)
            sheet.column_dimensions[get_column_letter(idx)].width = max(width, 11)
    workbook.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description="OSTKUS 账期费用级对账")
    parser.add_argument("--account", action="append", required=True, help="OSTKUS 账期 xlsx 路径，可多次传入")
    parser.add_argument("--out", default="", help="输出 xlsx 路径")
    parser.add_argument("--no-en", action="store_true", help="不调用 EN API，只输出账期侧明细")
    args = parser.parse_args()

    accounts: dict[str, pd.DataFrame] = {}
    for raw_path in args.account:
        path = Path(raw_path)
        accounts[str(path)] = load_account(path)

    docs: list[dict] = []
    if not args.no_en:
        key, secret = load_en_credentials()
        if not key or not secret:
            print("EN 凭证缺失，请检查 EN_API/.env；本次按 --no-en 处理。", file=sys.stderr)
        else:
            headers = en_headers(key, secret)
            base_ids = {row["base_os"] for frame in accounts.values() for _, row in frame.iterrows() if row["base_os"]}
            names = discover_en_names(base_ids, headers)
            docs = fetch_en_docs(names, headers)
            print(f"EN: {len(base_ids)} 个基础订单号，发现 {len(names)} 条 Tongtool Order。")

    output = Path(args.out) if args.out else ROOT / "platform_account_reconciliation" / "out" / "OSTKUS费用级核对.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    write_workbook(accounts, docs, output)
    print(f"OK: {output}")


if __name__ == "__main__":
    main()
