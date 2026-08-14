from pathlib import Path

from openpyxl import load_workbook

from amazon_pairing.candidates_v2 import CandidateScore
from amazon_pairing.v2 import V2Decision
from amazon_pairing.v2_report import write_v2_workbook


def candidate(sku, score=90, conflicts=0, strong=False):
    return CandidateScore(
        sku=sku,
        name=f"name-{sku}",
        object_type="finished_product",
        score=score,
        evidence=("asin",),
        hard_conflicts=conflicts,
        is_strong_conflict=strong,
    )


def test_v2_workbook_has_expected_sheets_and_rows(tmp_path: Path):
    output = tmp_path / "v2.xlsx"
    records = [
        (
            {"sku": "A", "asin": "A1", "title": "A title", "mainImage": ""},
            V2Decision("", "A", "A1", "strong_single", "finished_product", (), (candidate("KS-A"),), ("asin",)),
        ),
        (
            {"sku": "B", "asin": "B1", "title": "B title", "mainImage": ""},
            V2Decision("", "B", "B1", "candidate", "cover", ("no_filler",), (candidate("KS-COVER", 45),), ("parent_sku",)),
        ),
        (
            {"sku": "C", "asin": "C1", "title": "2 pcs", "mainImage": ""},
            V2Decision("", "C", "C1", "special_with_candidate", "combo", ("set_count",), (candidate("KS-SET", 20, 1),), ("parent_sku",)),
        ),
    ]
    summary = {"输入": 3, "强证据建议": 1, "候选审核": 1, "对象专项": 1, "无候选": 0}

    write_v2_workbook(output, records, summary)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "运行汇总",
        "强证据建议",
        "Top候选审核",
        "低证据候选",
        "冲突候选审核",
        "对象专项",
        "无候选",
        "证据审计",
        "反馈说明",
    ]
    assert workbook["强证据建议"].max_row == 2
    assert workbook["Top候选审核"].max_row == 2
    assert workbook["对象专项"].max_row == 2
    assert workbook["强证据建议"].data_validations.count == 1
