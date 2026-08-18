from pathlib import Path

import openpyxl

from amazon_pairing.attributes import AttributeValue, ListingAttributes
from amazon_pairing.candidates import CandidateProduct, ListingQuery
from amazon_pairing.ranking import (
    PairExample,
    RankingModel,
    evaluate_rankings,
    grouped_split,
)
from amazon_pairing.report import ReviewRecord, write_review_workbook


EMPTY_ATTRS = ListingAttributes(
    AttributeValue(), AttributeValue(), AttributeValue(), AttributeValue()
)


def product(sku: str) -> CandidateProduct:
    return CandidateProduct(sku, "KS0001", sku, EMPTY_ATTRS)


def query(msku: str, asin: str = "") -> ListingQuery:
    return ListingQuery(msku, msku, ("KS0001",), EMPTY_ATTRS)


def test_grouped_split_keeps_same_msku_and_asin_cluster_together():
    examples = [
        PairExample("q1", "MSKU-A", "ASIN-X", product("SKU-1"), 1, {}),
        PairExample("q2", "MSKU-B", "ASIN-X", product("SKU-2"), 1, {}),
        PairExample("q3", "MSKU-C", "ASIN-Y", product("SKU-3"), 1, {}),
        PairExample("q4", "MSKU-D", "ASIN-Z", product("SKU-4"), 1, {}),
    ]

    train, validation = grouped_split(examples, validation_fraction=0.5, seed=7)
    sides = {example.query_id: "train" for example in train}
    sides.update({example.query_id: "validation" for example in validation})

    assert sides["q1"] == sides["q2"]
    assert train and validation


def test_ranking_metrics_report_recall_and_mrr_by_query():
    rankings = {
        "q1": [("A", 0.9, 1), ("B", 0.8, 0)],
        "q2": [("C", 0.9, 0), ("D", 0.8, 1)],
        "q3": [("E", 0.9, 0), ("F", 0.8, 0)],
    }

    metrics = evaluate_rankings(rankings, at=(1, 2))

    assert metrics["queries"] == 3.0
    assert metrics["recall_at_1"] == 1 / 3
    assert metrics["recall_at_2"] == 2 / 3
    assert metrics["mrr"] == 0.5


def test_lightgbm_ranker_trains_predicts_and_reloads(tmp_path: Path):
    examples = []
    for query_index in range(8):
        query_id = f"q{query_index}"
        examples.extend(
            [
                PairExample(
                    query_id,
                    f"MSKU-{query_index}",
                    f"ASIN-{query_index}",
                    product(f"RIGHT-{query_index}"),
                    1,
                    {"exact_target": 1.0, "reliable_conflicts": 0.0},
                ),
                PairExample(
                    query_id,
                    f"MSKU-{query_index}",
                    f"ASIN-{query_index}",
                    product(f"WRONG-{query_index}"),
                    0,
                    {"exact_target": 0.0, "reliable_conflicts": 1.0},
                ),
            ]
        )

    model = RankingModel(seed=11).fit(examples)
    scores = model.predict(examples[:2])
    assert scores[0] > scores[1]

    path = tmp_path / "中文目录" / "ranker.txt"
    model.save(path)
    restored = RankingModel.load(path)
    restored_scores = restored.predict(examples[:2])
    assert restored_scores[0] > restored_scores[1]


def test_review_workbook_contains_fixed_feedback_contract(tmp_path: Path):
    output = tmp_path / "review.xlsx"
    records = [
        ReviewRecord(
            shop="US Store",
            marketplace="US",
            msku="MSKU-1",
            asin="B0001",
            title="Grey triangle pillow",
            image_url="https://example.invalid/image.jpg",
            route="ordinary",
            candidates=(("KS0001-A", "候选A", 0.91, "size/color match"),),
        )
    ]

    write_review_workbook(
        output,
        records,
        summary={"输入Listing": 1, "普通商品": 1},
        special_records=[],
        no_candidate_records=[],
        quarantined_records=[],
        metrics={"recall_at_3": 0.8},
    )

    workbook = openpyxl.load_workbook(output)
    assert workbook.sheetnames == [
        "运行汇总",
        "高可信精确证据",
        "智能候选审核",
        "特殊对象暂缓",
        "无可靠候选",
        "隔离历史数据",
        "模型评估",
        "反馈说明",
    ]
    sheet = workbook["智能候选审核"]
    headers = [cell.value for cell in sheet[1]]
    assert "人工结论" in headers
    assert "正确SKU另填" in headers
    feedback_cell = sheet.cell(row=2, column=headers.index("人工结论") + 1)
    assert feedback_cell.data_type == "n"
    assert sheet.data_validations.count == 1
