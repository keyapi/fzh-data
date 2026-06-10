# -*- coding: utf-8 -*-
"""Excel export — multi-SKU listing flat files from AssetCollection."""

from __future__ import annotations

import io
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from models import AssetCollection, Asset

_COLUMN_MAP = {
    "amazon": [
        "main_image_url", "other_image_url1", "other_image_url2",
        "other_image_url3", "other_image_url4", "other_image_url5",
        "other_image_url6", "other_image_url7", "other_image_url8",
    ],
    "wayfair": [
        "main_image_url", "image_url_2", "image_url_3",
        "image_url_4", "image_url_5", "image_url_6",
    ],
    "shopify": [
        "image_src", "image_src_2", "image_src_3", "image_src_4",
    ],
    "home24": [
        "media_main_image", "media_detail_1", "media_detail_2",
        "media_detail_3", "media_detail_4",
    ],
}


def _get_asset_url(asset: Asset | None) -> str:
    if not asset:
        return ""
    return f"/files/{Path(asset.stored_path).name}"


def export_collection_to_excel(
    session: Session, collection_id: str, platform: str = "amazon"
) -> bytes:
    """Multi-SKU export: one row per SKU, images ordered by position."""
    coll = session.query(AssetCollection).filter_by(id=collection_id).first()
    if not coll:
        raise ValueError("Collection not found")

    items = sorted(coll.items, key=lambda x: x.position)
    columns = _COLUMN_MAP.get(platform, _COLUMN_MAP["amazon"])

    # Group items by SKU, maintaining position order per SKU
    sku_images: dict[str, list[tuple[int, str]]] = defaultdict(list)
    sku_position: dict[str, int] = {}
    for idx, item in enumerate(items):
        asset = session.query(Asset).filter_by(id=item.asset_id).first()
        if not asset:
            continue
        # Get SKU from asset's primary product link
        sku = None
        if asset.product_links:
            primary = next(
                (pl for pl in asset.product_links if pl.is_primary), None
            )
            sku = (primary or asset.product_links[0]).product_sku
        if not sku:
            sku = "_unlinked"
        url = _get_asset_url(asset)
        if sku not in sku_position:
            sku_position[sku] = len(sku_position)
        sku_images[sku].append((idx, url))

    wb = Workbook()
    ws = wb.active
    ws.title = "Listing"

    # Headers
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    ws.cell(row=1, column=1, value="SKU").font = Font(bold=True)
    ws.cell(row=1, column=1).fill = header_fill
    for i, col_name in enumerate(columns, 2):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Data rows — sort SKUs by first appearance order
    sorted_skus = sorted(sku_images.keys(), key=lambda s: sku_position[s])
    for row_idx, sku in enumerate(sorted_skus, 2):
        ws.cell(row=row_idx, column=1, value=sku)
        for img_idx, (_, url) in enumerate(sorted(sku_images[sku], key=lambda x: x[0])):
            if img_idx < len(columns):
                ws.cell(row=row_idx, column=img_idx + 2, value=url)

    # Column widths
    ws.column_dimensions["A"].width = 20
    for i in range(len(columns)):
        col_letter = chr(ord("B") + i)
        ws.column_dimensions[col_letter].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
