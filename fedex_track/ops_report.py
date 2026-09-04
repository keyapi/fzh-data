"""fedex_track 运营异常报表生成器（v2：统一中文/EN、单一配色、表与分类对应）。

读 fedex 全量 summary(v2) + 通途原始表，输出一个**多工作表 + 配色 + 汇总**的 Excel，
站在 Amazon 运营角度做异常判定与动作建议。

口径：起点=建标时间；确认发货=站点收件时间(FedEx 首次取件扫描)；延迟用**营业日**计
（排除周末 + 美国联邦节假日），贴近 Amazon LSR（ship-by=下单日+处理时间，只算工作日）。
"""

from __future__ import annotations

import datetime as _dt
import re

import numpy as np
import pandas as pd

# ── 阈值 / 口径（可调）─────────────────────────────────────────
HANDLING_DAYS = 1
TRANSIT_SLOW_DAYS = 6
TRANSIT_SEVERE_DAYS = 12
STUCK_DAYS = 7
STUCK_SEVERE_DAYS = 14
MISSING_AFTER_DAYS = 3
US_HOLIDAYS = ["2026-01-01","2026-01-19","2026-02-16","2026-05-25","2026-06-19",
               "2026-07-03","2026-09-07","2026-10-12","2026-11-11","2026-11-26","2026-12-25"]

# ── 单一归类与配色来源 ─────────────────────────────────────────
# key → (中文, 英文, 基础色)
CLASS = {
    "missing_not_handed": ("漏发/未交接", "not_handed", "red"),
    "fresh_no_pickup":    ("建标未收件", "label_no_pickup", "yellow"),
    "late_handover":      ("迟发", "late_handover", "orange"),
    "fedex_slow":         ("FedEx延误", "fedex_slow", "orange"),
    "stuck":              ("卡件", "stuck", "red"),
    "cancelled":          ("已取消", "cancelled", "gray"),
    "not_found":          ("数据异常/查无", "not_found", "gray"),
    "reused_no_label":    ("复用旧票(缺建标)", "reused_no_label", "blue"),
    "delivered_ok":       ("正常交付", "delivered_ok", "green"),
    "in_transit":         ("在途", "in_transit", "yellow"),
}
# 等级 → 颜色（覆盖 分类基础色；用于首列与分级上色）
LEVEL_COLOR = {
    "漏发/未交接": "red", "卡件": "red", "严重延误": "red", "重度迟发": "red",
    "中度迟发": "orange", "FedEx延误": "orange",
    "轻度迟发": "yellow", "建标未收件": "yellow", "在途": "yellow",
    "已取消": "gray", "数据异常/查无": "gray", "复用旧票(缺建标)": "blue",
    "正常交付": "green", "准时": "green",
}
COLOR = {"green": "C6EFCE", "yellow": "FFEB9C", "orange": "FCD5B4", "red": "FFC7CE",
         "gray": "D9D9D9", "blue": "DDEBF7", "dark": "404040", "header": "2F5597"}
SHEET_OF_CLASS = {
    "missing_not_handed": "漏发未交接", "fresh_no_pickup": "漏发未交接",
    "late_handover": "迟发", "fedex_slow": "承运异常", "stuck": "承运异常",
    "cancelled": "取消·其他", "not_found": "取消·其他",
}
ACTION = {
    "missing_not_handed": "通知仓库/货代核查是否漏发或未交接；未发出→取消或补发，已交但无扫描→开FedEx trace",
    "fresh_no_pickup": "持续关注是否转漏发；与仓确认交接；超过N天按漏发处理",
    "late_handover": "交接SLA复盘；对已延误买家提前告知/安抚；根因治理（仓、货代排程）",
    "fedex_slow": "记录；超过严重阈值→开FedEx trace/索赔",
    "stuck": "开FedEx trace调查；超14天→索赔+重发/退款",
    "cancelled": "确认取消原因（我方撤单/货代/FedEx），按需退款或重下",
    "not_found": "核查通途源数据（如拼接号/非FedEx号）；非FedEx单改走对应渠道",
    "reused_no_label": "复用号旧票，从建标开始的对应票为准；无需处理",
    "delivered_ok": "—", "in_transit": "—",
}


def bizdays(d1, d2):
    if pd.isna(d1) or pd.isna(d2):
        return None
    try:
        return int(np.busday_count(np.datetime64(d1.date()), np.datetime64(d2.date()), holidays=US_HOLIDAYS))
    except Exception:
        return None


def _to_ts(v):
    try:
        d = pd.to_datetime(v, errors="coerce")
        return d
    except Exception:
        return pd.NaT


# 通途标识列
TT_PICK = {
    "订单号": "订单号", "包裹号": "包裹号", "渠道账号": "渠道账号", "渠道": "渠道",
    "平台SKU": "平台SKU", "通途SKU": "通途SKU", "产品名称": "产品名称", "品类": "品类",
    "发货仓库": "发货仓库", "执行发货人": "执行发货人", "是否补发货": "是否补发货",
    "销售站点": "销售站点", "买家姓名": "买家姓名", "国家/地区": "国家/地区", "城市": "城市",
    "省/州": "省/州", "邮编": "邮编", "订单总售价": "订单总售价", "利润": "利润",
    "发货日期": "发货日期", "发货时间": "发货时间", "邮寄方式": "邮寄方式",
}


def load_tt_identity(xlsx: str) -> dict[str, dict]:
    df = pd.read_excel(xlsx, sheet_name=0)
    tc = next(c for c in df.columns if "跟踪号" in str(c))
    num = df[tc].fillna("").astype(str).str.upper().str.replace(r"[-\s_]", "", regex=True)
    out: dict[str, dict] = {}
    for val, row in zip(num, df.to_dict("records")):
        if not val or val in out:
            continue
        ident = {}
        for k, src in TT_PICK.items():
            v = row.get(src)
            ident[k] = "" if pd.isna(v) else str(v).strip()
        out[val] = ident
    return out


def _cat(dev, pu, label, ship, now):
    """返回分类 key。"""
    if pu is pd.NaT:
        if dev is not pd.NaT:
            return "reused_no_label" if label is pd.NaT else "delivered_ok"
        if ship is not pd.NaT and (now - ship).days > MISSING_AFTER_DAYS:
            return "missing_not_handed"
        return "fresh_no_pickup"
    if label is pd.NaT:
        return "reused_no_label"
    if dev is pd.NaT:
        return "stuck" if (now - label).days > STUCK_DAYS else "in_transit"
    if (dev - pu).total_seconds() / 86400 > 7:
        return "fedex_slow"
    if ship is not pd.NaT and (pu.date() - ship.date()).days > 2:
        return "late_handover"
    return "delivered_ok"


def build(summary_csv: str, tt_xlsx: str, out_xlsx: str):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    S = pd.read_csv(summary_csv, dtype=str).fillna("")
    tt = load_tt_identity(tt_xlsx)
    now = pd.Timestamp(_dt.datetime.now())

    COLS = ["跟踪号", "分类(EN)", "分类(中文)", "等级", "订单号", "包裹号", "渠道账号", "渠道",
            "平台SKU", "通途SKU", "产品名称", "发货仓库", "执行发货人", "是否补发货", "销售站点",
            "买家姓名", "国家", "城市", "省/州", "发货日期", "建标时间", "站点收件时间", "交付时间",
            "日历延迟(天)", "营业日延迟(天)", "承运延迟(营业日)", "Amazon是否判迟", "所属Sheet", "建议动作", "处理状态"]

    rows = []
    for _, r in S.iterrows():
        n = r["跟踪号"]; dev = _to_ts(r["交付时间"]); pu = _to_ts(r["站点收件时间"]); label = _to_ts(r["建标时间"])
        ship = pd.NaT
        m = re.search(r"发货日期=([0-9-]+)", r["备注"])
        if m:
            ship = pd.to_datetime(m.group(1), errors="coerce")
        cat_key = _cat(dev, pu, label, ship, now)
        if r["已取消"] == "是" and dev is pd.NaT:
            cat_key = "cancelled"
        if r["当前状态"] == "查无此号":
            cat_key = "not_found"
        zh, en, base = CLASS[cat_key]
        overdue = None
        if cat_key == "late_handover":
            bd = bizdays(label, pu)
            overdue = bd - HANDLING_DAYS if bd is not None else None
        trans = bizdays(pu, dev) if (cat_key == "fedex_slow" and dev is not pd.NaT and pu is not pd.NaT) else None
        # 等级
        if cat_key == "late_handover":
            level = "准时" if (overdue is not None and overdue <= 0) else ("轻度迟发" if overdue <= 2 else ("中度迟发" if overdue <= 5 else "重度迟发"))
        elif cat_key == "fedex_slow":
            level = "严重延误" if (trans and trans > TRANSIT_SEVERE_DAYS) else "FedEx延误"
        else:
            level = zh
        amazon = "" if overdue is None else ("是" if overdue > 0 else "否")
        ident = tt.get(n, {})
        caldays = int((pu - label).days) if (pu is not pd.NaT and label is not pd.NaT) else ""
        rows.append({
            "跟踪号": n, "分类(EN)": en, "分类(中文)": zh, "等级": level,
            "订单号": ident.get("订单号", ""), "包裹号": ident.get("包裹号", ""),
            "渠道账号": ident.get("渠道账号", ""), "渠道": ident.get("渠道", ""),
            "平台SKU": ident.get("平台SKU", ""), "通途SKU": ident.get("通途SKU", ""),
            "产品名称": ident.get("产品名称", "")[:40], "发货仓库": ident.get("发货仓库", ""),
            "执行发货人": ident.get("执行发货人", ""), "是否补发货": ident.get("是否补发货", ""),
            "销售站点": ident.get("销售站点", ""), "买家姓名": ident.get("买家姓名", ""),
            "国家": ident.get("国家/地区", ""), "城市": ident.get("城市", ""), "省/州": ident.get("省/州", ""),
            "发货日期": ship.date() if ship is not pd.NaT else "",
            "建标时间": label, "站点收件时间": pu, "交付时间": dev,
            "日历延迟(天)": caldays, "营业日延迟(天)": overdue, "承运延迟(营业日)": trans,
            "Amazon是否判迟": amazon, "所属Sheet": SHEET_OF_CLASS.get(cat_key, ""),
            "建议动作": ACTION[cat_key], "处理状态": "",
        })
    df = pd.DataFrame(rows, columns=COLS)

    # 异常子集（排除 正常/在途/复用旧票）
    anomaly_cats = ["missing_not_handed", "fresh_no_pickup", "late_handover", "fedex_slow", "stuck", "cancelled", "not_found"]
    df["_key"] = df["分类(EN)"].map({v[1]: k for k, v in CLASS.items()})
    def pick(*keys):
        return df[df["_key"].isin(keys)]

    wb = openpyxl.Workbook()
    thin = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    thick = Side(style="thin", color="808080")
    thin_b = Border(*[thick] * 4)

    # ── 总览 ──
    ws = wb.active; ws.title = "总览"
    ws.cell(row=1, column=1, value="FedEx 运营异常总览（Amazon 口径 · 营业日）").font = Font(size=16, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=Color(COLOR_DARK) if False else "404040")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 24
    # KPI
    n_total = len(df)
    n_anom = len(pick(*anomaly_cats))
    kpi = [("总单(票)", n_total), ("异常(票)", n_anom), ("漏发/未交接", int(sum(df["_key"] == "missing_not_handed"))),
           ("迟发", int(sum(df["_key"] == "late_handover"))), ("FedEx延误", int(sum(df["_key"] == "fedex_slow"))),
           ("卡件", int(sum(df["_key"] == "stuck"))), ("已取消", int(sum(df["_key"] == "cancelled"))),
           ("数据异常", int(sum(df["_key"] == "not_found")))]
    for i, (k, v) in enumerate(kpi):
        col = 1 + i * 2
        ws.cell(row=3, column=col, value=k).font = Font(bold=True)
        c = ws.cell(row=4, column=col, value=v); c.font = Font(size=14, bold=True); c.alignment = Alignment(horizontal="center")
        if k in ("漏发/未交接", "卡件"):
            c.fill = PatternFill("solid", fgColor="FFC7CE")
    # 分类表(中文，与 Sheet 对应)
    r = 6
    ws.cell(row=r, column=1, value="分类(中文)").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=r, column=2, value="数量").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=r, column=3, value="所属Sheet").font = Font(bold=True, color="FFFFFF")
    for c in (1, 2, 3):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="2F5597")
    order = ["missing_not_handed", "fresh_no_pickup", "late_handover", "fedex_slow", "stuck",
             "cancelled", "not_found", "reused_no_label", "in_transit", "delivered_ok"]
    for k in order:
        cnt = int(sum(df["_key"] == k))
        r += 1
        zh, en, col = CLASS[k]
        ws.cell(row=r, column=1, value=zh)
        ws.cell(row=r, column=2, value=cnt).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=SHEET_OF_CLASS.get(k, "全部明细"))
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=COLOR[col])
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLOR[col])
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=COLOR[col])
    # 图例
    r += 2
    ws.cell(row=r, column=1, value="配色图例").font = Font(bold=True)
    legend = [("red", "需立即处理", "漏发/卡件/严重延误/重度迟发"), ("orange", "需关注", "中度迟发/FedEx延误"),
              ("yellow", "留意", "轻度迟发/建标未收件/在途"), ("blue", "信息", "复用号旧票"),
              ("gray", "已取消/数据异常", "取消 或 查无"), ("green", "正常/准时", "正常交付")]
    r += 1
    for col, lab, note in legend:
        ws.cell(row=r, column=1, value=lab).fill = PatternFill("solid", fgColor=COLOR[col])
        ws.cell(row=r, column=2, value=note).alignment = Alignment(wrap_text=True)
        r += 1
    for c in range(1, 4):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # ── 明细工作表 ──
    detail_plan = [
        ("异常处理", pick(*anomaly_cats), True, "异常子集：漏发/迟发/FedEx延误/卡件/取消/数据异常，含建议动作"),
        ("漏发未交接", pick("missing_not_handed", "fresh_no_pickup"), True, "建议：通知仓库/货代核查漏发或未交接"),
        ("迟发", pick("late_handover"), True, "含 营业日延迟 与 Amazon是否判迟；请按实际 handling 校准"),
        ("承运异常", pick("fedex_slow", "stuck"), True, "FedEx延误→记录/严重开trace；卡件→开trace/索赔"),
        ("取消·其他", pick("cancelled", "not_found"), True, "已取消 / 数据异常·查无"),
        ("全部明细", df, False, "全量（含正常/在途/复用旧票），审计与追溯"),
    ]
    widths = [16, 12, 14, 12, 14, 12, 12, 10, 12, 12, 24, 12, 10, 8, 10, 12, 8, 10, 8, 12, 18, 18, 18, 10, 10, 10, 10, 12, 58, 10]
    for name, data, colorize, note in detail_plan:
        ws2 = wb.create_sheet(name)
        if note:
            ws2.cell(row=1, column=1, value=note).font = Font(italic=True, color="808080")
            ws2.freeze_panes = "A3"
            start = 3
        else:
            start = 1
        for c, h in enumerate(COLS, 1):
            cell = ws2.cell(row=start, column=c, value=h)
            cell.fill = PatternFill("solid", fgColor="2F5597")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, (_, rd) in enumerate(data.iterrows()):
            rr = start + 1 + i
            for c, h in enumerate(COLS, 1):
                cell = ws2.cell(row=rr, column=c, value=rd[h])
                cell.border = thin
            if colorize:
                col = LEVEL_COLOR.get(rd["等级"], "green")
                ws2.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=COLOR[col])
        for c, w in enumerate(widths, 1):
            ws2.column_dimensions[get_column_letter(c)].width = w

    # ── 口径说明 ──
    ws3 = wb.create_sheet("口径说明")
    notes = [
        ("FedEx 异常运营报表口径说明（v2）", ""),
        ("1. 起点与确认", "起点=建标时间(面单创建)；确认发货=站点收件时间(FedEx首次取件扫描)。"),
        ("2. 迟发(Amazon口径)", "ship-by=下单日+处理时间(营业日)；周末与美国联邦假日不计入。营业日延迟=建标→收件营业日数-处理时间(默认1天)。"),
        ("3. 处理时间", f"默认 {HANDLING_DAYS} 个营业日，请按实际改脚本顶部 HANDLING_DAYS。"),
        ("4. 营业日排除", "周六日 + 2026美国联邦假日：1/1,1/19,2/16,5/25,6/19,7/3,9/7,10/12,11/11,11/26,12/25。"),
        ("5. 判定定义", "漏发/未交接=有发货日期但一直无站点收件且>{}天；建标未收件=建标但近期无站点收件；迟发=建标→收件营业日>处理时间；FedEx延误=收件→交付营业日>{}；卡件=在途且>{}天无扫描；复用旧票=有收件却无建标(复用号第2票)；取消/查无=单列。".format(MISSING_AFTER_DAYS, TRANSIT_SLOW_DAYS, STUCK_DAYS)),
        ("6. 动作", "漏发/建标未收件→通知仓库核查；迟发→SLA复盘+安抚；FedEx延误→记录/严重开trace；卡件→开FedEx trace/索赔；取消→确认原因；查无→核查源数据。"),
        ("7. 配色", "绿=正常/准时；黄=留意(轻度迟发/建标未收件/在途)；橙=关注(中度迟发/FedEx延误)；红=需立即处理(漏发/卡件/严重/重度迟发)；蓝=信息(复用旧票)；灰=取消/数据异常。"),
        ("8. 多票", "FedEx 复用跟踪号，同号多票。各票独立判定，跟踪号 [n] 区分；无建标的一票标为'复用旧票(缺建标)'，以有建标那票为准。"),
        ("9. 工作表关系", "「异常处理」=「全部明细」中**异常子集**（漏发/迟发/延误/卡件/取消/数据异常），同列多一『建议动作/处理状态』，便于排待办；其余为按类细分。「全部明细」为全量审计。"),
        ("10. 数据来源", "fedex_full_20260904_v2.summary.csv + 通途非FBA订单202608.xlsx。"),
    ]
    for i, (k, v) in enumerate(notes, 1):
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=i, column=1).alignment = Alignment(vertical="top")
        ws3.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.cell(row=i, column=2).fill = PatternFill("solid", fgColor="DDEBF7")
        ws3.row_dimensions[i].height = 34
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 118

    wb.save(out_xlsx)
    return df


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--summary", default="fedex_track_output/fedex_full_20260904_v2.summary.csv")
    p.add_argument("--tt", default="D:\\Work\\王忠于\\成本核算\\通途非FBA订单202608 202609030947 无需填0售价 加预估尾程.xlsx")
    p.add_argument("--out", default="fedex_track_output/fedex_ops_report_20260904.xlsx")
    a = p.parse_args()
    df = build(a.summary, a.tt, a.out)
    print("total rows", len(df))
    print(df["分类(中文)"].value_counts().to_dict())
    print("written", a.out)
